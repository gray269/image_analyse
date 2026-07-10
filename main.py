from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime, timedelta
from collections import Counter
import calendar
import hashlib
import json
import math
import os
import queue
import re
import threading
import time
import sys
import subprocess
import tempfile
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser
from tkinter.scrolledtext import ScrolledText

try:
    from PIL import Image, ImageTk, ImageChops, ImageDraw, ImageFilter
    Image.MAX_IMAGE_PIXELS = None
except ImportError:
    Image = None
    ImageTk = None

try:
    import numpy as np
except ImportError:
    np = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    Workbook = None
    PatternFill = Font = Alignment = Border = Side = None


SECTION_RE = re.compile(r"^\[(.+?)\]\s*$")
PRODUCT_TS_RE = re.compile(r"^(?P<product>\d{4,})(?:_(?P<ts>\d{14}))?.*$")
PICTURE_RE = re.compile(r"(picture[_ -]?\d+)", re.IGNORECASE)

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}
SKIP_SECTIONS_IN_COMPARE = {"PARAMETERS"}

INTERVALS = {
    "2 jours": 2,
    "1 semaine": 7,
    "1 mois": 31,
    "3 mois": 93,
    "6 mois": 186,
    "Tout": None,
}

FUZZY_ALIGN_SECTIONS = {
    "DARK SPOT",
    "WHITE SPOT",
    "SURFACE DEFECT",
    "FUNCTIONAL DEFECT",
    "GROWING SPOT",
}

GEOMETRIC_SECTIONS_PREFERRED = {
    "DARK SPOT",
    "WHITE SPOT",
    "SURFACE DEFECT",
    "FUNCTIONAL DEFECT",
    "ECE",
    "ECS",
    "AE",
    "OPTICAL FEEDBACK",
    "HALO",
    "EP OFF",
    "TA",
    "TB",
    "TBDELAY",
    "GROWING SPOT",
    "INHO",
}

PI_VAL = math.pi

CACHE_VERSION = 6
CACHE_DIR = Path.home() / ".comparateur_defauts_cache"
INDEX_CACHE_PATH = CACHE_DIR / "index_cache_v6.json"
SETTINGS_PATH = CACHE_DIR / "settings_v6.json"
FOLLOW_TARGETS_PATH = CACHE_DIR / "follow_targets_v7.txt"
THUMB_CACHE_DIR = CACHE_DIR / "thumbs"

DEFAULT_SETTINGS = {
    "resolution_um_per_pixel": 5.4,
    "max_files_per_type": 1500,
    "preview_max_size": 1800,
    "show_zone_boundaries": True,
    "zones": [
        {"name": "Z1", "max_um": 5800.0},
        {"name": "Z2", "max_um": 14700.0},
        {"name": "Z3", "max_um": 17700.0},
        {"name": "Hors zone", "max_um": None},
    ],
    "size_buckets": [
        {"name": "<75", "min_um": 0.0, "max_um": 75.0},
        {"name": "a", "min_um": 75.0, "max_um": 150.0},
        {"name": "b", "min_um": 150.0, "max_um": 230.0},
        {"name": "c", "min_um": 230.0, "max_um": 300.0},
    ],
    "grade_columns": [
        {"label": "GradeChickenWire", "key": "CW"},
        {"label": "GradeInho", "key": "INHO"},
    ],
}
APP_SETTINGS = json.loads(json.dumps(DEFAULT_SETTINGS))


@dataclass
class TextRun:
    path: Path
    product: str
    timestamp: datetime | None
    sections: dict[str, list[str]]


@dataclass
class ImageRun:
    path: Path
    product: str
    timestamp: datetime | None
    picture_name: str


@dataclass
class DefectPoint:
    section: str
    radius: float
    angle: float
    size: float
    defect_type: str
    location: str
    raw: str


@dataclass
class PassageSummary:
    run: TextRun
    grade_cw: str
    grade_inho: str
    grades: dict[str, str]
    counts: dict[str, int]
    total_points: int
    over_300_count: int
    recurrent_count: int
    location_counter: Counter
    section_counter: Counter
    recurrent_points: list[DefectPoint] = field(default_factory=list)
    all_points: list[DefectPoint] = field(default_factory=list)


@dataclass
class CellLabel:
    text: str
    bg: str = "white"
    fg: str = "black"
    bold: bool = False
    anchor: str = "center"
    width: int = 10


def normalize_settings(settings: dict) -> dict:
    merged = json.loads(json.dumps(DEFAULT_SETTINGS))
    if isinstance(settings, dict):
        for key, value in settings.items():
            merged[key] = value

    try:
        merged["resolution_um_per_pixel"] = float(merged.get("resolution_um_per_pixel", 5.4))
    except Exception:
        merged["resolution_um_per_pixel"] = 5.4

    try:
        merged["max_files_per_type"] = int(merged.get("max_files_per_type", 1500))
    except Exception:
        merged["max_files_per_type"] = 1500
    if merged["max_files_per_type"] < 0:
        merged["max_files_per_type"] = 0

    try:
        merged["preview_max_size"] = int(merged.get("preview_max_size", 1800))
    except Exception:
        merged["preview_max_size"] = 1800
    if merged["preview_max_size"] < 200:
        merged["preview_max_size"] = 200

    zones = []
    for z in merged.get("zones", []):
        if not isinstance(z, dict):
            continue
        name = str(z.get("name", "")).strip()
        if not name:
            continue
        max_um = z.get("max_um")
        if max_um in ("", None, "None", "none", "Hors zone"):
            max_um = None
        else:
            try:
                max_um = float(max_um)
            except Exception:
                max_um = None
        zones.append({"name": name, "max_um": max_um})
    if not zones:
        zones = json.loads(json.dumps(DEFAULT_SETTINGS["zones"]))
    merged["zones"] = zones

    buckets = []
    for b in merged.get("size_buckets", []):
        if not isinstance(b, dict):
            continue
        name = str(b.get("name", "")).strip()
        if not name:
            continue
        try:
            min_um = float(b.get("min_um", 0.0))
            max_um = float(b.get("max_um", 0.0))
        except Exception:
            continue
        if max_um <= min_um:
            continue
        buckets.append({"name": name, "min_um": min_um, "max_um": max_um})
    if not buckets:
        buckets = json.loads(json.dumps(DEFAULT_SETTINGS["size_buckets"]))
    merged["size_buckets"] = buckets

    grades = []
    for g in merged.get("grade_columns", []):
        if not isinstance(g, dict):
            continue
        label = str(g.get("label", "")).strip()
        key = str(g.get("key", "")).strip()
        if label and key:
            grades.append({"label": label, "key": key})
    if not grades:
        grades = json.loads(json.dumps(DEFAULT_SETTINGS["grade_columns"]))
    merged["grade_columns"] = grades

    return merged


def load_app_settings() -> dict:
    try:
        if SETTINGS_PATH.exists():
            with SETTINGS_PATH.open("r", encoding="utf-8") as f:
                return normalize_settings(json.load(f))
    except Exception as exc:
        print(f"Paramètres illisibles : {exc}")
    return normalize_settings(DEFAULT_SETTINGS)


def save_app_settings(settings: dict):
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        with SETTINGS_PATH.open("w", encoding="utf-8") as f:
            json.dump(normalize_settings(settings), f, ensure_ascii=False, indent=2)
    except Exception as exc:
        print(f"Impossible d'écrire les paramètres : {exc}")


def apply_global_settings(settings: dict):
    global APP_SETTINGS
    APP_SETTINGS = normalize_settings(settings)


def get_zones() -> list[dict]:
    return APP_SETTINGS.get("zones", DEFAULT_SETTINGS["zones"])


def get_size_buckets() -> list[dict]:
    return APP_SETTINGS.get("size_buckets", DEFAULT_SETTINGS["size_buckets"])


def get_grade_columns() -> list[dict]:
    return APP_SETTINGS.get("grade_columns", DEFAULT_SETTINGS["grade_columns"])


def get_max_files_per_type() -> int:
    try:
        return int(APP_SETTINGS.get("max_files_per_type", 1500))
    except Exception:
        return 1500


def get_preview_max_size() -> int:
    try:
        return int(APP_SETTINGS.get("preview_max_size", 1800))
    except Exception:
        return 1800


def zones_to_text() -> str:
    lines = []
    for zone in get_zones():
        max_um = zone.get("max_um")
        if max_um is None:
            max_txt = ""
        else:
            max_txt = f"{float(max_um):g}"
        lines.append(f"{zone.get('name', '')};{max_txt}")
    return "\n".join(lines)


def size_buckets_to_text() -> str:
    lines = []
    for b in get_size_buckets():
        lines.append(f"{b.get('name', '')};{float(b.get('min_um', 0)):g};{float(b.get('max_um', 0)):g}")
    return "\n".join(lines)


def grade_columns_to_text() -> str:
    return "\n".join(f"{g.get('label', '')};{g.get('key', '')}" for g in get_grade_columns())


def parse_zones_text(text: str) -> list[dict]:
    zones = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(";")]
        name = parts[0] if parts else ""
        if not name:
            continue
        max_um = None
        if len(parts) >= 2 and parts[1]:
            max_um = float(parts[1].replace(",", "."))
        zones.append({"name": name, "max_um": max_um})
    if not zones:
        raise ValueError("Aucune zone valide.")
    return zones


def parse_size_buckets_text(text: str) -> list[dict]:
    buckets = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(";")]
        if len(parts) < 3:
            raise ValueError(f"Classe taille invalide : {line}")
        name = parts[0]
        min_um = float(parts[1].replace(",", "."))
        max_um = float(parts[2].replace(",", "."))
        if max_um <= min_um:
            raise ValueError(f"Max <= min pour la classe : {line}")
        buckets.append({"name": name, "min_um": min_um, "max_um": max_um})
    if not buckets:
        raise ValueError("Aucune classe de taille valide.")
    return buckets


def parse_grade_columns_text(text: str) -> list[dict]:
    grades = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split(";")]
        if len(parts) < 2:
            raise ValueError(f"Colonne grade invalide : {line}")
        grades.append({"label": parts[0], "key": parts[1]})
    if not grades:
        raise ValueError("Aucune colonne grade valide.")
    return grades


def compact_column_label(zone_name: str, size_name: str) -> str:
    if zone_name.lower().startswith("hors"):
        return f"Hors {size_name}"
    return f"{zone_name}{size_name}"


apply_global_settings(load_app_settings())


def read_text_file(path: Path) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return path.read_text(errors="ignore")


def parse_product_timestamp_from_token(token: str) -> tuple[str | None, datetime | None]:
    stem = Path(token).stem
    m = PRODUCT_TS_RE.match(stem)
    if not m:
        return None, None

    product = m.group("product")
    ts = None
    if m.group("ts"):
        try:
            ts = datetime.strptime(m.group("ts"), "%Y%m%d%H%M%S")
        except ValueError:
            ts = None
    return product, ts


def parse_horodatage(value: str) -> datetime | None:
    value = value.strip()
    for fmt in ("%d/%m/%y %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y%m%d%H%M%S"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    return None


def parse_txt_file(path: Path) -> TextRun:
    raw = read_text_file(path)
    sections: dict[str, list[str]] = {}
    current_section = None

    for raw_line in raw.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        m = SECTION_RE.match(line)
        if m:
            current_section = m.group(1).strip()
            # La section PARAMETERS est très longue et inutile pour le rapport / comparaison principale.
            # On ne la garde pas en mémoire pour accélérer les scans serveur.
            if current_section.upper() != "PARAMETERS":
                sections.setdefault(current_section, [])
            continue

        if current_section and current_section.upper() != "PARAMETERS":
            sections[current_section].append(line)

    product_from_name, timestamp_from_name = parse_product_timestamp_from_token(path.stem)
    product = product_from_name
    timestamp = timestamp_from_name

    for line in sections.get("CONDITIONS", []):
        if line.startswith("Numéro tube="):
            product = line.split("=", 1)[1].strip() or product
        elif line.startswith("Horodatage="):
            # Horodatage interne prioritaire : le nom de fichier n'est pas fiable.
            internal_ts = parse_horodatage(line.split("=", 1)[1])
            if internal_ts is not None:
                timestamp = internal_ts

    if product is None:
        product = path.stem.split("_", 1)[0]

    if timestamp is None:
        timestamp = datetime.fromtimestamp(path.stat().st_mtime)

    return TextRun(path=path, product=product, timestamp=timestamp, sections=sections)


def parse_picture_name(path: Path) -> str:
    m = PICTURE_RE.search(path.stem)
    if m:
        return m.group(1).lower().replace(" ", "_").replace("-", "_")
    return path.stem.lower()


def parse_image_file(path: Path) -> ImageRun | None:
    product = None
    timestamp = None

    tokens = [path.parent.name, path.parent.parent.name, path.stem]
    for token in tokens:
        p, ts = parse_product_timestamp_from_token(token)
        if p:
            product = p
            timestamp = ts
            break

    if product is None:
        return None

    if timestamp is None:
        timestamp = datetime.fromtimestamp(path.stat().st_mtime)

    return ImageRun(
        path=path,
        product=product,
        timestamp=timestamp,
        picture_name=parse_picture_name(path),
    )


def scan_text_folder(folder: Path) -> list[TextRun]:
    runs = []
    for path in folder.rglob("*.txt"):
        try:
            runs.append(parse_txt_file(path))
        except Exception as exc:
            print(f"Erreur TXT {path}: {exc}")
    return runs


def scan_image_folder(folder: Path) -> list[ImageRun]:
    runs = []
    for path in folder.rglob("*"):
        if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS:
            try:
                img = parse_image_file(path)
                if img is not None:
                    runs.append(img)
            except Exception as exc:
                print(f"Erreur image {path}: {exc}")
    return runs



def load_index_cache() -> dict:
    try:
        if INDEX_CACHE_PATH.exists():
            with INDEX_CACHE_PATH.open("r", encoding="utf-8") as f:
                data = json.load(f)
            if data.get("version") == CACHE_VERSION:
                return data.get("files", {})
    except Exception as exc:
        print(f"Cache index illisible : {exc}")
    return {}


def save_index_cache(files_cache: dict):
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        data = {
            "version": CACHE_VERSION,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "files": files_cache,
        }
        tmp = INDEX_CACHE_PATH.with_suffix(".tmp")
        with tmp.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        tmp.replace(INDEX_CACHE_PATH)
    except Exception as exc:
        print(f"Impossible d'écrire le cache index : {exc}")


def dt_to_cache(dt: datetime | None) -> str | None:
    return dt.isoformat() if dt else None


def dt_from_cache(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return None


def text_run_to_cache(run: TextRun, stat: os.stat_result) -> dict:
    return {
        "kind": "txt",
        "mtime_ns": int(stat.st_mtime_ns),
        "size": int(stat.st_size),
        "product": run.product,
        "timestamp": dt_to_cache(run.timestamp),
        "sections": run.sections,
    }


def image_run_to_cache(run: ImageRun, stat: os.stat_result) -> dict:
    return {
        "kind": "image",
        "mtime_ns": int(stat.st_mtime_ns),
        "size": int(stat.st_size),
        "product": run.product,
        "timestamp": dt_to_cache(run.timestamp),
        "picture_name": run.picture_name,
    }


def text_run_from_cache(path_str: str, entry: dict) -> TextRun:
    return TextRun(
        path=Path(path_str),
        product=str(entry.get("product", "")),
        timestamp=dt_from_cache(entry.get("timestamp")),
        sections=entry.get("sections", {}) or {},
    )


def image_run_from_cache(path_str: str, entry: dict) -> ImageRun:
    return ImageRun(
        path=Path(path_str),
        product=str(entry.get("product", "")),
        timestamp=dt_from_cache(entry.get("timestamp")),
        picture_name=str(entry.get("picture_name", "")),
    )


def is_cache_entry_valid(path: Path, entry: dict, kind: str, stat: os.stat_result) -> bool:
    return (
        entry.get("kind") == kind
        and int(entry.get("mtime_ns", -1)) == int(stat.st_mtime_ns)
        and int(entry.get("size", -1)) == int(stat.st_size)
    )


def path_under_root(path_str: str, root: Path) -> bool:
    # Important pour le mode "cache seulement" : ne pas appeler Path.resolve(),
    # car cela peut réveiller le serveur ou ralentir fortement sur lecteurs réseau.
    try:
        root_s = os.path.normcase(os.path.abspath(str(root)))
        path_s = os.path.normcase(os.path.abspath(str(path_str)))
    except Exception:
        root_s = str(root).lower()
        path_s = str(path_str).lower()
    return path_s.startswith(root_s)


def cache_records_for_roots(files_cache: dict, text_dir: Path, image_dir: Path | None, product_filter: str = "") -> tuple[list[TextRun], list[ImageRun]]:
    product_filter = product_filter.strip()
    text_runs: list[TextRun] = []
    image_runs: list[ImageRun] = []

    for path_str, entry in files_cache.items():
        product = str(entry.get("product", ""))
        if product_filter and product_filter not in product and product_filter not in Path(path_str).name:
            continue

        if entry.get("kind") == "txt" and path_under_root(path_str, text_dir):
            text_runs.append(text_run_from_cache(path_str, entry))
        elif image_dir is not None and entry.get("kind") == "image" and path_under_root(path_str, image_dir):
            image_runs.append(image_run_from_cache(path_str, entry))

    return text_runs, image_runs

def iter_files_fast(root: Path, suffixes: set[str], cancel_event: threading.Event,
                    product_filter: str = "", kind: str = "", progress_callback=None):
    """Parcours plus rapide que Path.rglob, avec filtrage optionnel.

    Le filtrage module ne supprime pas totalement le coût de parcours réseau,
    mais il évite de parser les fichiers non pertinents.
    """
    stack = [str(root)]
    visited = 0
    product_filter = product_filter.strip()

    while stack:
        if cancel_event.is_set():
            return
        current = stack.pop()
        try:
            with os.scandir(current) as it:
                for entry in it:
                    if cancel_event.is_set():
                        return
                    try:
                        if entry.is_dir(follow_symlinks=False):
                            stack.append(entry.path)
                        elif entry.is_file(follow_symlinks=False):
                            ext = Path(entry.name).suffix.lower()
                            if ext not in suffixes:
                                continue

                            if product_filter:
                                check_text = entry.name if kind == "txt" else entry.path
                                if product_filter not in check_text:
                                    continue

                            visited += 1
                            if progress_callback and visited % 200 == 0:
                                progress_callback(f"Scan {kind} : {visited} fichier(s) parcouru(s)…", visited)
                            yield Path(entry.path)
                    except OSError:
                        continue
        except OSError:
            continue



def file_datetime_from_name_or_mtime(path: Path) -> datetime | None:
    """Date rapide utilisée pour le filtrage scan.

    Malgré le nom historique de la fonction, on n'utilise plus la date dans le nom.
    Le scan filtre uniquement sur la date de création fichier, sans ouvrir le TXT.
    L'Horodatage interne reste uniquement utilisé pour l'affichage rapport/comparaison.
    """
    return file_creation_datetime(path)


def entry_mtime_datetime(entry) -> datetime | None:
    try:
        return datetime.fromtimestamp(entry.stat(follow_symlinks=False).st_mtime)
    except Exception:
        return None


def date_in_interval(dt: datetime | None, start_dt: datetime | None, end_dt: datetime | None) -> bool:
    if dt is None:
        return True
    if start_dt is not None and dt < start_dt:
        return False
    if end_dt is not None and dt >= end_dt + timedelta(days=1):
        return False
    return True


def iter_txt_files_newest_first(root: Path, cancel_event: threading.Event,
                                product_filter: str = "",
                                start_dt: datetime | None = None,
                                end_dt: datetime | None = None,
                                progress_callback=None):
    """Charge complètement le dossier TXT, puis trie par date.

    Objectif : éviter que l'ordre naturel du serveur, souvent anciens fichiers d'abord,
    impose l'ordre de traitement. On charge d'abord tous les chemins TXT accessibles,
    puis seulement après on trie du plus récent au plus ancien.
    """
    product_filter = product_filter.strip()
    stack = [Path(root)]
    all_txt: list[tuple[datetime, Path]] = []
    no_date_txt: list[Path] = []
    visited_dirs = 0

    if progress_callback:
        progress_callback("Chargement complet du dossier TXT avant tri par date…", 0)

    # Étape 1 : chargement complet des chemins TXT.
    while stack:
        if cancel_event.is_set():
            return

        current = stack.pop()
        visited_dirs += 1

        try:
            # list(...) force le chargement complet du dossier courant avant de continuer.
            entries = list(os.scandir(current))
        except OSError:
            continue

        for entry in entries:
            if cancel_event.is_set():
                return
            try:
                if entry.is_dir(follow_symlinks=False):
                    stack.append(Path(entry.path))
                elif entry.is_file(follow_symlinks=False):
                    if Path(entry.name).suffix.lower() != ".txt":
                        continue
                    if product_filter and product_filter not in entry.name:
                        continue

                    path = Path(entry.path)
                    dt = file_datetime_from_name_or_mtime(path)
                    if dt is None:
                        no_date_txt.append(path)
                    else:
                        all_txt.append((dt, path))

                    total = len(all_txt) + len(no_date_txt)
                    if progress_callback and total % 1000 == 0:
                        progress_callback(
                            f"Chargement TXT : {total} fichier(s) trouvé(s), {visited_dirs} dossier(s)…",
                            total,
                        )
            except OSError:
                continue

    if progress_callback:
        progress_callback(f"Tri par date de {len(all_txt) + len(no_date_txt)} fichier(s) TXT…", len(all_txt) + len(no_date_txt))

    # Étape 2 : tri global du plus récent au plus ancien.
    all_txt.sort(key=lambda item: item[0], reverse=True)

    # Étape 3 : on fournit seulement les fichiers compatibles avec l'intervalle demandé.
    yielded = 0
    for dt, path in all_txt:
        if cancel_event.is_set():
            return

        if end_dt is not None and dt >= end_dt + timedelta(days=1):
            continue

        if start_dt is not None and dt < start_dt:
            # Comme la liste est triée du plus récent au plus ancien,
            # tous les fichiers suivants seront encore plus vieux.
            break

        yielded += 1
        if progress_callback and yielded % 500 == 0:
            progress_callback(f"TXT dans l'intervalle fichier transmis : {yielded}", yielded)
        yield path

    # Les fichiers sans date rapide sont rares. On les transmet à la fin pour que
    # l'Horodatage= interne puisse trancher, sans bloquer les fichiers datés.
    for path in no_date_txt:
        if cancel_event.is_set():
            return
        yielded += 1
        yield path



def image_product_from_path_fast(path: Path) -> str | None:
    """Extraction légère du module depuis le chemin image, sans ouvrir l'image."""
    tokens = []
    try:
        tokens.extend([path.parent.name, path.parent.parent.name, path.parent.parent.parent.name, path.stem])
    except Exception:
        tokens.append(path.stem)

    for token in tokens:
        product, _ts = parse_product_timestamp_from_token(str(token))
        if product:
            return product
    return None



def file_creation_datetime(path: Path) -> datetime | None:
    """Date utilisée pour filtrer le scan.

    Important : le scan ne doit PAS ouvrir le TXT pour lire Horodatage=
    et ne doit PAS utiliser la date contenue dans le nom du fichier.
    On utilise uniquement la date de création du fichier côté OS,
    avec mtime en fallback si la création n'est pas disponible.
    """
    try:
        stat = path.stat()
        return datetime.fromtimestamp(getattr(stat, "st_ctime", stat.st_mtime))
    except Exception:
        try:
            return datetime.fromtimestamp(path.stat().st_mtime)
        except Exception:
            return None


# Ancien nom gardé pour compatibilité interne : il retourne maintenant seulement la date de création.
def file_datetime_from_name_or_ctime(path: Path) -> datetime | None:
    return file_creation_datetime(path)


def unique_paths(paths):
    seen = set()
    out = []
    for p in paths:
        try:
            key = str(Path(p).resolve())
        except Exception:
            key = str(p)
        if key not in seen:
            seen.add(key)
            out.append(Path(p))
    return out



def fast_directory_glob(directory: Path, pattern: str, want_dirs: bool | None = None) -> list[Path]:
    """Glob non récursif optimisé pour éviter le scan profond serveur."""
    directory = Path(directory)
    if not directory.exists():
        return []

    if os.name == "nt":
        try:
            import ctypes
            from ctypes import wintypes

            FILE_ATTRIBUTE_DIRECTORY = 0x10
            INVALID_HANDLE_VALUE = ctypes.c_void_p(-1).value

            class WIN32_FIND_DATAW(ctypes.Structure):
                _fields_ = [
                    ("dwFileAttributes", wintypes.DWORD),
                    ("ftCreationTime", wintypes.FILETIME),
                    ("ftLastAccessTime", wintypes.FILETIME),
                    ("ftLastWriteTime", wintypes.FILETIME),
                    ("nFileSizeHigh", wintypes.DWORD),
                    ("nFileSizeLow", wintypes.DWORD),
                    ("dwReserved0", wintypes.DWORD),
                    ("dwReserved1", wintypes.DWORD),
                    ("cFileName", wintypes.WCHAR * 260),
                    ("cAlternateFileName", wintypes.WCHAR * 14),
                    ("dwFileType", wintypes.DWORD),
                    ("dwCreatorType", wintypes.DWORD),
                    ("wFinderFlags", wintypes.WORD),
                ]

            kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
            find_first = kernel32.FindFirstFileW
            find_first.argtypes = [wintypes.LPCWSTR, ctypes.POINTER(WIN32_FIND_DATAW)]
            find_first.restype = wintypes.HANDLE
            find_next = kernel32.FindNextFileW
            find_next.argtypes = [wintypes.HANDLE, ctypes.POINTER(WIN32_FIND_DATAW)]
            find_next.restype = wintypes.BOOL
            find_close = kernel32.FindClose
            find_close.argtypes = [wintypes.HANDLE]
            find_close.restype = wintypes.BOOL

            data = WIN32_FIND_DATAW()
            handle = find_first(str(directory / pattern), ctypes.byref(data))
            if handle == INVALID_HANDLE_VALUE:
                return []

            out = []
            try:
                while True:
                    name = data.cFileName
                    if name not in (".", ".."):
                        is_dir = bool(data.dwFileAttributes & FILE_ATTRIBUTE_DIRECTORY)
                        if want_dirs is None or is_dir == want_dirs:
                            out.append(directory / name)
                    if not find_next(handle, ctypes.byref(data)):
                        break
            finally:
                find_close(handle)
            return out
        except Exception:
            pass

    try:
        out = []
        for p in directory.glob(pattern):
            try:
                if want_dirs is None or p.is_dir() == want_dirs:
                    out.append(p)
            except OSError:
                continue
        return out
    except OSError:
        return []


def follow_txt_candidate_paths(text_dir: Path, target: str, progress_callback=None) -> list[Path]:
    """Mode suivi ultra-rapide : pas de scan récursif serveur.

    On lit uniquement dans le dossier TXT sélectionné :
    - NUMERO.txt
    - NUMERO*.txt pour les fichiers avec date/suffixe
    """
    target = str(target).strip()
    if not target:
        return []

    text_dir = Path(text_dir)
    candidates = []

    direct = text_dir / f"{target}.txt"
    try:
        if direct.is_file():
            candidates.append(direct)
    except OSError:
        pass

    candidates.extend(fast_directory_glob(text_dir, f"{target}*.txt", want_dirs=False))
    candidates = [p for p in unique_paths(candidates) if p.suffix.lower() == ".txt" and p.name.startswith(target)]

    if progress_callback:
        progress_callback(
            f"Mode suivi TXT {target} : {len(candidates)} fichier(s) trouvé(s) directement dans le dossier TXT.",
            len(candidates),
        )
    return candidates



def follow_image_candidate_roots(image_dir: Path, product: str, progress_callback=None) -> list[Path]:
    """Recherche ultra-rapide du dossier images du module.

    Principe voulu :
    - dossier_images/NUMERO
    - dossier_images/NUMERO*
    puis fallback rapide seulement si besoin :
    - dossier_images/QIA/NUMERO*

    Important : on ne scanne plus les dossiers niveau 1 du répertoire image,
    car sur serveur cela peut redevenir lent.
    """
    product = str(product).strip()
    if not product:
        return []

    image_dir = Path(image_dir)
    roots = []

    # 1) Cas le plus rapide : dossier exact.
    direct = image_dir / product
    try:
        if direct.is_dir():
            roots.append(direct)
    except OSError:
        pass

    # 2) Même dossier uniquement : NUMERO*.
    if not roots:
        roots.extend(fast_directory_glob(image_dir, f"{product}*", want_dirs=True))

    # 3) Fallback rapide si le dossier image sélectionné est au-dessus d'un dossier QIA.
    # Pas de scan profond.
    if not roots:
        roots.extend(fast_directory_glob(image_dir / "QIA", f"{product}*", want_dirs=True))

    roots = unique_paths(roots)

    if progress_callback:
        progress_callback(
            f"Mode suivi images {product} : {len(roots)} dossier(s) trouvé(s) directement dans le dossier images.",
            len(roots),
        )
    return roots


def follow_image_candidate_files(image_dir: Path, product: str) -> list[Path]:
    """Fallback non récursif : images directement dans le dossier image.

    On ne cherche que les fichiers du type NUMERO* dans le dossier image sélectionné.
    """
    product = str(product).strip()
    if not product:
        return []

    image_dir = Path(image_dir)
    candidates = []
    for ext in IMAGE_EXTENSIONS:
        candidates.extend(fast_directory_glob(image_dir, f"{product}*{ext}", want_dirs=False))
        candidates.extend(fast_directory_glob(image_dir / "QIA", f"{product}*{ext}", want_dirs=False))
    return unique_paths(candidates)

def iter_images_under_roots_for_follow(roots: list[Path], cancel_event: threading.Event,
                                       progress_callback=None):
    """Parcourt uniquement les dossiers module déjà trouvés."""
    visited = 0
    stack = [str(r) for r in roots if r.exists()]

    while stack:
        if cancel_event.is_set():
            return
        current = stack.pop()
        try:
            with os.scandir(current) as it:
                for entry in it:
                    if cancel_event.is_set():
                        return
                    try:
                        if entry.is_dir(follow_symlinks=False):
                            stack.append(entry.path)
                        elif entry.is_file(follow_symlinks=False):
                            if Path(entry.name).suffix.lower() in IMAGE_EXTENSIONS:
                                visited += 1
                                if progress_callback and visited % 100 == 0:
                                    progress_callback(f"Mode suivi images : {visited} image(s) lue(s) dans le dossier module…", visited)
                                yield Path(entry.path)
                    except OSError:
                        continue
        except OSError:
            continue


def scan_follow_targets_indexed(text_dir: Path, image_dir: Path | None, follow_targets: list[str],
                                cache_only: bool = False,
                                cancel_event: threading.Event | None = None,
                                progress_callback=None,
                                max_files_per_type: int | None = None,
                                start_dt: datetime | None = None,
                                end_dt: datetime | None = None,
                                txt_only: bool = False) -> tuple[list[TextRun], list[ImageRun], dict]:
    """Scan optimisé pour le mode suivi.

    Contrairement au scan standard, il ne parcourt pas les 200 000 fichiers.
    Pour chaque module, il cherche directement les chemins du type :
    - dossier_txt\\numero_module*.txt
    - dossier_images\\numero_module*\\images
    et applique les bornes de dates avant parsing quand une date rapide est disponible.
    """
    cancel_event = cancel_event or threading.Event()
    files_cache = load_index_cache()

    if max_files_per_type is None:
        max_files_per_type = get_max_files_per_type()
    try:
        max_files_per_type = int(max_files_per_type)
    except Exception:
        max_files_per_type = 1500

    targets = [str(t).strip() for t in follow_targets if str(t).strip()]
    targets = list(dict.fromkeys(targets))

    parsed_txt = 0
    parsed_images = 0
    cache_hits = 0
    visited_txt = 0
    visited_images = 0
    skipped_txt_outside_date = 0
    skipped_images_outside_date = 0

    if cache_only:
        text_runs_all = []
        image_runs_all = []
        for target in targets:
            tr, ir = cache_records_for_roots(files_cache, text_dir, image_dir, target)
            text_runs_all.extend(tr)
            image_runs_all.extend(ir)

        text_runs = [r for r in text_runs_all if date_in_interval(file_creation_datetime(r.path), start_dt, end_dt)]
        allowed_products = {r.product for r in text_runs}
        image_runs = [img for img in image_runs_all if img.product in allowed_products]
        image_runs = [img for img in image_runs if date_in_interval(file_creation_datetime(img.path), start_dt, end_dt)]

        return text_runs, image_runs, {
            "from_cache_only": True,
            "follow_optimized": True,
            "parsed_txt": 0,
            "parsed_images": 0,
            "cache_hits": len(text_runs) + len(image_runs),
            "visited_txt": len(text_runs),
            "visited_images": len(image_runs),
            "max_files_per_type": max_files_per_type,
            "follow_count": len(targets),
            "selected_txt": len(text_runs),
            "skipped_txt_outside_date": 0,
            "skipped_images_outside_date": 0,
        }

    text_runs: list[TextRun] = []

    for target_idx, target in enumerate(targets, start=1):
        if cancel_event.is_set():
            break

        if progress_callback:
            progress_callback(f"Mode suivi optimisé TXT : {target} ({target_idx}/{len(targets)})", target_idx)

        candidates = follow_txt_candidate_paths(text_dir, target, progress_callback=progress_callback)

        # Filtre date rapide avant parsing : nom de fichier, sinon date de création fichier.
        dated_candidates = []
        for p in candidates:
            quick_dt = file_datetime_from_name_or_ctime(p)
            if not date_in_interval(quick_dt, start_dt, end_dt):
                skipped_txt_outside_date += 1
                continue
            dated_candidates.append((quick_dt or datetime.min, p))

        dated_candidates.sort(key=lambda x: x[0], reverse=True)

        kept_for_target = 0
        for _dt, path in dated_candidates:
            if cancel_event.is_set():
                break
            if max_files_per_type > 0 and kept_for_target >= max_files_per_type:
                break

            visited_txt += 1

            try:
                stat = path.stat()
                key = str(path)
                entry = files_cache.get(key)

                if entry and is_cache_entry_valid(path, entry, "txt", stat):
                    run = text_run_from_cache(key, entry)
                    cache_hits += 1
                else:
                    run = parse_txt_file(path)
                    files_cache[key] = text_run_to_cache(run, stat)
                    parsed_txt += 1

                # Sécurité module : on accepte si le target est dans le produit ou dans le nom du fichier.
                if target not in run.product and target not in run.path.name:
                    continue

                # Ne pas filtrer avec Horodatage= : le mode suivi filtre uniquement
                # avec la date de création fichier avant ouverture du TXT.
                text_runs.append(run)
                kept_for_target += 1

            except Exception as exc:
                print(f"Erreur TXT suivi {path}: {exc}")

    # Déduplication des TXT.
    unique_text = {}
    for run in text_runs:
        unique_text[str(run.path)] = run
    text_runs = sorted(unique_text.values(), key=lambda r: (r.product, r.timestamp or datetime.min))
    allowed_products = {r.product for r in text_runs}

    image_runs: list[ImageRun] = []
    if allowed_products and not txt_only and image_dir is not None:
        for product_idx, product in enumerate(sorted(allowed_products), start=1):
            if cancel_event.is_set():
                break

            if progress_callback:
                progress_callback(f"Mode suivi optimisé images : {product} ({product_idx}/{len(allowed_products)})", product_idx)

            roots = follow_image_candidate_roots(image_dir, product, progress_callback=progress_callback)
            paths = list(iter_images_under_roots_for_follow(roots, cancel_event, progress_callback=progress_callback))

            # Fallback ciblé : fichiers à la racine ou 1/2 niveaux si aucun dossier module n'a été trouvé.
            if not paths:
                paths = follow_image_candidate_files(image_dir, product)

            paths = unique_paths(paths)

            for path in paths:
                if cancel_event.is_set():
                    break

                quick_dt = file_datetime_from_name_or_ctime(path)
                if not date_in_interval(quick_dt, start_dt, end_dt):
                    skipped_images_outside_date += 1
                    continue

                visited_images += 1

                try:
                    stat = path.stat()
                    key = str(path)
                    entry = files_cache.get(key)

                    if entry and is_cache_entry_valid(path, entry, "image", stat):
                        img = image_run_from_cache(key, entry)
                        cache_hits += 1
                    else:
                        img = parse_image_file(path)
                        if img is None:
                            continue
                        files_cache[key] = image_run_to_cache(img, stat)
                        parsed_images += 1

                    if img.product not in allowed_products:
                        continue

                    # Date déjà filtrée avant parsing avec la création fichier.
                    image_runs.append(img)

                except Exception as exc:
                    print(f"Erreur image suivi {path}: {exc}")

    unique_images = {}
    for run in image_runs:
        unique_images[str(run.path)] = run
    image_runs = sorted(unique_images.values(), key=lambda r: (r.product, r.timestamp or datetime.min))

    save_index_cache(files_cache)
    return text_runs, image_runs, {
        "from_cache_only": False,
        "follow_optimized": True,
        "parsed_txt": parsed_txt,
        "parsed_images": parsed_images,
        "cache_hits": cache_hits,
        "visited_txt": visited_txt,
        "visited_images": visited_images,
        "max_files_per_type": max_files_per_type,
        "follow_count": len(targets),
        "selected_txt": len(text_runs),
        "skipped_txt_outside_date": skipped_txt_outside_date,
        "skipped_images_outside_date": skipped_images_outside_date,
    }


def scan_folders_indexed(text_dir: Path, image_dir: Path | None, product_filter: str = "",
                         cache_only: bool = False, cancel_event: threading.Event | None = None,
                         progress_callback=None, max_files_per_type: int | None = None,
                         start_dt: datetime | None = None,
                         end_dt: datetime | None = None,
                         txt_only: bool = False) -> tuple[list[TextRun], list[ImageRun], dict]:
    """Scan en deux étapes.

    1) Le dossier TXT est la source de vérité.
    2) Le dossier TXT est chargé complètement avant le tri.
    3) Le parcours TXT est ensuite fait du plus récent au plus ancien.
    4) Les TXT sont gardés uniquement si leur date de création fichier est dans l'intervalle demandé.
    5) Le scan s'arrête à la limite de TXT retenus, par défaut 1500.
    6) Les images sont ensuite recherchées seulement pour ces modules.
    """
    cancel_event = cancel_event or threading.Event()
    files_cache = load_index_cache()
    product_filter = product_filter.strip()

    if max_files_per_type is None:
        max_files_per_type = get_max_files_per_type()
    try:
        max_files_per_type = int(max_files_per_type)
    except Exception:
        max_files_per_type = 1500

    parsed_txt = 0
    parsed_images = 0
    cache_hits = 0
    visited_txt = 0
    skipped_txt_outside_date = 0
    visited_images = 0
    skipped_images_not_in_txt_modules = 0

    if cache_only:
        text_runs, image_runs = cache_records_for_roots(files_cache, text_dir, image_dir, product_filter)
        # Filtre de scan = date de création fichier, pas Horodatage=.
        text_runs = [r for r in text_runs if date_in_interval(file_creation_datetime(r.path), start_dt, end_dt)]
        text_runs = sorted(text_runs, key=lambda r: file_creation_datetime(r.path) or datetime.min)
        if max_files_per_type > 0:
            text_runs = text_runs[:max_files_per_type]
        allowed_products = {r.product for r in text_runs}
        if txt_only or image_dir is None:
            image_runs = []
        else:
            image_runs = [img for img in image_runs if img.product in allowed_products]
            image_runs = [img for img in image_runs if date_in_interval(file_creation_datetime(img.path), start_dt, end_dt)]
        return text_runs, image_runs, {
            "from_cache_only": True,
            "parsed_txt": 0,
            "parsed_images": 0,
            "cache_hits": len(text_runs) + len(image_runs),
            "visited_txt": len(text_runs),
            "visited_images": len(image_runs),
            "max_files_per_type": max_files_per_type,
            "skipped_txt_outside_date": 0,
            "skipped_images_not_in_txt_modules": 0,
            "selected_txt": len(text_runs),
        }

    text_runs: list[TextRun] = []
    if progress_callback:
        progress_callback("Scan TXT : chargement complet puis tri par date…", 0)

    for path in iter_txt_files_newest_first(text_dir, cancel_event, product_filter, start_dt, end_dt, progress_callback):
        if cancel_event.is_set():
            break

        visited_txt += 1

        if max_files_per_type > 0 and len(text_runs) >= max_files_per_type:
            if progress_callback:
                progress_callback(f"Limite TXT retenus atteinte ({max_files_per_type}).", len(text_runs))
            break

        try:
            stat = path.stat()
            key = str(path)
            entry = files_cache.get(key)

            if entry and is_cache_entry_valid(path, entry, "txt", stat):
                run = text_run_from_cache(key, entry)
                cache_hits += 1
            else:
                run = parse_txt_file(path)
                files_cache[key] = text_run_to_cache(run, stat)
                parsed_txt += 1

            if product_filter and product_filter not in run.product and product_filter not in run.path.name:
                continue

            # Ne pas filtrer ici avec Horodatage=.
            # Le filtrage date du scan a déjà été fait avant ouverture du TXT,
            # uniquement avec la date de création fichier.
            text_runs.append(run)

            if progress_callback and len(text_runs) % 50 == 0:
                progress_callback(
                    f"TXT retenus dans l'intervalle : {len(text_runs)} / {max_files_per_type}",
                    len(text_runs),
                )

        except Exception as exc:
            print(f"Erreur TXT {path}: {exc}")

    text_runs = sorted(text_runs, key=lambda r: r.timestamp or datetime.min)
    allowed_products = {r.product for r in text_runs}

    if not allowed_products:
        save_index_cache(files_cache)
        return text_runs, [], {
            "from_cache_only": False,
            "parsed_txt": parsed_txt,
            "parsed_images": 0,
            "cache_hits": cache_hits,
            "visited_txt": visited_txt,
            "visited_images": 0,
            "max_files_per_type": max_files_per_type,
            "skipped_txt_outside_date": skipped_txt_outside_date,
            "skipped_images_not_in_txt_modules": 0,
            "selected_txt": len(text_runs),
        }

    if txt_only or image_dir is None:
        save_index_cache(files_cache)
        return text_runs, [], {
            "from_cache_only": False,
            "txt_only": True,
            "parsed_txt": parsed_txt,
            "parsed_images": 0,
            "cache_hits": cache_hits,
            "visited_txt": visited_txt,
            "visited_images": 0,
            "max_files_per_type": max_files_per_type,
            "skipped_txt_outside_date": skipped_txt_outside_date,
            "skipped_images_not_in_txt_modules": 0,
            "selected_txt": len(text_runs),
        }

    if progress_callback:
        progress_callback(f"Scan images : recherche des images correspondant aux {len(allowed_products)} module(s) TXT retenus…", 0)

    image_runs: list[ImageRun] = []
    max_images_to_keep = max_files_per_type * 20 if max_files_per_type > 0 else 0

    for path in iter_files_fast(image_dir, IMAGE_EXTENSIONS, cancel_event, "", "image", progress_callback):
        if cancel_event.is_set():
            break

        # Filtre date du scan : création fichier uniquement, sans ouvrir l'image.
        if not date_in_interval(file_creation_datetime(path), start_dt, end_dt):
            continue

        product = image_product_from_path_fast(path)
        if product not in allowed_products:
            skipped_images_not_in_txt_modules += 1
            continue

        visited_images += 1

        try:
            stat = path.stat()
            key = str(path)
            entry = files_cache.get(key)

            if entry and is_cache_entry_valid(path, entry, "image", stat):
                img = image_run_from_cache(key, entry)
                cache_hits += 1
            else:
                img = parse_image_file(path)
                if img is None:
                    continue
                files_cache[key] = image_run_to_cache(img, stat)
                parsed_images += 1

            if img.product not in allowed_products:
                skipped_images_not_in_txt_modules += 1
                continue

            image_runs.append(img)

            if max_images_to_keep > 0 and len(image_runs) >= max_images_to_keep:
                if progress_callback:
                    progress_callback(f"Limite images de sécurité atteinte ({max_images_to_keep}).", len(image_runs))
                break

        except Exception as exc:
            print(f"Erreur image {path}: {exc}")

    save_index_cache(files_cache)
    return text_runs, image_runs, {
        "from_cache_only": False,
        "parsed_txt": parsed_txt,
        "parsed_images": parsed_images,
        "cache_hits": cache_hits,
        "visited_txt": visited_txt,
        "visited_images": visited_images,
        "max_files_per_type": max_files_per_type,
        "skipped_txt_outside_date": skipped_txt_outside_date,
        "skipped_images_not_in_txt_modules": skipped_images_not_in_txt_modules,
        "selected_txt": len(text_runs),
    }



def preview_cache_path(path: Path) -> Path:
    preview_max_size = get_preview_max_size()
    try:
        stat = path.stat()
        token = f"{path}|{stat.st_mtime_ns}|{stat.st_size}|{preview_max_size}"
    except Exception:
        token = f"{path}|{preview_max_size}"
    digest = hashlib.sha1(token.encode("utf-8", errors="ignore")).hexdigest()
    return THUMB_CACHE_DIR / f"{digest}.jpg"


def load_preview_or_create(path: Path):
    """Charge une prévisualisation locale. La pleine résolution reste disponible via bouton."""
    if Image is None:
        return None
    THUMB_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = preview_cache_path(path)

    if cache_path.exists():
        return Image.open(cache_path).convert("RGB")

    img = Image.open(path).convert("RGB")
    preview = img.copy()
    preview_max_size = get_preview_max_size()
    preview.thumbnail((preview_max_size, preview_max_size), Image.Resampling.LANCZOS)
    try:
        preview.save(cache_path, "JPEG", quality=92)
    except Exception:
        pass
    return preview



def filter_by_interval(records, days: int | None):
    # Conservé par compatibilité. La V6 utilise plutôt filter_by_date_range().
    if days is None:
        return list(records)

    timestamps = [r.timestamp for r in records if r.timestamp is not None]
    if not timestamps:
        return list(records)

    last_date = max(timestamps)
    limit = last_date - timedelta(days=days)
    return [r for r in records if r.timestamp is None or r.timestamp >= limit]


def parse_date_entry(value: str) -> datetime | None:
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    return None


def filter_by_date_range(records, start_dt: datetime | None, end_dt: datetime | None):
    if start_dt is None and end_dt is None:
        return list(records)
    # fin inclusive : on ajoute un jour si l'utilisateur a choisi une date sans heure
    end_exclusive = end_dt + timedelta(days=1) if end_dt is not None else None
    out = []
    for r in records:
        ts = getattr(r, "timestamp", None)
        if ts is None:
            out.append(r)
            continue
        if start_dt is not None and ts < start_dt:
            continue
        if end_exclusive is not None and ts >= end_exclusive:
            continue
        out.append(r)
    return out


def date_label(dt: datetime | None) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "date inconnue"


def run_label(run: TextRun) -> str:
    return f"{date_label(run.timestamp)}  |  {run.path.name}"


def split_fields(line: str) -> tuple[list[str], str]:
    if ";" in line:
        return line.split(";"), ";"
    if "=" in line:
        return line.split("="), "="
    return [line], ""


def to_float(value: str) -> float | None:
    try:
        return float(value.replace(",", "."))
    except ValueError:
        return None


def line_signature(line: str) -> tuple[str, ...]:
    fields, _sep = split_fields(line)
    sig = []
    for field in fields:
        if to_float(field) is None:
            sig.append(field.strip().upper())
    return tuple(sig)


def numeric_values(line: str) -> list[float]:
    fields, _sep = split_fields(line)
    nums = []
    for field in fields:
        value = to_float(field)
        if value is not None:
            nums.append(value)
    return nums


def can_fuzzy_align(section: str, ref_lines: list[str], comp_lines: list[str]) -> bool:
    if section.upper() not in FUZZY_ALIGN_SECTIONS:
        return False
    lines = ref_lines + comp_lines
    if not lines:
        return False
    return any(line_signature(line) for line in lines)


def fuzzy_score(ref_line: str, comp_line: str) -> float:
    ref_sig = line_signature(ref_line)
    comp_sig = line_signature(comp_line)

    if ref_sig and comp_sig and ref_sig != comp_sig:
        return 10_000.0

    ref_nums = numeric_values(ref_line)
    comp_nums = numeric_values(comp_line)
    if not ref_nums or not comp_nums:
        return 10_000.0

    score = 0.0
    if len(ref_nums) >= 1 and len(comp_nums) >= 1:
        score += abs(ref_nums[0] - comp_nums[0]) / 25.0
    if len(ref_nums) >= 2 and len(comp_nums) >= 2:
        score += abs(ref_nums[1] - comp_nums[1]) / 0.8
    if len(ref_nums) >= 3 and len(comp_nums) >= 3:
        score += abs(ref_nums[2] - comp_nums[2]) / 40.0
    return score


def align_lines(section: str, ref_lines: list[str], comp_lines: list[str]) -> list[tuple[str | None, str | None]]:
    if not can_fuzzy_align(section, ref_lines, comp_lines):
        max_len = max(len(ref_lines), len(comp_lines))
        return [
            (
                ref_lines[i] if i < len(ref_lines) else None,
                comp_lines[i] if i < len(comp_lines) else None,
            )
            for i in range(max_len)
        ]

    unused_ref = set(range(len(ref_lines)))
    pairs: list[tuple[int | None, int | None]] = []

    for comp_idx, comp_line in enumerate(comp_lines):
        best_ref_idx = None
        best_score = 10_000.0
        for ref_idx in unused_ref:
            score = fuzzy_score(ref_lines[ref_idx], comp_line)
            if score < best_score:
                best_score = score
                best_ref_idx = ref_idx
        if best_ref_idx is not None and best_score <= 8.0:
            pairs.append((best_ref_idx, comp_idx))
            unused_ref.remove(best_ref_idx)
        else:
            pairs.append((None, comp_idx))

    for ref_idx in sorted(unused_ref):
        insert_at = len(pairs)
        for i, (paired_ref_idx, _paired_comp_idx) in enumerate(pairs):
            if paired_ref_idx is not None and paired_ref_idx > ref_idx:
                insert_at = i
                break
        pairs.insert(insert_at, (ref_idx, None))

    return [
        (
            ref_lines[ref_idx] if ref_idx is not None else None,
            comp_lines[comp_idx] if comp_idx is not None else None,
        )
        for ref_idx, comp_idx in pairs
    ]


def section_order(ref: TextRun, comp: TextRun) -> list[str]:
    order = [s for s in ref.sections.keys() if s not in SKIP_SECTIONS_IN_COMPARE]
    for s in comp.sections.keys():
        if s not in order and s not in SKIP_SECTIONS_IN_COMPARE:
            order.append(s)
    return order


def normalize_angle(a: float) -> float:
    while a > PI_VAL:
        a -= 2 * PI_VAL
    while a <= -PI_VAL:
        a += 2 * PI_VAL
    return a


def angular_distance(a1: float, a2: float) -> float:
    return abs(normalize_angle(a1 - a2))


def fit_text(text: str, max_len: int = 32) -> str:
    text = str(text)
    if len(text) <= max_len:
        return text
    return text[: max_len - 1] + "…"


def color_by_zone(radius: float) -> str:
    colors = ["#00b050", "#ffc000", "#ff0000", "#7030a0", "#0070c0", "#c000c0"]
    zone_name = classify_zone(radius)
    for idx, zone in enumerate(get_zones()):
        if str(zone.get("name", "")).lower() == str(zone_name).lower():
            return colors[idx % len(colors)]
    return "#7030a0"


def extract_grade_value(run: TextRun, grade_key: str) -> str:
    for line in run.sections.get("GRADE", []):
        if line.upper().startswith(grade_key.upper() + "="):
            return line.split("=", 1)[1].strip()
    return ""


def is_probably_geometric_section(section: str, lines: list[str]) -> bool:
    if not lines:
        return False
    if section.upper() in {"PARAMETERS", "CONDITIONS", "GRADE", "CHICKENWIRE", "INFORMATION", "DECHET/DEFAUT", "RETARD", "TREATMENT"}:
        return False
    if section.upper() in GEOMETRIC_SECTIONS_PREFERRED:
        return True

    valid = 0
    for line in lines[:10]:
        fields, _ = split_fields(line)
        if len(fields) >= 2 and to_float(fields[0]) is not None and to_float(fields[1]) is not None:
            valid += 1
    return valid > 0


def parse_defect_point(section: str, line: str) -> DefectPoint | None:
    fields, sep = split_fields(line)
    if sep != ";" or len(fields) < 3:
        return None

    radius = to_float(fields[0])
    angle = to_float(fields[1])
    size = to_float(fields[2])
    if radius is None or angle is None or size is None:
        return None

    defect_type = fields[4].strip() if len(fields) >= 5 else section
    location = fields[5].strip() if len(fields) >= 6 else ""

    return DefectPoint(
        section=section,
        radius=radius,
        angle=angle,
        size=size,
        defect_type=defect_type,
        location=location,
        raw=line,
    )


def parse_geometric_points_by_section(run: TextRun) -> dict[str, list[DefectPoint]]:
    out: dict[str, list[DefectPoint]] = {}
    for section, lines in run.sections.items():
        if not is_probably_geometric_section(section, lines):
            continue
        pts = []
        for line in lines:
            pt = parse_defect_point(section, line)
            if pt is not None:
                pts.append(pt)
        if pts:
            out[section] = pts
    return out


def parse_summary_points(run: TextRun) -> list[DefectPoint]:
    grouped = parse_geometric_points_by_section(run)
    pts: list[DefectPoint] = []
    preferred_order = [
        "DARK SPOT",
        "WHITE SPOT",
        "SURFACE DEFECT",
        "FUNCTIONAL DEFECT",
        "GROWING SPOT",
        "EP OFF",
        "ECE",
        "ECS",
        "AE",
        "OPTICAL FEEDBACK",
        "HALO",
        "TA",
        "TB",
        "TBDelay",
        "INHO",
    ]
    for name in preferred_order:
        for section_name, section_pts in grouped.items():
            if section_name.upper() == name.upper():
                pts.extend(section_pts)
    # ajoute les autres éventuelles sections sans doublon
    for section_name, section_pts in grouped.items():
        if all(existing.raw != p.raw or existing.section != p.section for existing in pts for p in section_pts):
            pass
    return pts


def classify_zone(radius_um: float) -> str:
    for zone in get_zones():
        max_um = zone.get("max_um")
        if max_um is None:
            return str(zone.get("name", "Hors zone"))
        try:
            if radius_um < float(max_um):
                return str(zone.get("name", "Zone"))
        except Exception:
            continue
    return "Hors zone"


def classify_size(size_um: float) -> str | None:
    for bucket in get_size_buckets():
        try:
            name = str(bucket.get("name", ""))
            mini = float(bucket.get("min_um", 0))
            maxi = float(bucket.get("max_um", 0))
        except Exception:
            continue
        if mini <= size_um < maxi:
            return name
    return None


def summary_column_names() -> list[str]:
    cols = []
    for zone in get_zones():
        for bucket in get_size_buckets():
            cols.append(compact_column_label(str(zone.get("name", "")), str(bucket.get("name", ""))))
    return cols


def init_count_dict() -> dict[str, int]:
    return {name: 0 for name in summary_column_names()}


def point_matches_for_recurrence(cur: DefectPoint, old: DefectPoint,
                                 radius_tol: float = 300.0,
                                 angle_tol: float = 0.12,
                                 size_tol: float = 40.0) -> bool:
    if cur.section.upper() != old.section.upper():
        return False
    if abs(cur.radius - old.radius) > radius_tol:
        return False
    if angular_distance(cur.angle, old.angle) > angle_tol:
        return False
    if abs(cur.size - old.size) > size_tol:
        return False
    if cur.defect_type and old.defect_type and cur.defect_type.upper() != old.defect_type.upper():
        return False
    # La localisation est informative : si elle existe des deux côtés, on l'impose.
    if cur.location and old.location and cur.location.upper() != old.location.upper():
        return False
    return True


def summarize_run(run: TextRun, previous_runs: list[TextRun]) -> PassageSummary:
    grade_cw = extract_grade_value(run, "CW")
    grade_inho = extract_grade_value(run, "INHO")
    grades = {
        str(col.get("label", "")): extract_grade_value(run, str(col.get("key", "")))
        for col in get_grade_columns()
    }

    points = parse_summary_points(run)
    previous_points: list[DefectPoint] = []
    for old_run in previous_runs:
        previous_points.extend(parse_summary_points(old_run))

    counts = init_count_dict()
    locations = Counter()
    sections = Counter()
    recurrent_points: list[DefectPoint] = []
    over_300_count = 0

    for pt in points:
        sections[pt.section] += 1
        if pt.location:
            locations[pt.location] += 1
        zone = classify_zone(pt.radius)
        size_bucket = classify_size(pt.size)
        if size_bucket is not None:
            counts[compact_column_label(zone, size_bucket)] += 1
        elif pt.size > 300:
            over_300_count += 1

        if any(point_matches_for_recurrence(pt, prev) for prev in previous_points):
            recurrent_points.append(pt)

    return PassageSummary(
        run=run,
        grade_cw=grade_cw,
        grade_inho=grade_inho,
        grades=grades,
        counts=counts,
        total_points=len(points),
        over_300_count=over_300_count,
        recurrent_count=len(recurrent_points),
        location_counter=locations,
        section_counter=sections,
        recurrent_points=recurrent_points,
        all_points=points,
    )


def build_product_summaries(runs: list[TextRun]) -> list[PassageSummary]:
    runs = sorted(runs, key=lambda r: r.timestamp or datetime.min)
    summaries: list[PassageSummary] = []
    previous: list[TextRun] = []
    for run in runs:
        summaries.append(summarize_run(run, previous))
        previous.append(run)
    return summaries


def format_counter(counter: Counter, empty_text: str = "") -> str:
    if not counter:
        return empty_text
    return ", ".join(f"{k}:{v}" for k, v in sorted(counter.items(), key=lambda x: (-x[1], x[0])))


def percent_color(value: float) -> str:
    if value == 0:
        return "#d9ead3"
    if value < 30:
        return "#fff2cc"
    return "#f4cccc"


def grade_cell_color(value: str) -> str:
    v = to_float(value or "")
    if v is None:
        return "white"
    if v >= 5:
        return "#d9ead3"
    if v >= 3:
        return "#fff2cc"
    return "#f4cccc"


def count_cell_color(value: int) -> str:
    if value == 0:
        return "#d9ead3"
    if value == 1:
        return "#fff2cc"
    return "#f4cccc"


class SyncedImageGroup:
    def __init__(self):
        self.viewers: list["SyncedImageCanvas"] = []
        self.zoom = 1.0
        self.center_x = 0.0
        self.center_y = 0.0
        self.auto_fit = True
        self._fit_after_id = None
        self.full_resolution = False

    def add_viewer(self, viewer: "SyncedImageCanvas"):
        self.viewers.append(viewer)

    def schedule_fit(self, parent: tk.Widget):
        if not self.auto_fit:
            return
        if self._fit_after_id is not None:
            try:
                parent.after_cancel(self._fit_after_id)
            except Exception:
                pass
        self._fit_after_id = parent.after(120, self.fit_all)

    def fit_all(self):
        ready = [v for v in self.viewers if v.image is not None and v.canvas.winfo_width() > 20 and v.canvas.winfo_height() > 20]
        if not ready:
            return

        zooms = []
        for v in ready:
            w, h = v.image.size
            cw = max(v.canvas.winfo_width() - 12, 20)
            ch = max(v.canvas.winfo_height() - 12, 20)
            zooms.append(min(cw / w, ch / h))

        self.zoom = max(min(zooms), 0.01)
        first = ready[0]
        self.center_x = first.image.size[0] / 2
        self.center_y = first.image.size[1] / 2
        self.draw_all()

    def reset_fit(self):
        self.auto_fit = True
        self.fit_all()

    def set_full_resolution(self, enabled: bool):
        self.full_resolution = enabled
        for viewer in self.viewers:
            viewer.load_image(full_resolution=enabled)
        self.reset_fit()

    def zoom_at(self, viewer: "SyncedImageCanvas", x: int, y: int, factor: float):
        if viewer.image is None:
            return
        img_x, img_y = viewer.canvas_to_image(x, y)
        self.auto_fit = False
        self.zoom = max(0.05, min(self.zoom * factor, 20.0))
        self.center_x = img_x
        self.center_y = img_y
        self.draw_all()

    def zoom_center(self, factor: float):
        self.auto_fit = False
        self.zoom = max(0.05, min(self.zoom * factor, 20.0))
        self.draw_all()

    def pan(self, dx: float, dy: float):
        self.auto_fit = False
        if self.zoom <= 0:
            return
        self.center_x -= dx / self.zoom
        self.center_y -= dy / self.zoom
        self.draw_all()

    def draw_all(self):
        for viewer in self.viewers:
            viewer.draw()


class SyncedImageCanvas(ttk.Frame):
    def __init__(self, parent: tk.Widget, group: SyncedImageGroup, record: ImageRun,
                 rotation_store: dict[str, int]):
        super().__init__(parent, padding=4, relief="ridge")
        self.group = group
        self.record = record
        self.rotation_store = rotation_store
        self.image = None
        self.base_image = None
        self.tk_image = None
        self.image_item = None
        self._drag_start = None
        self.is_full_resolution = False
        self.source_size = None

        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        title = f"{date_label(record.timestamp)}  —  {record.path.name}"
        ttk.Label(self, text=title, anchor="center").grid(row=0, column=0, sticky="ew", pady=(0, 2))

        self.canvas = tk.Canvas(self, bg="black", highlightthickness=0)
        self.canvas.grid(row=1, column=0, sticky="nsew")

        btn_row = ttk.Frame(self)
        btn_row.grid(row=2, column=0, sticky="ew", pady=(4, 0))
        ttk.Button(btn_row, text="Tourner 180°", command=self.toggle_rotation).pack(side="left", padx=2)
        ttk.Button(btn_row, text="Réinitialiser", command=self.reset_rotation).pack(side="left", padx=2)
        ttk.Button(btn_row, text="Zones ON/OFF", command=self.toggle_zones).pack(side="left", padx=2)
        self.rotation_label = ttk.Label(btn_row, text=self.rotation_text())
        self.rotation_label.pack(side="right", padx=4)

        self.load_image(full_resolution=self.group.full_resolution)

        self.canvas.bind("<Configure>", self.on_configure)
        self.canvas.bind("<MouseWheel>", self.on_mousewheel)
        self.canvas.bind("<Button-4>", lambda e: self.group.zoom_at(self, e.x, e.y, 1.15))
        self.canvas.bind("<Button-5>", lambda e: self.group.zoom_at(self, e.x, e.y, 1 / 1.15))
        self.canvas.bind("<ButtonPress-1>", self.on_drag_start)
        self.canvas.bind("<B1-Motion>", self.on_drag_move)

        self.group.add_viewer(self)

    def rotation_key(self) -> str:
        return str(self.record.path)

    def current_rotation(self) -> int:
        return self.rotation_store.get(self.rotation_key(), 0)

    def rotation_text(self) -> str:
        mode = "pleine résolution" if self.is_full_resolution else "prévisualisation"
        return f"Rotation : {self.current_rotation()}° | {mode}"

    def load_image(self, full_resolution: bool = False):
        self.is_full_resolution = full_resolution
        try:
            if full_resolution:
                self.base_image = Image.open(self.record.path).convert("RGB")
                self.source_size = self.base_image.size
            else:
                try:
                    with Image.open(self.record.path) as src_img:
                        self.source_size = src_img.size
                except Exception:
                    self.source_size = None
                self.base_image = load_preview_or_create(self.record.path)
                if self.source_size is None and self.base_image is not None:
                    self.source_size = self.base_image.size
            self.apply_rotation_to_image()
        except Exception as exc:
            self.base_image = None
            self.image = None
            self.canvas.delete("all")
            self.canvas.create_text(10, 10, anchor="nw", fill="white", text=f"Erreur ouverture image : {exc}")

    def apply_rotation_to_image(self):
        if self.base_image is None:
            return
        rotation = self.current_rotation() % 360
        if rotation == 180:
            self.image = self.base_image.rotate(180, expand=True)
        else:
            self.image = self.base_image.copy()
        self.rotation_label.configure(text=self.rotation_text())

    def toggle_rotation(self):
        cur = self.current_rotation()
        self.rotation_store[self.rotation_key()] = 180 if cur == 0 else 0
        self.apply_rotation_to_image()
        self.group.draw_all()

    def reset_rotation(self):
        self.rotation_store[self.rotation_key()] = 0
        self.apply_rotation_to_image()
        self.group.draw_all()

    def on_configure(self, _event=None):
        self.group.schedule_fit(self)
        if not self.group.auto_fit:
            self.draw()

    def on_mousewheel(self, event):
        factor = 1.15 if event.delta > 0 else 1 / 1.15
        self.group.zoom_at(self, event.x, event.y, factor)

    def on_drag_start(self, event):
        self._drag_start = (event.x, event.y)

    def on_drag_move(self, event):
        if self._drag_start is None:
            self._drag_start = (event.x, event.y)
            return
        old_x, old_y = self._drag_start
        dx = event.x - old_x
        dy = event.y - old_y
        self._drag_start = (event.x, event.y)
        self.group.pan(dx, dy)

    def canvas_to_image(self, x: float, y: float) -> tuple[float, float]:
        ox, oy = self.current_offset()
        if self.group.zoom <= 0:
            return self.group.center_x, self.group.center_y
        return (x - ox) / self.group.zoom, (y - oy) / self.group.zoom

    def current_offset(self) -> tuple[float, float]:
        cw = self.canvas.winfo_width()
        ch = self.canvas.winfo_height()
        ox = cw / 2 - self.group.center_x * self.group.zoom
        oy = ch / 2 - self.group.center_y * self.group.zoom
        return ox, oy

    def toggle_zones(self):
        APP_SETTINGS["show_zone_boundaries"] = not bool(APP_SETTINGS.get("show_zone_boundaries", True))
        save_app_settings(APP_SETTINGS)
        self.group.draw_all()

    def draw_zone_boundaries(self, ox: float, oy: float, zoom: float):
        if not bool(APP_SETTINGS.get("show_zone_boundaries", True)):
            return
        if self.image is None:
            return

        try:
            resolution = float(APP_SETTINGS.get("resolution_um_per_pixel", 5.4))
        except Exception:
            resolution = 5.4
        if resolution <= 0:
            return

        img_w, img_h = self.image.size
        if self.source_size and self.source_size[0] > 0:
            preview_scale = img_w / float(self.source_size[0])
        else:
            preview_scale = 1.0

        cx = ox + (img_w * zoom) / 2.0
        cy = oy + (img_h * zoom) / 2.0
        colors = ["#00ffff", "#ffff00", "#ff9900", "#ff0000", "#00ff00"]

        drawn = 0
        for idx, zone in enumerate(get_zones()):
            max_um = zone.get("max_um")
            if max_um is None:
                continue
            try:
                radius_px_original = float(max_um) / resolution
            except Exception:
                continue
            radius_display = radius_px_original * preview_scale * zoom
            if radius_display <= 1:
                continue

            color = colors[idx % len(colors)]
            self.canvas.create_oval(
                cx - radius_display,
                cy - radius_display,
                cx + radius_display,
                cy + radius_display,
                outline=color,
                width=2,
                dash=(8, 4),
            )
            self.canvas.create_text(
                cx + radius_display + 6,
                cy,
                anchor="w",
                fill=color,
                text=str(zone.get("name", "")),
                font=("Arial", 9, "bold"),
            )
            drawn += 1

        if drawn:
            self.canvas.create_text(
                8,
                26,
                anchor="nw",
                fill="#00ffff",
                text="Zones affichées",
                font=("Arial", 9, "bold"),
            )

    def draw(self):
        if self.image is None:
            return
        w, h = self.image.size
        zoom = max(self.group.zoom, 0.01)
        display_w = max(1, int(w * zoom))
        display_h = max(1, int(h * zoom))

        resized = self.image.resize((display_w, display_h), Image.Resampling.LANCZOS)
        self.tk_image = ImageTk.PhotoImage(resized)
        ox, oy = self.current_offset()

        self.canvas.delete("all")
        self.image_item = self.canvas.create_image(ox, oy, image=self.tk_image, anchor="nw")
        self.draw_zone_boundaries(ox, oy, zoom)
        mode = "full" if self.is_full_resolution else "preview"
        self.canvas.create_text(8, 8, anchor="nw", fill="white", text=f"zoom {zoom:.2f}x | {self.current_rotation()}° | {mode}")



def load_follow_targets() -> list[str]:
    try:
        if FOLLOW_TARGETS_PATH.exists():
            return [line.strip() for line in FOLLOW_TARGETS_PATH.read_text(encoding="utf-8").splitlines() if line.strip()]
    except Exception as exc:
        print(f"Suivi illisible : {exc}")
    return []


def save_follow_targets(targets: list[str]):
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cleaned = []
        seen = set()
        for t in targets:
            t = str(t).strip()
            if t and t not in seen:
                cleaned.append(t)
                seen.add(t)
        FOLLOW_TARGETS_PATH.write_text("\n".join(cleaned), encoding="utf-8")
    except Exception as exc:
        print(f"Impossible d'écrire la liste de suivi : {exc}")


def image_tokens_for_sync(path: Path) -> list[str]:
    out = []
    try:
        out.extend([path.parent.name, path.parent.parent.name, path.parent.parent.parent.name, path.stem])
    except Exception:
        out.append(path.stem)
    # dédoublonnage en gardant l'ordre
    seen = set()
    cleaned = []
    for token in out:
        token = str(token)
        if token and token not in seen:
            cleaned.append(token)
            seen.add(token)
    return cleaned


def synchronize_image_timestamps_with_texts(text_runs: list[TextRun], image_runs: list[ImageRun]):
    """Force les images à utiliser l'horodatage fiable issu du TXT quand c'est possible."""
    by_product: dict[str, list[TextRun]] = {}
    by_stem: dict[str, TextRun] = {}

    for run in text_runs:
        by_product.setdefault(run.product, []).append(run)
        by_stem[run.path.stem] = run

    for runs in by_product.values():
        runs.sort(key=lambda r: r.timestamp or datetime.min)

    for img in image_runs:
        tokens = image_tokens_for_sync(img.path)
        matched = None

        for token in tokens:
            if token in by_stem:
                matched = by_stem[token]
                break

        product_runs = by_product.get(img.product, [])
        if matched is None and len(product_runs) == 1:
            matched = product_runs[0]

        if matched is None and product_runs and img.timestamp is not None:
            matched = min(
                product_runs,
                key=lambda r: abs(((r.timestamp or datetime.min) - img.timestamp).total_seconds())
                if r.timestamp is not None else float("inf")
            )

        if matched is not None and matched.timestamp is not None:
            img.timestamp = matched.timestamp


def merge_unique_runs(text_runs_lists: list[list[TextRun]], image_runs_lists: list[list[ImageRun]]) -> tuple[list[TextRun], list[ImageRun]]:
    text_by_path: dict[str, TextRun] = {}
    image_by_path: dict[str, ImageRun] = {}
    for runs in text_runs_lists:
        for run in runs:
            text_by_path[str(run.path)] = run
    for runs in image_runs_lists:
        for run in runs:
            image_by_path[str(run.path)] = run
    return list(text_by_path.values()), list(image_by_path.values())




def parse_multi_folder_value(value: str) -> list[Path]:
    """Accepte plusieurs dossiers séparés par ; ou retour ligne."""
    value = value or ""
    raw_parts = []
    for chunk in value.replace("\r", "\n").replace(";", "\n").split("\n"):
        part = chunk.strip().strip('"')
        if part:
            raw_parts.append(part)

    out = []
    seen = set()
    for part in raw_parts:
        p = Path(part)
        key = str(p)
        if key not in seen:
            out.append(p)
            seen.add(key)
    return out


def format_multi_folder_value(paths: list[Path]) -> str:
    return "; ".join(str(p) for p in paths)


def add_path_to_multi_value(current: str, new_folder: str) -> str:
    paths = parse_multi_folder_value(current)
    p = Path(new_folder)
    if str(p) not in {str(x) for x in paths}:
        paths.append(p)
    return format_multi_folder_value(paths)


def make_scan_pairs(text_dirs: list[Path], image_dirs: list[Path], txt_only: bool) -> list[tuple[Path, Path | None]]:
    """Crée les couples TXT/Image à scanner.

    - TXT uniquement : un scan par dossier TXT.
    - Même nombre de dossiers TXT et image : pairing 1 à 1, pratique pour comparer deux bancs.
    - Un seul dossier image : il est utilisé avec tous les dossiers TXT.
    - Sinon : toutes les combinaisons.
    """
    if txt_only or not image_dirs:
        return [(td, None) for td in text_dirs]

    if len(text_dirs) == len(image_dirs):
        return list(zip(text_dirs, image_dirs))

    if len(image_dirs) == 1:
        return [(td, image_dirs[0]) for td in text_dirs]

    if len(text_dirs) == 1:
        return [(text_dirs[0], img) for img in image_dirs]

    return [(td, img) for td in text_dirs for img in image_dirs]


def merge_scan_stats(stats_list: list[dict], txt_only: bool = False, follow_optimized: bool = False) -> dict:
    out = {
        "from_cache_only": bool(stats_list and all(s.get("from_cache_only") for s in stats_list)),
        "txt_only": txt_only,
        "follow_optimized": follow_optimized or any(s.get("follow_optimized") for s in stats_list),
        "parsed_txt": 0,
        "parsed_images": 0,
        "cache_hits": 0,
        "visited_txt": 0,
        "visited_images": 0,
        "skipped_txt_outside_date": 0,
        "skipped_images_outside_date": 0,
        "skipped_images_not_in_txt_modules": 0,
        "selected_txt": 0,
        "max_files_per_type": get_max_files_per_type(),
    }
    for st in stats_list:
        for key in (
            "parsed_txt", "parsed_images", "cache_hits", "visited_txt", "visited_images",
            "skipped_txt_outside_date", "skipped_images_outside_date",
            "skipped_images_not_in_txt_modules", "selected_txt",
        ):
            out[key] += int(st.get(key, 0))
    return out


def scan_multi_folders_indexed(text_dirs: list[Path], image_dirs: list[Path],
                               product_filter: str = "",
                               cache_only: bool = False,
                               cancel_event: threading.Event | None = None,
                               progress_callback=None,
                               max_files_per_type: int | None = None,
                               start_dt: datetime | None = None,
                               end_dt: datetime | None = None,
                               txt_only: bool = False) -> tuple[list[TextRun], list[ImageRun], dict]:
    text_lists = []
    image_lists = []
    stats_list = []
    pairs = make_scan_pairs(text_dirs, image_dirs, txt_only)

    for idx_pair, (td, imgd) in enumerate(pairs, start=1):
        if cancel_event and cancel_event.is_set():
            break
        if progress_callback:
            img_txt = "TXT seul" if imgd is None else str(imgd)
            progress_callback(f"Scan dossier {idx_pair}/{len(pairs)} : TXT={td} | Images={img_txt}", idx_pair)
        tr, ir, st = scan_folders_indexed(
            text_dir=td,
            image_dir=imgd,
            product_filter=product_filter,
            cache_only=cache_only,
            cancel_event=cancel_event,
            progress_callback=progress_callback,
            max_files_per_type=max_files_per_type,
            start_dt=start_dt,
            end_dt=end_dt,
            txt_only=txt_only or imgd is None,
        )
        text_lists.append(tr)
        image_lists.append(ir)
        stats_list.append(st)

    text_runs, image_runs = merge_unique_runs(text_lists, image_lists)
    stats = merge_scan_stats(stats_list, txt_only=txt_only, follow_optimized=False)
    stats["multi_text_dirs"] = len(text_dirs)
    stats["multi_image_dirs"] = len(image_dirs)
    return text_runs, image_runs, stats


def scan_follow_targets_multi_indexed(text_dirs: list[Path], image_dirs: list[Path],
                                      follow_targets: list[str],
                                      cache_only: bool = False,
                                      cancel_event: threading.Event | None = None,
                                      progress_callback=None,
                                      max_files_per_type: int | None = None,
                                      start_dt: datetime | None = None,
                                      end_dt: datetime | None = None,
                                      txt_only: bool = False) -> tuple[list[TextRun], list[ImageRun], dict]:
    text_lists = []
    image_lists = []
    stats_list = []
    pairs = make_scan_pairs(text_dirs, image_dirs, txt_only)

    for idx_pair, (td, imgd) in enumerate(pairs, start=1):
        if cancel_event and cancel_event.is_set():
            break
        if progress_callback:
            img_txt = "TXT seul" if imgd is None else str(imgd)
            progress_callback(f"Mode suivi dossier {idx_pair}/{len(pairs)} : TXT={td} | Images={img_txt}", idx_pair)
        tr, ir, st = scan_follow_targets_indexed(
            text_dir=td,
            image_dir=imgd,
            follow_targets=follow_targets,
            cache_only=cache_only,
            cancel_event=cancel_event,
            progress_callback=progress_callback,
            max_files_per_type=max_files_per_type,
            start_dt=start_dt,
            end_dt=end_dt,
            txt_only=txt_only or imgd is None,
        )
        text_lists.append(tr)
        image_lists.append(ir)
        stats_list.append(st)

    text_runs, image_runs = merge_unique_runs(text_lists, image_lists)
    stats = merge_scan_stats(stats_list, txt_only=txt_only, follow_optimized=True)
    stats["multi_text_dirs"] = len(text_dirs)
    stats["multi_image_dirs"] = len(image_dirs)
    stats["follow_count"] = len(follow_targets)
    return text_runs, image_runs, stats



class DefectReportWindow(tk.Toplevel):
    def __init__(self, parent: tk.Widget, ref_run: TextRun, comp_run: TextRun):
        super().__init__(parent)
        self.title(f"Visualisation défauts — {date_label(comp_run.timestamp)} — {comp_run.path.name}")
        self.geometry("1180x850")
        self.minsize(900, 650)

        self.ref_run = ref_run
        self.comp_run = comp_run
        self.visual_scale = 1.0
        self._redraw_after = None

        container = ttk.Frame(self)
        container.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(container, bg="white")
        vbar = ttk.Scrollbar(container, orient="vertical", command=self.canvas.yview)
        hbar = ttk.Scrollbar(container, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=vbar.set, xscrollcommand=hbar.set)

        self.canvas.grid(row=0, column=0, sticky="nsew")
        vbar.grid(row=0, column=1, sticky="ns")
        hbar.grid(row=1, column=0, sticky="ew")
        container.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)

        self.canvas.bind("<Configure>", self.on_resize)
        self.draw_report()

    def on_resize(self, _event=None):
        if self._redraw_after is not None:
            try:
                self.after_cancel(self._redraw_after)
            except Exception:
                pass
        self._redraw_after = self.after(120, self.draw_report)

    def sc(self, value: float) -> float:
        return value * self.visual_scale

    def draw_text(self, x: float, y: float, text: str, *, bold: bool = False, size: int = 10, fill: str = "black", anchor: str = "nw"):
        font_size = max(7, int(round(size * self.visual_scale)))
        font = ("Arial", font_size, "bold" if bold else "normal")
        return self.canvas.create_text(self.sc(x), self.sc(y), text=text, fill=fill, anchor=anchor, font=font)

    def draw_circle(self, center_x: float, center_y: float, radius: float, color: str, width: float):
        self.canvas.create_oval(
            self.sc(center_x - radius),
            self.sc(center_y - radius),
            self.sc(center_x + radius),
            self.sc(center_y + radius),
            outline=color,
            width=max(1, int(round(width * self.visual_scale))),
        )

    def draw_line(self, x1, y1, x2, y2, **kwargs):
        if "width" in kwargs:
            kwargs["width"] = max(1, int(round(kwargs["width"] * self.visual_scale)))
        self.canvas.create_line(self.sc(x1), self.sc(y1), self.sc(x2), self.sc(y2), **kwargs)

    def draw_rect(self, x1, y1, x2, y2, **kwargs):
        self.canvas.create_rectangle(self.sc(x1), self.sc(y1), self.sc(x2), self.sc(y2), **kwargs)

    def draw_oval(self, x1, y1, x2, y2, **kwargs):
        self.canvas.create_oval(self.sc(x1), self.sc(y1), self.sc(x2), self.sc(y2), **kwargs)

    def draw_report(self):
        self.canvas.delete("all")
        canvas_width = max(self.canvas.winfo_width(), 1)
        self.visual_scale = max(1.0, min(canvas_width / 1120.0, 2.4))

        self.draw_text(20, 16, "Visualisation défauts 2D", bold=True, size=16)
        zones_txt = " | ".join(
            f"{z.get('name')}<{float(z.get('max_um'))/1000:g} mm" if z.get("max_um") is not None else str(z.get("name"))
            for z in get_zones()
        )
        sizes_txt = " | ".join(f"{b.get('name')}={float(b.get('min_um')):g}-{float(b.get('max_um')):g} µm" for b in get_size_buckets())
        self.draw_text(20, 42, f"Zones rapport : {zones_txt}", size=9)
        self.draw_text(20, 60, f"Tailles : {sizes_txt}", size=9)

        left_h = self.draw_run_column(self.ref_run, x_left=20, y_top=96, title="Référence")
        right_h = self.draw_run_column(self.comp_run, x_left=560, y_top=96, title="Fichier comparé")

        total_h = max(left_h, right_h) + 130
        self.draw_line(540, 90, 540, total_h, fill="#333333", width=2)
        self.canvas.configure(scrollregion=(0, 0, self.sc(1080), self.sc(total_h + 40)))

    def draw_run_column(self, run: TextRun, x_left: float, y_top: float, title: str) -> float:
        y = y_top
        self.draw_text(x_left, y, title, bold=True, size=13)
        y += 22
        self.draw_text(x_left, y, f"Fichier : {fit_text(run.path.name, 58)}", bold=True, size=9)
        y += 17
        self.draw_text(x_left, y, f"Module : {run.product}", size=9)
        y += 16
        self.draw_text(x_left, y, f"Date : {date_label(run.timestamp)}", size=9)
        y += 26

        sections = parse_geometric_points_by_section(run)
        if not sections:
            self.draw_text(x_left, y, "Aucune section géométrique exploitable trouvée.", size=10, fill="#b00000")
            return y + 40

        ordered_names = []
        for name in ["DARK SPOT", "WHITE SPOT", "GROWING SPOT", "EP OFF", "SURFACE DEFECT", "FUNCTIONAL DEFECT", "ECE", "ECS", "AE", "OPTICAL FEEDBACK", "HALO", "TA", "TB", "TBDelay", "INHO"]:
            for existing in sections.keys():
                if existing.upper() == name.upper() and existing not in ordered_names:
                    ordered_names.append(existing)
        for existing in sections.keys():
            if existing not in ordered_names:
                ordered_names.append(existing)

        for section_name in ordered_names:
            y = self.draw_section_block(section_name, sections[section_name], x_left, y)
            y += 18
        return y

    def draw_section_block(self, section_name: str, pts: list[DefectPoint], x_left: float, y_top: float) -> float:
        block_w = 500
        block_h = 285
        center_x = x_left + 160
        center_y = y_top + 158
        radius_px = 96

        finite_zone_limits = [float(z.get("max_um")) for z in get_zones() if z.get("max_um") is not None]
        base_radius = max(finite_zone_limits + [p.radius for p in pts] + [1.0])
        max_radius = base_radius * 1.05
        scale = radius_px / max(max_radius, 1.0)

        self.draw_rect(x_left, y_top, x_left + block_w, y_top + block_h, outline="#d0d0d0", fill="#fbfbfb")
        self.draw_text(x_left + 8, y_top + 8, f"[{section_name}] — {len(pts)} défaut(s)", bold=True, size=10)

        colors = ["#00b050", "#ffc000", "#ff0000", "#7030a0", "#0070c0", "#c000c0"]
        legend_y = y_top + 58
        previous_um = 0.0
        for idx, zone in enumerate(get_zones()):
            name = str(zone.get("name", ""))
            max_um = zone.get("max_um")
            color = colors[idx % len(colors)]
            if max_um is None:
                label = f"{name} : > {previous_um/1000:g} mm"
                self.draw_circle(center_x, center_y, base_radius * scale, color, 1.5)
            else:
                label = f"{name} : < {float(max_um)/1000:g} mm"
                previous_um = float(max_um)
                self.draw_circle(center_x, center_y, float(max_um) * scale, color, 1.5)
            self.draw_text(x_left + 330, legend_y, label, size=8, fill=color)
            legend_y += 17

        self.draw_line(center_x - radius_px, center_y, center_x + radius_px, center_y, fill="#777777", dash=(4, 3))
        self.draw_line(center_x, center_y - radius_px, center_x, center_y + radius_px, fill="#777777", dash=(4, 3))

        for pt in pts:
            x = center_x + pt.radius * scale * math.cos(pt.angle)
            y = center_y - pt.radius * scale * math.sin(pt.angle)
            diam = max(4.0, min(11.0, 3.0 + pt.size / 25.0))
            fill = "#7030a0" if str(classify_zone(pt.radius)).lower().startswith("hors") else color_by_zone(pt.radius)
            self.draw_oval(x - diam / 2, y - diam / 2, x + diam / 2, y + diam / 2, fill=fill, outline="black", width=1)
            if pt.location:
                self.draw_text(x + 5, y - 5, pt.location, size=7)

        y_legend = y_top + 208
        for line in [p.raw for p in pts[:4]]:
            self.draw_text(x_left + 8, y_legend, fit_text(line, 72), size=7, fill="#333333")
            y_legend += 14
        if len(pts) > 4:
            self.draw_text(x_left + 8, y_legend, f"… +{len(pts) - 4} ligne(s)", size=7, fill="#333333")
        return y_top + block_h



class DatePicker(tk.Toplevel):
    def __init__(self, parent: tk.Widget, target_var: tk.StringVar, initial: datetime | None = None):
        super().__init__(parent)
        self.title("Choisir une date")
        self.resizable(False, False)
        self.target_var = target_var
        self.current = initial or parse_date_entry(target_var.get()) or datetime.today()
        self.year = self.current.year
        self.month = self.current.month
        self.transient(parent)
        self.grab_set()

        self.header = ttk.Frame(self, padding=6)
        self.header.pack(fill="x")
        ttk.Button(self.header, text="<", width=3, command=self.prev_month).pack(side="left")
        self.title_label = ttk.Label(self.header, text="", width=22, anchor="center", font=("Arial", 10, "bold"))
        self.title_label.pack(side="left", padx=6)
        ttk.Button(self.header, text=">", width=3, command=self.next_month).pack(side="left")

        self.days_frame = ttk.Frame(self, padding=6)
        self.days_frame.pack(fill="both", expand=True)
        self.render()

    def prev_month(self):
        self.month -= 1
        if self.month == 0:
            self.month = 12
            self.year -= 1
        self.render()

    def next_month(self):
        self.month += 1
        if self.month == 13:
            self.month = 1
            self.year += 1
        self.render()

    def choose(self, day: int):
        self.target_var.set(f"{self.year:04d}-{self.month:02d}-{day:02d}")
        self.destroy()

    def render(self):
        for child in self.days_frame.winfo_children():
            child.destroy()

        self.title_label.configure(text=f"{calendar.month_name[self.month]} {self.year}")

        for col, name in enumerate(["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]):
            ttk.Label(self.days_frame, text=name, anchor="center", width=5).grid(row=0, column=col, padx=1, pady=1)

        month_cal = calendar.Calendar(firstweekday=0).monthdayscalendar(self.year, self.month)
        for r, week in enumerate(month_cal, start=1):
            for c, day in enumerate(week):
                if day == 0:
                    ttk.Label(self.days_frame, text="", width=5).grid(row=r, column=c, padx=1, pady=1)
                else:
                    ttk.Button(self.days_frame, text=str(day), width=5, command=lambda d=day: self.choose(d)).grid(row=r, column=c, padx=1, pady=1)



class DefectCompareApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Comparateur défauts QIA")
        self.geometry("1550x920")
        self.minsize(1180, 760)

        self.text_dir_var = tk.StringVar()
        self.image_dir_var = tk.StringVar()
        self.interval_var = tk.StringVar(value="Tout")
        self.start_date_var = tk.StringVar()
        self.end_date_var = tk.StringVar()
        self.reference_var = tk.StringVar()
        self.product_filter_var = tk.StringVar()
        self.cache_only_var = tk.BooleanVar(value=False)
        self.follow_mode_var = tk.BooleanVar(value=False)
        self.txt_only_var = tk.BooleanVar(value=False)
        self.status_var = tk.StringVar(value="Prêt.")
        self.follow_targets_text: ScrolledText | None = None
        self._scan_cancel_event: threading.Event | None = None
        self._scan_queue: queue.Queue | None = None
        self._scan_window: tk.Toplevel | None = None
        self._loaded_views: set[tuple] = set()
        self._picture_notebooks: dict[tk.Widget, dict] = {}

        self.text_runs: list[TextRun] = []
        self.image_runs: list[ImageRun] = []
        self.products: list[str] = []
        self.current_product: str | None = None
        self.reference_by_product: dict[str, str] = {}
        self.ref_label_to_path: dict[str, str] = {}
        self.image_rotations: dict[str, int] = {}

        self.report_detail_text: ScrolledText | None = None
        self._report_row_summaries: list[PassageSummary] = []

        self.protocol("WM_DELETE_WINDOW", self.close_app)
        self._build_ui()

    def _build_ui(self):
        top = ttk.Frame(self, padding=8)
        top.pack(fill="x")

        ttk.Button(top, text="Ajouter dossier TXT", command=self.pick_text_folder).grid(row=0, column=0, padx=4, pady=4)
        ttk.Entry(top, textvariable=self.text_dir_var, width=78).grid(row=0, column=1, padx=4, pady=4, sticky="ew")

        ttk.Button(top, text="Ajouter dossier images", command=self.pick_image_folder).grid(row=1, column=0, padx=4, pady=4)
        ttk.Entry(top, textvariable=self.image_dir_var, width=78).grid(row=1, column=1, padx=4, pady=4, sticky="ew")

        date_box = ttk.LabelFrame(top, text="Intervalle de dates", padding=4)
        date_box.grid(row=0, column=2, rowspan=2, padx=8, pady=2, sticky="nsew")
        ttk.Label(date_box, text="Début").grid(row=0, column=0, padx=2)
        ttk.Entry(date_box, textvariable=self.start_date_var, width=12).grid(row=0, column=1, padx=2)
        ttk.Button(date_box, text="📅", width=3, command=lambda: self.open_date_picker(self.start_date_var)).grid(row=0, column=2, padx=2)
        ttk.Label(date_box, text="Fin").grid(row=1, column=0, padx=2)
        ttk.Entry(date_box, textvariable=self.end_date_var, width=12).grid(row=1, column=1, padx=2)
        ttk.Button(date_box, text="📅", width=3, command=lambda: self.open_date_picker(self.end_date_var)).grid(row=1, column=2, padx=2)
        ttk.Button(date_box, text="Appliquer au scan", command=self.scan).grid(row=0, column=3, rowspan=2, padx=6)
        ttk.Button(date_box, text="Tout", command=self.clear_date_range).grid(row=0, column=4, rowspan=2, padx=2)

        ttk.Label(top, text="Module précis :").grid(row=2, column=0, padx=4, pady=4, sticky="e")
        ttk.Entry(top, textvariable=self.product_filter_var, width=22).grid(row=2, column=1, padx=4, pady=4, sticky="w")

        ttk.Checkbutton(
            top,
            text="Utiliser seulement l'index cache local",
            variable=self.cache_only_var,
        ).grid(row=2, column=1, padx=(190, 4), pady=4, sticky="w")

        ttk.Button(top, text="Scanner / Analyser", command=self.scan).grid(row=1, column=3, padx=4)
        ttk.Checkbutton(top, text="Mode suivi", variable=self.follow_mode_var).grid(row=1, column=4, padx=4, sticky="w")
        ttk.Checkbutton(top, text="TXT uniquement", variable=self.txt_only_var).grid(row=2, column=4, padx=4, sticky="w")
        ttk.Button(top, text="Charger module précis", command=self.scan_filtered_product).grid(row=2, column=3, padx=4)

        ttk.Label(top, textvariable=self.status_var).grid(row=3, column=0, columnspan=4, padx=4, pady=(2, 0), sticky="w")

        top.columnconfigure(1, weight=1)

        main = ttk.PanedWindow(self, orient="horizontal")
        main.pack(fill="both", expand=True, padx=8, pady=8)

        left = ttk.Frame(main, padding=6)
        main.add(left, weight=1)
        ttk.Label(left, text="Numéros modules disponibles").pack(anchor="w")
        self.product_list = tk.Listbox(left, height=30, exportselection=False)
        self.product_list.pack(fill="both", expand=True, pady=6)
        self.product_list.bind("<<ListboxSelect>>", self.on_product_select)

        right = ttk.Frame(main, padding=6)
        main.add(right, weight=6)
        self.tabs = ttk.Notebook(right)
        self.tabs.pack(fill="both", expand=True)
        self.tabs.bind("<<NotebookTabChanged>>", self.on_main_tab_changed)

        # Onglets principaux : Rapport, outil d'orientation, Suivi, Paramètres.
        # Comparaison TXT et Images deviennent des sous-onglets accessibles depuis Rapport.
        self.report_main_tab = ttk.Frame(self.tabs)
        self.overlay_tab = ttk.Frame(self.tabs)
        self.follow_tab = ttk.Frame(self.tabs)
        self.settings_tab = ttk.Frame(self.tabs)

        self.tabs.add(self.report_main_tab, text="Rapport")
        self.tabs.add(self.overlay_tab, text="Analyse d’image")
        self.tabs.add(self.follow_tab, text="Suivi")
        self.tabs.add(self.settings_tab, text="Paramètres")

        self.report_subtabs = ttk.Notebook(self.report_main_tab)
        self.report_subtabs.pack(fill="both", expand=True)

        self.report_tab = ttk.Frame(self.report_subtabs)
        self.text_tab = ttk.Frame(self.report_subtabs)
        self.image_tab = ttk.Frame(self.report_subtabs)

        self.report_subtabs.add(self.report_tab, text="Rapport")
        self.report_subtabs.add(self.text_tab, text="Comparaison TXT")
        self.report_subtabs.add(self.image_tab, text="Images")
        self.report_subtabs.bind("<<NotebookTabChanged>>", self.on_report_subtab_changed)

    def pick_text_folder(self):
        folder = filedialog.askdirectory(title="Ajouter un dossier contenant des fichiers TXT")
        if folder:
            self.text_dir_var.set(add_path_to_multi_value(self.text_dir_var.get(), folder))

    def pick_image_folder(self):
        folder = filedialog.askdirectory(title="Ajouter un dossier racine des images")
        if folder:
            self.image_dir_var.set(add_path_to_multi_value(self.image_dir_var.get(), folder))

    def clear_frame(self, frame):
        for child in frame.winfo_children():
            child.destroy()

    def close_app(self):
        """Ferme proprement l'application et les fenêtres secondaires."""
        try:
            if self._scan_cancel_event is not None:
                self._scan_cancel_event.set()
        except Exception:
            pass

        try:
            if self._scan_window is not None and self._scan_window.winfo_exists():
                self._scan_window.destroy()
        except Exception:
            pass

        # Ferme les fenêtres Toplevel ouvertes : date picker, visualisation, scan, etc.
        try:
            for win in list(self.winfo_children()):
                if isinstance(win, tk.Toplevel):
                    try:
                        win.destroy()
                    except Exception:
                        pass
        except Exception:
            pass

        try:
            self.quit()
        except Exception:
            pass

        try:
            self.destroy()
        except Exception:
            pass

    def open_date_picker(self, target_var: tk.StringVar):
        DatePicker(self, target_var)

    def clear_date_range(self):
        self.start_date_var.set("")
        self.end_date_var.set("")
        self.scan()

    def selected_date_range(self) -> tuple[datetime | None, datetime | None]:
        start = parse_date_entry(self.start_date_var.get())
        end = parse_date_entry(self.end_date_var.get())
        if start and end and start > end:
            # On inverse plutôt que de bloquer : pratique si l'utilisateur clique vite.
            start, end = end, start
        return start, end

    def set_date_range_from_loaded_data(self):
        if self.start_date_var.get().strip() or self.end_date_var.get().strip():
            return
        dates = [r.timestamp for r in self.text_runs if r.timestamp is not None]
        if not dates:
            dates = [r.timestamp for r in self.image_runs if r.timestamp is not None]
        if not dates:
            return
        self.start_date_var.set(min(dates).strftime("%Y-%m-%d"))
        self.end_date_var.set(max(dates).strftime("%Y-%m-%d"))


    def scan(self):
        self.start_scan(product_filter="")

    def scan_filtered_product(self):
        product_filter = self.product_filter_var.get().strip()
        if not product_filter:
            messagebox.showwarning("Module précis", "Entre un numéro de module avant de lancer ce chargement.")
            return
        self.start_scan(product_filter=product_filter)

    def start_scan(self, product_filter: str = ""):
        text_dirs = parse_multi_folder_value(self.text_dir_var.get())
        image_dirs = parse_multi_folder_value(self.image_dir_var.get())
        txt_only = bool(self.txt_only_var.get())

        valid_text_dirs = [p for p in text_dirs if p.exists() and p.is_dir()]
        valid_image_dirs = [p for p in image_dirs if p.exists() and p.is_dir()]

        if not valid_text_dirs:
            messagebox.showerror("Erreur", "Choisis au moins un dossier TXT valide.")
            return

        if not txt_only and not valid_image_dirs:
            messagebox.showerror("Erreur", "Choisis au moins un dossier images valide ou coche TXT uniquement.")
            return

        if self._scan_cancel_event is not None:
            messagebox.showinfo("Scan en cours", "Un scan est déjà en cours.")
            return

        self._scan_cancel_event = threading.Event()
        self._scan_queue = queue.Queue()

        self._scan_window = tk.Toplevel(self)
        self._scan_window.title("Scan en cours")
        self._scan_window.geometry("620x140")
        self._scan_window.transient(self)
        self._scan_window.grab_set()

        if txt_only:
            mode_label = f"Scan TXT uniquement — {len(valid_text_dirs)} dossier(s) TXT…"
        else:
            mode_label = f"Scan multi-dossiers — {len(valid_text_dirs)} dossier(s) TXT / {len(valid_image_dirs)} dossier(s) images…"
        ttk.Label(self._scan_window, text=mode_label).pack(anchor="w", padx=12, pady=(12, 4))
        progress_label = ttk.Label(self._scan_window, textvariable=self.status_var)
        progress_label.pack(anchor="w", padx=12, pady=4)
        bar = ttk.Progressbar(self._scan_window, mode="indeterminate")
        bar.pack(fill="x", padx=12, pady=8)
        bar.start(12)
        ttk.Button(self._scan_window, text="Annuler", command=self.cancel_scan).pack(anchor="e", padx=12, pady=4)

        cache_only = self.cache_only_var.get()
        follow_targets = self.get_follow_targets() if self.follow_mode_var.get() and not product_filter else []
        if self.follow_mode_var.get() and not product_filter and not follow_targets:
            self.end_scan_ui()
            messagebox.showwarning("Mode suivi", "Le mode suivi est coché, mais la liste Suivi est vide.")
            return

        def progress(message: str, _count: int = 0):
            if self._scan_queue:
                self._scan_queue.put(("progress", message))

        def worker():
            try:
                start_dt, end_dt = self.selected_date_range()

                if follow_targets:
                    if txt_only:
                        progress(f"Mode suivi TXT uniquement : lecture directe dans {len(valid_text_dirs)} dossier(s) TXT…", 0)
                    else:
                        progress(f"Mode suivi multi-dossiers : {len(valid_text_dirs)} dossier(s) TXT / {len(valid_image_dirs)} dossier(s) images…", 0)

                    text_runs, image_runs, stats = scan_follow_targets_multi_indexed(
                        text_dirs=valid_text_dirs,
                        image_dirs=valid_image_dirs,
                        follow_targets=follow_targets,
                        cache_only=cache_only,
                        cancel_event=self._scan_cancel_event,
                        progress_callback=progress,
                        max_files_per_type=get_max_files_per_type(),
                        start_dt=start_dt,
                        end_dt=end_dt,
                        txt_only=txt_only,
                    )
                else:
                    text_runs, image_runs, stats = scan_multi_folders_indexed(
                        text_dirs=valid_text_dirs,
                        image_dirs=valid_image_dirs,
                        product_filter=product_filter,
                        cache_only=cache_only,
                        cancel_event=self._scan_cancel_event,
                        progress_callback=progress,
                        max_files_per_type=get_max_files_per_type(),
                        start_dt=start_dt,
                        end_dt=end_dt,
                        txt_only=txt_only,
                    )

                if self._scan_cancel_event and self._scan_cancel_event.is_set():
                    self._scan_queue.put(("cancelled", None))
                else:
                    stats["txt_only"] = txt_only
                    stats["multi_text_dirs"] = len(valid_text_dirs)
                    stats["multi_image_dirs"] = len(valid_image_dirs)
                    self._scan_queue.put(("done", text_runs, image_runs, stats))
            except Exception as exc:
                self._scan_queue.put(("error", str(exc)))

        threading.Thread(target=worker, daemon=True).start()
        self.after(120, self.poll_scan_queue)

    def cancel_scan(self):
        if self._scan_cancel_event is not None:
            self._scan_cancel_event.set()
            self.status_var.set("Annulation demandée…")

    def poll_scan_queue(self):
        if self._scan_queue is None:
            return

        try:
            while True:
                item = self._scan_queue.get_nowait()
                kind = item[0]

                if kind == "progress":
                    self.status_var.set(item[1])

                elif kind == "done":
                    _kind, text_runs, image_runs, stats = item
                    self.finish_scan(text_runs, image_runs, stats)
                    return

                elif kind == "cancelled":
                    self.end_scan_ui()
                    self.status_var.set("Scan annulé.")
                    return

                elif kind == "error":
                    self.end_scan_ui()
                    messagebox.showerror("Erreur scan", item[1])
                    self.status_var.set("Erreur pendant le scan.")
                    return

        except queue.Empty:
            pass

        self.after(120, self.poll_scan_queue)

    def end_scan_ui(self):
        if self._scan_window is not None:
            try:
                self._scan_window.grab_release()
                self._scan_window.destroy()
            except Exception:
                pass
        self._scan_window = None
        self._scan_cancel_event = None
        self._scan_queue = None

    def finish_scan(self, text_runs: list[TextRun], image_runs: list[ImageRun], stats: dict):
        self.end_scan_ui()

        synchronize_image_timestamps_with_texts(text_runs, image_runs)

        # Les TXT ont déjà été filtrés par date et limités pendant le scan.
        # Le TXT reste la source de vérité : la liste des modules vient uniquement des TXT sélectionnés.
        raw_txt_count = int(stats.get("visited_txt", len(text_runs)))
        raw_img_count = int(stats.get("visited_images", len(image_runs)))
        start_dt, end_dt = self.selected_date_range()
        text_products = {r.product for r in text_runs}
        ignored_image_only_products = []
        image_runs = [r for r in image_runs if r.product in text_products]
        # Ne pas refiltrer ici par Horodatage/timestamp synchronisé.
        # Le scan a déjà filtré les fichiers par date de création fichier.

        self.text_runs = text_runs
        self.image_runs = image_runs
        self.reference_by_product.clear()
        self._loaded_views.clear()
        self._picture_notebooks.clear()

        self.products = sorted(text_products)
        self.product_list.delete(0, "end")

        for product in self.products:
            txt_count = sum(1 for r in self.text_runs if r.product == product)
            img_count = sum(1 for r in self.image_runs if r.product == product)
            self.product_list.insert("end", f"{product}    |    {txt_count} TXT    |    {img_count} images")

        self.clear_frame(self.report_tab)
        self.clear_frame(self.text_tab)
        self.clear_frame(self.image_tab)
        if hasattr(self, "follow_tab"):
            self.clear_frame(self.follow_tab)

        if self.products:
            self.product_list.selection_clear(0, "end")
            self.product_list.selection_set(0)
            self.product_list.activate(0)
            self.current_product = self.products[0]
            self.select_report_main("Rapport")
            self.load_active_tab()
        else:
            self.current_product = None
            ttk.Label(
                self.report_tab,
                text="Aucun module avec fichier TXT trouvé dans le scan / l'intervalle demandé.",
                foreground="red",
                font=("Arial", 10, "bold"),
            ).pack(anchor="w", padx=8, pady=8)

        mode = "cache seul" if stats.get("from_cache_only") else ("TXT uniquement" if stats.get("txt_only") else ("suivi optimisé" if stats.get("follow_optimized") else "scan serveur"))
        date_info = ""
        if self.date_filter_is_active():
            date_info = f" | filtre date appliqué au scan"
        ignored_info = ""
        if ignored_image_only_products:
            ignored_info = f" | modules avec images sans TXT ignorés: {len(ignored_image_only_products)}"
        self.status_var.set(
            f"Terminé ({mode}) — modules: {len(self.products)} | "
            f"cache: {stats.get('cache_hits', 0)} | "
            f"TXT parsés: {stats.get('parsed_txt', 0)} | "
            f"images indexées: {stats.get('parsed_images', 0)} | "
            f"TXT analysés: {stats.get('visited_txt', 0)} | "
            f"TXT retenus: {stats.get('selected_txt', len(self.text_runs))}/{stats.get('max_files_per_type', get_max_files_per_type())} | "
            f"hors intervalle interne: {stats.get('skipped_txt_outside_date', 0)}"
            f"{date_info}"
            f"{ignored_info}"
        )

        messagebox.showinfo("Scan terminé", f"{len(self.products)} module(s) trouvé(s).")

    def selected_product(self) -> str | None:
        sel = self.product_list.curselection()
        if not sel:
            return None
        return self.products[sel[0]]

    def interval_days(self) -> int | None:
        return None

    def product_text_runs(self, product: str) -> list[TextRun]:
        runs = [r for r in self.text_runs if r.product == product]
        return sorted(runs, key=lambda r: r.timestamp or datetime.min)

    def product_image_runs(self, product: str) -> list[ImageRun]:
        runs = [r for r in self.image_runs if r.product == product]
        return sorted(runs, key=lambda r: (r.picture_name, r.timestamp or datetime.min))

    def date_filter_is_active(self) -> bool:
        return bool(self.start_date_var.get().strip() or self.end_date_var.get().strip())

    def unfiltered_product_text_count(self, product: str) -> int:
        return sum(1 for r in self.text_runs if r.product == product)

    def unfiltered_product_image_count(self, product: str) -> int:
        return sum(1 for r in self.image_runs if r.product == product)

    def active_tab_name(self) -> str:
        try:
            return self.tabs.tab(self.tabs.select(), "text")
        except Exception:
            return "Rapport"

    def active_report_subtab_name(self) -> str:
        try:
            return self.report_subtabs.tab(self.report_subtabs.select(), "text")
        except Exception:
            return "Rapport"

    def current_view_key(self, tab_name: str) -> tuple[str, str, str] | None:
        product = self.current_product or self.selected_product()
        if not product:
            return None
        return (product, self.start_date_var.get(), self.end_date_var.get(), tab_name)

    def mark_views_dirty(self):
        self._loaded_views.clear()
        self._picture_notebooks.clear()
        self.clear_frame(self.report_tab)
        self.clear_frame(self.text_tab)
        self.clear_frame(self.image_tab)

    def select_report_main(self, subtab: str = "Rapport"):
        try:
            self.tabs.select(self.report_main_tab)
        except Exception:
            pass
        try:
            target = {
                "Rapport": self.report_tab,
                "Comparaison TXT": self.text_tab,
                "Images": self.image_tab,
            }.get(subtab, self.report_tab)
            self.report_subtabs.select(target)
        except Exception:
            pass

    def refresh_current_product(self):
        product = self.current_product or self.selected_product()
        if not product:
            return
        self.current_product = product
        self.mark_views_dirty()
        self.select_report_main("Rapport")
        self.load_active_tab()

    def on_main_tab_changed(self, _event=None):
        self.load_active_tab()

    def on_report_subtab_changed(self, _event=None):
        if self.active_tab_name() == "Rapport":
            self.load_active_tab()

    def load_active_tab(self):
        tab_name = self.active_tab_name()

        if tab_name == "Paramètres":
            key = ("settings",)
            if key not in self._loaded_views:
                self.render_settings()
                self._loaded_views.add(key)
            return

        if tab_name == "Suivi":
            key = ("follow",)
            if key not in self._loaded_views:
                self.render_follow_tab()
                self._loaded_views.add(key)
            return

        if tab_name == "Analyse d’image":
            key = ("image_analysis",)
            if key not in self._loaded_views:
                self.render_overlay_tab()
                self._loaded_views.add(key)
            return

        # Onglet principal Rapport : on charge seulement le sous-onglet demandé.
        if tab_name == "Rapport":
            tab_name = self.active_report_subtab_name()

        product = self.current_product or self.selected_product()
        if not product:
            return

        key = self.current_view_key(tab_name)
        if key in self._loaded_views:
            return

        if tab_name == "Rapport":
            self.render_report(product)
        elif tab_name == "Comparaison TXT":
            self.render_text_compare(product)
        elif tab_name == "Images":
            self.render_images(product)

        if key:
            self._loaded_views.add(key)

    def on_product_select(self, _event=None):
        product = self.selected_product()
        if not product:
            return
        self.current_product = product
        self.mark_views_dirty()
        # On revient toujours sur le rapport pour éviter de charger immédiatement les images.
        self.select_report_main("Rapport")
        self.load_active_tab()

    def render_report(self, product: str):
        self.clear_frame(self.report_tab)
        runs = self.product_text_runs(product)

        if not runs:
            total_unfiltered = self.unfiltered_product_text_count(product)
            ttk.Label(self.report_tab, text="Aucun fichier TXT disponible pour ce module.").pack(anchor="w", padx=8, pady=8)
            return

        summaries = build_product_summaries(runs)
        self._report_row_summaries = summaries

        actions = ttk.Frame(self.report_tab, padding=6)
        actions.pack(fill="x")
        ttk.Button(actions, text="Exporter tout le rapport Excel", command=self.export_all_reports_excel).pack(side="right")

        info = ttk.Frame(self.report_tab, padding=6)
        info.pack(fill="x")
        ttk.Label(info, text=f"Module {product} — {len(runs)} passage(s) analysé(s)", font=("Arial", 11, "bold")).pack(side="left")

        zones_txt = " | ".join(
            f"{z.get('name')}<{float(z.get('max_um'))/1000:g} mm" if z.get("max_um") is not None else str(z.get("name"))
            for z in get_zones()
        )
        sizes_txt = " | ".join(f"{b.get('name')}={float(b.get('min_um')):g}-{float(b.get('max_um')):g} µm" for b in get_size_buckets())
        ttk.Label(info, text=f"Zones : {zones_txt}    Tailles : {sizes_txt}").pack(side="left", padx=16)

        main_pane = ttk.PanedWindow(self.report_tab, orient="vertical")
        main_pane.pack(fill="both", expand=True, padx=6, pady=6)

        top_frame = ttk.Frame(main_pane)
        main_pane.add(top_frame, weight=3)
        bottom_frame = ttk.Frame(main_pane)
        main_pane.add(bottom_frame, weight=2)

        canvas = tk.Canvas(top_frame, bg="white", highlightthickness=0)
        vbar = ttk.Scrollbar(top_frame, orient="vertical", command=canvas.yview)
        hbar = ttk.Scrollbar(top_frame, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=vbar.set, xscrollcommand=hbar.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vbar.grid(row=0, column=1, sticky="ns")
        hbar.grid(row=1, column=0, sticky="ew")
        top_frame.rowconfigure(0, weight=1)
        top_frame.columnconfigure(0, weight=1)

        table_frame = tk.Frame(canvas, bg="white")
        canvas.create_window((0, 0), window=table_frame, anchor="nw")
        table_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        grade_cols = get_grade_columns()
        columns = [
            ("DateMaj", 16),
            ("Fichier", 18),
            ("Banc", 12),
        ]
        columns += [(g.get("label", ""), max(8, min(14, len(str(g.get("label", ""))) + 1))) for g in grade_cols]

        for zone in get_zones():
            for bucket in get_size_buckets():
                header = compact_column_label(str(zone.get("name", "")), str(bucket.get("name", "")))
                columns.append((header, 6 if len(header) <= 6 else 8))
        columns += [
            ("Total", 6),
            (">300", 6),
            ("Réc.", 6),
            ("Loc.", 18),
        ]

        for c_idx, (name, width) in enumerate(columns):
            lbl = tk.Label(table_frame, text=name, bg="#fff200", fg="black", relief="solid", bd=1,
                           width=width, font=("Arial", 8, "bold"), anchor="center")
            lbl.grid(row=0, column=c_idx, sticky="nsew")

        for r_idx, summary in enumerate(summaries, start=1):
            row_cells = self.build_report_row_cells(summary)
            for c_idx, cell in enumerate(row_cells):
                lbl = tk.Label(
                    table_frame,
                    text=cell.text,
                    bg=cell.bg,
                    fg=cell.fg,
                    relief="solid",
                    bd=1,
                    width=cell.width,
                    font=("Arial", 8, "bold" if cell.bold else "normal"),
                    anchor=cell.anchor,
                )
                lbl.grid(row=r_idx, column=c_idx, sticky="nsew")
                lbl.bind("<Button-1>", lambda _e, idx=r_idx - 1: self.show_report_summary_details(idx))

        if summaries:
            presence_row = len(summaries) + 1
            presence = self.compute_presence_percentages(summaries)
            presence_cells = [
                CellLabel("Présence %", bg="#e7e6e6", bold=True, width=16),
                CellLabel("", bg="#e7e6e6", width=18),
                CellLabel("", bg="#e7e6e6", width=12),
            ]
            for g in grade_cols:
                presence_cells.append(CellLabel("", bg="#e7e6e6", width=max(8, min(14, len(str(g.get("label", ""))) + 1))))
            for zone in get_zones():
                for bucket in get_size_buckets():
                    col_name = compact_column_label(str(zone.get("name", "")), str(bucket.get("name", "")))
                    pct = presence.get(col_name, 0.0)
                    presence_cells.append(CellLabel(f"{pct:.0f}%", bg=percent_color(pct), width=6 if len(col_name) <= 6 else 8))
            presence_cells += [
                CellLabel("", bg="#e7e6e6", width=6),
                CellLabel("", bg="#e7e6e6", width=6),
                CellLabel(f"{presence.get('Nb récurrents', 0.0):.0f}%", bg=percent_color(presence.get('Nb récurrents', 0.0)), width=6),
                CellLabel("", bg="#e7e6e6", width=18),
            ]
            for c_idx, cell in enumerate(presence_cells):
                tk.Label(
                    table_frame,
                    text=cell.text,
                    bg=cell.bg,
                    fg=cell.fg,
                    relief="solid",
                    bd=1,
                    width=cell.width,
                    font=("Arial", 8, "bold"),
                    anchor=cell.anchor,
                ).grid(row=presence_row, column=c_idx, sticky="nsew")

        ttk.Label(bottom_frame, text="Détails du passage sélectionné", font=("Arial", 10, "bold")).pack(anchor="w", padx=4, pady=(4, 2))
        self.report_detail_text = ScrolledText(bottom_frame, wrap="word", font=("Consolas", 10), height=14)
        self.report_detail_text.pack(fill="both", expand=True, padx=4, pady=4)
        self.show_report_summary_details(0)

    def build_report_row_cells(self, summary: PassageSummary) -> list[CellLabel]:
        loc_text = format_counter(summary.location_counter, empty_text="")
        banc_name = self.infer_banc_name_from_run(summary.run)
        cells = [
            CellLabel(date_label(summary.run.timestamp), bg="white", anchor="w", width=16),
            CellLabel(summary.run.path.name, bg="white", anchor="w", width=18),
            CellLabel(banc_name, bg="white", anchor="w", width=12),
        ]

        for g in get_grade_columns():
            label = str(g.get("label", ""))
            val = summary.grades.get(label, "")
            cells.append(CellLabel(val or "", bg=grade_cell_color(val), width=max(8, min(14, len(label) + 1))))

        for zone in get_zones():
            for bucket in get_size_buckets():
                name = compact_column_label(str(zone.get("name", "")), str(bucket.get("name", "")))
                count = summary.counts.get(name, 0)
                cells.append(CellLabel(str(count), bg=count_cell_color(count), width=6 if len(name) <= 6 else 8))

        cells += [
            CellLabel(str(summary.total_points), bg=count_cell_color(summary.total_points), width=6),
            CellLabel(str(summary.over_300_count), bg=count_cell_color(summary.over_300_count), width=6),
            CellLabel(str(summary.recurrent_count), bg=count_cell_color(summary.recurrent_count), width=6),
            CellLabel(loc_text, bg="white", anchor="w", width=18),
        ]
        return cells

    def compute_presence_percentages(self, summaries: list[PassageSummary]) -> dict[str, float]:
        result: dict[str, float] = {}
        total = max(len(summaries), 1)
        for zone in get_zones():
            for bucket in get_size_buckets():
                col_name = compact_column_label(str(zone.get("name", "")), str(bucket.get("name", "")))
                count_present = sum(1 for s in summaries if s.counts.get(col_name, 0) > 0)
                result[col_name] = 100.0 * count_present / total
        recur_present = sum(1 for s in summaries if s.recurrent_count > 0)
        result["Nb récurrents"] = 100.0 * recur_present / total
        return result

    def infer_banc_name_from_run(self, run: TextRun) -> str:
        """Retourne le banc à afficher pour un passage.

        Priorité :
        1. Champ TXT [CONDITIONS] : ID Machine="QIA2" ou ID Machine=QIA2
        2. Fallback chemin image : .../NomDuBanc/QIA/...
        3. Fallback chemin image : parent/dossier
        """
        for line in run.sections.get("CONDITIONS", []):
            if line.strip().lower().startswith("id machine="):
                value = line.split("=", 1)[1].strip().strip('"').strip("'")
                if value:
                    return value

        return self.infer_banc_name(run.product)

    def infer_banc_name(self, product: str) -> str:
        runs = self.product_image_runs(product)
        counts = Counter()

        for run in runs:
            try:
                parts = list(run.path.parts)
                found = ""

                # Cas attendu : ...\NomDuBanc\QIA\...
                # On affiche donc "NomDuBanc\QIA".
                for i, part in enumerate(parts):
                    if str(part).upper() == "QIA" and i > 0:
                        found = f"{parts[i - 1]}\\{parts[i]}"
                        break

                # Fallback : dossier contenant l'image + son parent.
                # Exemple : ...\Parent\Dossier\image.tif -> "Parent\Dossier".
                if not found:
                    parent = run.path.parent
                    grand_parent = parent.parent
                    if parent.name and grand_parent.name and parent != grand_parent:
                        found = f"{grand_parent.name}\\{parent.name}"
                    elif parent.name:
                        found = parent.name

                if found:
                    counts[str(found)] += 1
            except Exception:
                pass

        return counts.most_common(1)[0][0] if counts else ""

    def report_headers_for_export(self) -> tuple[list[str], list[str]]:
        headers = ["DateMaj", "Fichier", "Banc"] + [str(g.get("label", "")) for g in get_grade_columns()]
        defect_headers = []
        for zone in get_zones():
            for bucket in get_size_buckets():
                defect_headers.append(compact_column_label(str(zone.get("name", "")), str(bucket.get("name", ""))))
        headers += defect_headers + ["Total", ">300", "Réc.", "Loc."]
        return headers, defect_headers

    def excel_fill_for_value(self, value):
        if PatternFill is None:
            return None
        try:
            val = float(str(value).replace("%", "").replace(",", "."))
        except Exception:
            val = 0.0
        if val <= 0:
            return PatternFill("solid", fgColor="FFFFFF")
        if val < 50:
            return PatternFill("solid", fgColor="FFC7CE")
        if val < 80:
            return PatternFill("solid", fgColor="FFEB9C")
        return PatternFill("solid", fgColor="C6EFCE")

    def export_all_reports_excel(self):
        if Workbook is None:
            messagebox.showerror("Export Excel", "openpyxl n'est pas installé. Remplace requirements.txt par la nouvelle version puis reconstruis l'exécutable.")
            return
        if not getattr(self, "products", None):
            messagebox.showwarning("Export Excel", "Aucun module à exporter.")
            return

        out_path = filedialog.asksaveasfilename(
            title="Exporter tout le rapport Excel",
            defaultextension=".xlsx",
            filetypes=[("Classeur Excel", "*.xlsx")]
        )
        if not out_path:
            return

        headers, defect_headers = self.report_headers_for_export()

        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport"

        fill_header = PatternFill("solid", fgColor="FFF200")
        fill_title = PatternFill("solid", fgColor="D9EAF7")
        fill_white = PatternFill("solid", fgColor="FFFFFF")
        fill_grey = PatternFill("solid", fgColor="E7E6E6")
        font_bold = Font(bold=True)
        font_title = Font(bold=True, size=12)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        module_summaries: dict[str, list[PassageSummary]] = {}
        total_modules = 0
        for product in self.products:
            runs = self.product_text_runs(product)
            if not runs:
                continue
            summaries = build_product_summaries(runs)
            if not summaries:
                continue
            module_summaries[product] = summaries
            total_modules += 1

        # Les références des cellules de présence par défaut seront collectées pendant l'écriture
        # des tableaux modules. La synthèse en haut est remplie à la fin avec des formules.
        presence_cell_refs: dict[str, list[str]] = {d: [] for d in defect_headers}
        summary_count_cells: dict[str, object] = {}
        summary_avg_cells: dict[str, object] = {}

        row = 1
        ws.cell(row=row, column=1, value="Synthèse de répétabilité par défaut")
        ws.cell(row=row, column=1).font = font_title
        ws.cell(row=row, column=1).fill = fill_title
        row += 1

        summary_headers = ["Indicateur", "Nombre modules total"] + defect_headers
        for col, header in enumerate(summary_headers, start=1):
            c = ws.cell(row=row, column=col, value=header)
            c.fill = fill_header
            c.font = font_bold
            c.alignment = align_center
            c.border = border
        row += 1

        ws.cell(row=row, column=1, value="Modules concernés")
        ws.cell(row=row, column=2, value=total_modules)
        ws.cell(row=row, column=1).font = font_bold
        ws.cell(row=row, column=1).fill = fill_title
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).alignment = align_center
        ws.cell(row=row, column=2).border = border
        for col, defect in enumerate(defect_headers, start=3):
            c = ws.cell(row=row, column=col, value="")
            c.alignment = align_center
            c.border = border
            c.fill = fill_white
            summary_count_cells[defect] = c
        row += 1

        ws.cell(row=row, column=1, value="Répétabilité moyenne (hors 0%)")
        ws.cell(row=row, column=2, value=total_modules)
        ws.cell(row=row, column=1).font = font_bold
        ws.cell(row=row, column=1).fill = fill_title
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).alignment = align_center
        ws.cell(row=row, column=2).border = border
        for col, defect in enumerate(defect_headers, start=3):
            c = ws.cell(row=row, column=col, value="")
            c.alignment = align_center
            c.border = border
            c.fill = fill_white
            summary_avg_cells[defect] = c
        row += 3

        # Tableaux par module, même logique que le tableau du soft + ligne de présence.
        for product in self.products:
            summaries = module_summaries.get(product)
            if not summaries:
                continue

            ws.cell(row=row, column=1, value=f"Module {product} — {len(summaries)} passage(s)")
            ws.cell(row=row, column=1).font = font_title
            ws.cell(row=row, column=1).fill = fill_title
            row += 1

            for col, header in enumerate(headers, start=1):
                c = ws.cell(row=row, column=col, value=header)
                c.fill = fill_header
                c.font = font_bold
                c.alignment = align_center
                c.border = border
            row += 1

            for summary in summaries:
                cells = self.build_report_row_cells(summary)
                for col, cell in enumerate(cells, start=1):
                    c = ws.cell(row=row, column=col, value=cell.text)
                    c.border = border
                    c.alignment = align_left if cell.anchor == "w" else align_center
                    bg = (cell.bg or "").replace("#", "").upper()
                    if bg:
                        try:
                            c.fill = PatternFill("solid", fgColor=bg)
                        except Exception:
                            c.fill = fill_white
                    else:
                        c.fill = fill_white
                row += 1

            presence = self.compute_presence_percentages(summaries)
            ws.cell(row=row, column=1, value="Présence %")
            ws.cell(row=row, column=1).font = font_bold
            ws.cell(row=row, column=1).fill = fill_grey
            ws.cell(row=row, column=1).border = border

            # Colonne de départ des défauts = DateMaj/Fichier/Banc + grades + 1.
            start_col = 3 + len(get_grade_columns()) + 1
            for idx, defect in enumerate(defect_headers, start=start_col):
                pct = presence.get(defect, 0.0)
                c = ws.cell(row=row, column=idx)
                if pct > 0:
                    # Valeur numérique Excel, formatée en pourcentage.
                    # Exemple : 33% est stocké 0.33, pas comme texte "33%".
                    c.value = pct / 100.0
                    c.number_format = "0%"
                    c.fill = self.excel_fill_for_value(pct)
                    # On collecte uniquement les cellules non nulles, pour que la formule montre
                    # clairement quelles cellules ont été prises en compte.
                    presence_cell_refs[defect].append(c.coordinate)
                else:
                    c.value = ""
                    c.fill = fill_white
                c.alignment = align_center
                c.border = border
            row += 3

        # Remplissage final de la synthèse avec des formules Excel.
        for defect in defect_headers:
            refs = presence_cell_refs.get(defect, [])
            count_cell = summary_count_cells[defect]
            avg_cell = summary_avg_cells[defect]

            if refs:
                count_cell.value = len(refs)
                avg_cell.value = f"=AVERAGE({';'.join(refs)})"
                avg_cell.number_format = "0%"
                # Couleur sur la base de la valeur calculée côté Python pour avoir un code couleur visible immédiatement.
                # La formule reste bien présente dans la cellule Excel.
                vals = []
                for ref in refs:
                    val = ws[ref].value
                    try:
                        fval = float(str(val).replace("%", "").replace(",", "."))
                        # Les cellules de présence sont stockées comme fraction Excel : 0.33 = 33%.
                        if 0.0 <= fval <= 1.0:
                            fval *= 100.0
                        vals.append(fval)
                    except Exception:
                        pass
                avg = sum(vals) / max(len(vals), 1)
                avg_cell.fill = self.excel_fill_for_value(avg)
            else:
                count_cell.value = 0
                avg_cell.value = ""
                avg_cell.fill = fill_white

            count_cell.alignment = align_center
            count_cell.border = border
            avg_cell.alignment = align_center
            avg_cell.border = border

        ws.freeze_panes = "A2"
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len = 0
            for c in col_cells:
                if c.value is not None:
                    max_len = max(max_len, len(str(c.value)))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 38)

        try:
            wb.save(out_path)
            messagebox.showinfo("Export Excel", f"Rapport exporté :\n{out_path}")
        except Exception as exc:
            messagebox.showerror("Export Excel", f"Impossible d'exporter :\n{exc}")

    def show_report_summary_details(self, index: int):
        if self.report_detail_text is None:
            return
        if index < 0 or index >= len(self._report_row_summaries):
            return
        summary = self._report_row_summaries[index]

        lines = []
        lines.append(f"Module : {summary.run.product}")
        lines.append(f"Fichier : {summary.run.path.name}")
        lines.append(f"Date : {date_label(summary.run.timestamp)}")
        lines.append("")
        for g in get_grade_columns():
            label = str(g.get("label", ""))
            val = summary.grades.get(label, "N/A") or "N/A"
            lines.append(f"{label:<18}: {val}")
        lines.append(f"Total défauts     : {summary.total_points}")
        lines.append(f"Défauts > 300 µm  : {summary.over_300_count}")
        lines.append(f"Nb récurrents     : {summary.recurrent_count}")
        lines.append("")
        lines.append("Localisations observées :")
        lines.append(format_counter(summary.location_counter, empty_text="Aucune localisation renseignée") or "Aucune localisation renseignée")
        lines.append("")
        lines.append("Sections observées :")
        lines.append(format_counter(summary.section_counter, empty_text="Aucune section géométrique") or "Aucune section géométrique")
        lines.append("")
        lines.append("Comptage par zone / taille :")
        for zone in get_zones():
            zone_name = str(zone.get("name", ""))
            row = "  " + zone_name + "  " + " | ".join(
                f"{bucket.get('name')}:{summary.counts.get(compact_column_label(zone_name, str(bucket.get('name', ''))), 0)}"
                for bucket in get_size_buckets()
            )
            lines.append(row)
        lines.append("")
        if summary.recurrent_points:
            lines.append("Défauts récurrents détectés (observation) :")
            for pt in summary.recurrent_points[:50]:
                lines.append(f"- [{pt.section}] {pt.raw}")
            if len(summary.recurrent_points) > 50:
                lines.append(f"… +{len(summary.recurrent_points) - 50} autre(s)")
        else:
            lines.append("Aucun défaut récurrent détecté dans cet intervalle.")
        lines.append("")
        lines.append("Rappel des paramètres actifs :")
        for zone in get_zones():
            if zone.get("max_um") is None:
                lines.append(f"- {zone.get('name')} : au-delà de la dernière zone")
            else:
                lines.append(f"- {zone.get('name')} : rayon < {float(zone.get('max_um'))/1000:g} mm")
        for bucket in get_size_buckets():
            lines.append(f"- {bucket.get('name')} : {float(bucket.get('min_um')):g} à {float(bucket.get('max_um')):g} µm")

        self.report_detail_text.configure(state="normal")
        self.report_detail_text.delete("1.0", "end")
        self.report_detail_text.insert("1.0", "\n".join(lines))
        self.report_detail_text.configure(state="disabled")



    # ------------------------------------------------------------------
    # Outil de superposition optique / masque hexagonal
    # ------------------------------------------------------------------
    def render_overlay_tab(self):
        self.clear_frame(self.overlay_tab)

        if Image is None or ImageTk is None:
            ttk.Label(self.overlay_tab, text="Pillow n'est pas installé : impossible d'utiliser l'outil d'analyse d'image.").pack(padx=10, pady=10)
            return

        self.overlay_image_path_var = getattr(self, "overlay_image_path_var", tk.StringVar())
        self.overlay_angle_var = getattr(self, "overlay_angle_var", tk.DoubleVar(value=0.0))
        self.overlay_scale_var = getattr(self, "overlay_scale_var", tk.DoubleVar(value=1.0))
        self.overlay_dx_var = getattr(self, "overlay_dx_var", tk.DoubleVar(value=0.0))
        self.overlay_dy_var = getattr(self, "overlay_dy_var", tk.DoubleVar(value=0.0))
        self.overlay_opacity_var = getattr(self, "overlay_opacity_var", tk.DoubleVar(value=1.0))
        self.overlay_mask_mode_var = getattr(self, "overlay_mask_mode_var", tk.StringVar(value="Bleu uniquement"))
        self.overlay_roi_enabled_var = getattr(self, "overlay_roi_enabled_var", tk.BooleanVar(value=True))
        self.overlay_roi_radius_var = getattr(self, "overlay_roi_radius_var", tk.DoubleVar(value=0.43))
        self.overlay_line_thin_var = getattr(self, "overlay_line_thin_var", tk.DoubleVar(value=0.0))
        self.overlay_detect_sensitivity_var = getattr(self, "overlay_detect_sensitivity_var", tk.DoubleVar(value=1.0))
        self.overlay_detect_percentile_var = getattr(self, "overlay_detect_percentile_var", tk.DoubleVar(value=90.0))
        self.overlay_detect_blur_var = getattr(self, "overlay_detect_blur_var", tk.DoubleVar(value=9.0))
        self.overlay_pitch_min_var = getattr(self, "overlay_pitch_min_var", tk.DoubleVar(value=18.0))
        self.overlay_pitch_max_var = getattr(self, "overlay_pitch_max_var", tk.DoubleVar(value=120.0))
        self.overlay_view_zoom_var = getattr(self, "overlay_view_zoom_var", tk.DoubleVar(value=1.0))
        self.overlay_status_var = getattr(self, "overlay_status_var", tk.StringVar(value="Charge une ou plusieurs images optiques."))
        self.overlay_selected_index_var = getattr(self, "overlay_selected_index_var", tk.StringVar(value=""))
        self.overlay_sticker_mode_var = getattr(self, "overlay_sticker_mode_var", tk.StringVar(value="none"))
        self.overlay_sticker_color_var = getattr(self, "overlay_sticker_color_var", tk.StringVar(value="#ff0000"))

        self.overlay_layers = getattr(self, "overlay_layers", [])
        self.overlay_selected_index = getattr(self, "overlay_selected_index", 0)
        self.overlay_stickers = getattr(self, "overlay_stickers", [])
        self.overlay_viewer = getattr(self, "overlay_viewer", None)
        self.overlay_canvas = getattr(self, "overlay_canvas", None)
        self.overlay_photo = None
        self.overlay_detection_photo = None
        self.overlay_display_scale = 1.0
        self.overlay_drag_last = None
        self.overlay_base_image = getattr(self, "overlay_base_image", None)
        self.overlay_base_preview_cache = None
        self.overlay_base_preview_key = None
        self.overlay_mask_alpha_cache = getattr(self, "overlay_mask_alpha_cache", None)
        self.overlay_mask_alpha_key = getattr(self, "overlay_mask_alpha_key", None)

        root = ttk.Frame(self.overlay_tab, padding=6)
        root.pack(fill="both", expand=True)

        controls = ttk.LabelFrame(root, text="Analyse d'image / détection de grillage", padding=6)
        controls.pack(fill="x")

        ttk.Button(controls, text="Charger image", command=self.overlay_pick_image).grid(row=0, column=0, padx=3, pady=3)
        ttk.Button(controls, text="Ajouter image", command=self.overlay_add_image).grid(row=0, column=1, padx=3, pady=3)
        ttk.Button(controls, text="Supprimer image", command=self.overlay_remove_selected_image).grid(row=0, column=2, padx=3, pady=3)
        ttk.Label(controls, text="Image :").grid(row=0, column=3, padx=(12,3), pady=3, sticky="e")
        self.overlay_layer_selector = ttk.Combobox(controls, textvariable=self.overlay_selected_index_var, state="readonly", width=32)
        self.overlay_layer_selector.grid(row=0, column=4, padx=3, pady=3, sticky="w")
        self.overlay_layer_selector.bind("<<ComboboxSelected>>", self.overlay_on_layer_selected)
        ttk.Button(controls, text="Ouvrir fenêtre image", command=self.overlay_open_viewer).grid(row=0, column=5, padx=3, pady=3)
        controls.columnconfigure(6, weight=1)

        ttk.Label(controls, textvariable=self.overlay_image_path_var).grid(row=1, column=0, columnspan=7, sticky="w", padx=3, pady=(0,3))

        actions = ttk.Frame(root)
        actions.pack(fill="x", pady=(3, 2))
        ttk.Button(actions, text="Détection grillage auto", command=self.overlay_detect_grid).pack(side="left", padx=3)
        ttk.Button(actions, text="Symétrie H", command=self.overlay_toggle_symmetry_h).pack(side="left", padx=3)
        ttk.Button(actions, text="Symétrie V", command=self.overlay_toggle_symmetry_v).pack(side="left", padx=3)
        ttk.Button(actions, text="Exporter PNG", command=self.overlay_export).pack(side="left", padx=3)
        ttk.Button(actions, text="Copier paramètres", command=self.overlay_copy_params).pack(side="left", padx=3)
        ttk.Button(actions, text="Zoom initial", command=self.overlay_reset_view_zoom).pack(side="left", padx=3)
        ttk.Button(actions, text="Reset image", command=self.overlay_reset).pack(side="left", padx=3)

        sticker_frame = ttk.LabelFrame(root, text="Stickers", padding=6)
        sticker_frame.pack(fill="x", pady=(2, 2))
        ttk.Button(sticker_frame, text="Croix", command=lambda: self.overlay_set_sticker_mode("cross")).pack(side="left", padx=3)
        ttk.Button(sticker_frame, text="Rond", command=lambda: self.overlay_set_sticker_mode("circle")).pack(side="left", padx=3)
        ttk.Button(sticker_frame, text="Couleur", command=self.overlay_choose_sticker_color).pack(side="left", padx=3)
        ttk.Button(sticker_frame, text="Effacer stickers", command=self.overlay_clear_stickers).pack(side="left", padx=3)
        ttk.Label(sticker_frame, text="Double-clic dans la fenêtre image pour poser le sticker.").pack(side="left", padx=10)

        sliders = ttk.Frame(root)
        sliders.pack(fill="x", pady=(2, 1))
        self.overlay_make_slider(sliders, "Rotation °", self.overlay_angle_var, -180.0, 180.0, 0, 0, resolution=0.2)
        self.overlay_make_slider(sliders, "Échelle", self.overlay_scale_var, 0.05, 8.0, 0, 3, resolution=0.01)
        self.overlay_make_slider(sliders, "Décalage X", self.overlay_dx_var, -5000.0, 5000.0, 1, 0, resolution=1.0)
        self.overlay_make_slider(sliders, "Décalage Y", self.overlay_dy_var, -5000.0, 5000.0, 1, 3, resolution=1.0)
        self.overlay_make_slider(sliders, "Opacité image", self.overlay_opacity_var, 0.05, 1.0, 2, 0, resolution=0.01)
        self.overlay_make_slider(sliders, "Sensibilité", self.overlay_detect_sensitivity_var, 0.25, 4.0, 2, 3, resolution=0.05)
        self.overlay_make_slider(sliders, "Seuil %", self.overlay_detect_percentile_var, 70.0, 99.5, 3, 0, resolution=0.5)
        self.overlay_make_slider(sliders, "Lissage", self.overlay_detect_blur_var, 3.0, 35.0, 3, 3, resolution=1.0)
        self.overlay_make_slider(sliders, "Pas min px", self.overlay_pitch_min_var, 5.0, 120.0, 4, 0, resolution=1.0)
        self.overlay_make_slider(sliders, "Pas max px", self.overlay_pitch_max_var, 10.0, 250.0, 4, 3, resolution=1.0)

        ttk.Label(root, text="La vue image est affichée dans une fenêtre dédiée. Les curseurs modifient uniquement l'image sélectionnée.").pack(anchor="w", pady=(0, 2))
        ttk.Label(root, textvariable=self.overlay_status_var).pack(anchor="w", pady=(0, 2))

        self.overlay_refresh_layer_selector()
        self.overlay_load_layer_to_vars()
        self.overlay_open_viewer()
        self.overlay_redraw()

    def overlay_make_slider(self, parent, label: str, variable: tk.DoubleVar, vmin: float, vmax: float, row: int, col: int = 0, resolution: float = 1.0):
        ttk.Label(parent, text=label, width=12).grid(row=row, column=col, padx=2, pady=1, sticky="w")
        scale = tk.Scale(
            parent,
            from_=vmin,
            to=vmax,
            orient="horizontal",
            resolution=resolution,
            variable=variable,
            showvalue=False,
            length=240,
            command=lambda _v: self.overlay_on_transform_changed(),
        )
        scale.grid(row=row, column=col + 1, padx=2, pady=1, sticky="ew")
        entry = ttk.Entry(parent, width=8)
        entry.grid(row=row, column=col + 2, padx=2, pady=1)

        def refresh_entry(*_args):
            try:
                entry.delete(0, "end")
                entry.insert(0, f"{float(variable.get()):.6g}")
            except Exception:
                pass

        variable.trace_add("write", lambda *_a: refresh_entry())
        refresh_entry()

        def commit(_event=None):
            try:
                variable.set(float(entry.get().replace(",", ".")))
            except Exception:
                refresh_entry()
            self.overlay_on_transform_changed()

        entry.bind("<Return>", commit)
        entry.bind("<FocusOut>", commit)
        parent.grid_columnconfigure(col + 1, weight=1)

    def overlay_on_transform_changed(self):
        self.overlay_sync_current_layer_from_vars()
        self.overlay_redraw()

    def overlay_refresh_layer_selector(self):
        if not hasattr(self, "overlay_layer_selector"):
            return
        values = [f"Image {i+1} — {Path(layer.get('path','')).name}" for i, layer in enumerate(getattr(self, "overlay_layers", []))]
        self.overlay_layer_selector["values"] = values
        if values:
            idx = max(0, min(getattr(self, "overlay_selected_index", 0), len(values) - 1))
            self.overlay_selected_index = idx
            self.overlay_selected_index_var.set(values[idx])
        else:
            self.overlay_selected_index = 0
            self.overlay_selected_index_var.set("")

    def overlay_get_selected_layer(self):
        layers = getattr(self, "overlay_layers", [])
        idx = getattr(self, "overlay_selected_index", 0)
        if 0 <= idx < len(layers):
            return layers[idx]
        return None

    def overlay_sync_current_layer_from_vars(self):
        layer = self.overlay_get_selected_layer()
        if layer is None:
            self.overlay_base_image = None
            self.overlay_detected_grid = None
            self.overlay_detected_texture = None
            return
        layer["angle"] = float(self.overlay_angle_var.get())
        layer["scale"] = float(self.overlay_scale_var.get())
        layer["dx"] = float(self.overlay_dx_var.get())
        layer["dy"] = float(self.overlay_dy_var.get())
        layer["opacity"] = float(self.overlay_opacity_var.get())
        self.overlay_base_image = layer["image"]
        self.overlay_detected_grid = layer.get("detected_grid")
        self.overlay_detected_texture = layer.get("detected_texture")
        self.overlay_image_path_var.set(str(layer.get("path", "")))

    def overlay_load_layer_to_vars(self, idx=None):
        layers = getattr(self, "overlay_layers", [])
        if idx is None:
            idx = getattr(self, "overlay_selected_index", 0)
        if not (0 <= idx < len(layers)):
            self.overlay_base_image = None
            self.overlay_detected_grid = None
            self.overlay_detected_texture = None
            self.overlay_image_path_var.set("")
            return
        layer = layers[idx]
        self.overlay_selected_index = idx
        self.overlay_angle_var.set(float(layer.get("angle", 0.0)))
        self.overlay_scale_var.set(float(layer.get("scale", 1.0)))
        self.overlay_dx_var.set(float(layer.get("dx", 0.0)))
        self.overlay_dy_var.set(float(layer.get("dy", 0.0)))
        self.overlay_opacity_var.set(float(layer.get("opacity", 1.0)))
        self.overlay_base_image = layer["image"]
        self.overlay_detected_grid = layer.get("detected_grid")
        self.overlay_detected_texture = layer.get("detected_texture")
        self.overlay_image_path_var.set(str(layer.get("path", "")))
        self.overlay_refresh_layer_selector()

    def overlay_on_layer_selected(self, _event=None):
        self.overlay_sync_current_layer_from_vars()
        values = list(self.overlay_layer_selector["values"]) if hasattr(self, "overlay_layer_selector") else []
        try:
            idx = values.index(self.overlay_selected_index_var.get())
        except Exception:
            idx = 0
        self.overlay_load_layer_to_vars(idx)
        self.overlay_redraw()

    def overlay_add_layer_from_path(self, path: str, replace_current: bool = False):
        img = Image.open(path).convert("RGB")
        layer = {
            "path": str(path),
            "image": img,
            "angle": 0.0,
            "scale": 1.0,
            "dx": 0.0,
            "dy": 0.0,
            "opacity": 1.0 if not getattr(self, "overlay_layers", []) else 0.65,
            "flip_h": False,
            "flip_v": False,
            "detected_grid": None,
            "detected_texture": None,
        }
        if replace_current and getattr(self, "overlay_layers", []):
            idx = getattr(self, "overlay_selected_index", 0)
            self.overlay_layers[idx] = layer
        else:
            self.overlay_layers.append(layer)
            self.overlay_selected_index = len(self.overlay_layers) - 1
        self.overlay_base_preview_cache = None
        self.overlay_base_preview_key = None
        self.overlay_refresh_layer_selector()
        self.overlay_load_layer_to_vars(self.overlay_selected_index)
        self.overlay_status_var.set(f"Image chargée : {Path(path).name} — {img.size[0]}x{img.size[1]}")
        self.overlay_open_viewer()
        self.overlay_redraw()

    def overlay_add_image(self):
        path = filedialog.askopenfilename(
            title="Ajouter une image optique",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff"), ("Tous les fichiers", "*.*")]
        )
        if path:
            try:
                self.overlay_add_layer_from_path(path, replace_current=False)
            except Exception as exc:
                messagebox.showerror("Image optique", f"Impossible de charger l'image :\n{exc}")

    def overlay_remove_selected_image(self):
        layers = getattr(self, "overlay_layers", [])
        if not layers:
            return
        idx = max(0, min(getattr(self, "overlay_selected_index", 0), len(layers) - 1))
        layers.pop(idx)
        if layers:
            self.overlay_selected_index = min(idx, len(layers) - 1)
            self.overlay_load_layer_to_vars(self.overlay_selected_index)
        else:
            self.overlay_selected_index = 0
            self.overlay_base_image = None
            self.overlay_detected_grid = None
            self.overlay_detected_texture = None
            self.overlay_image_path_var.set("")
        self.overlay_refresh_layer_selector()
        self.overlay_redraw()

    def overlay_toggle_symmetry_h(self):
        layer = self.overlay_get_selected_layer()
        if layer is None:
            return
        layer["flip_h"] = not bool(layer.get("flip_h", False))
        self.overlay_redraw()

    def overlay_toggle_symmetry_v(self):
        layer = self.overlay_get_selected_layer()
        if layer is None:
            return
        layer["flip_v"] = not bool(layer.get("flip_v", False))
        self.overlay_redraw()

    def overlay_choose_sticker_color(self):
        color = colorchooser.askcolor(color=self.overlay_sticker_color_var.get(), title="Choisir la couleur du sticker")[1]
        if color:
            self.overlay_sticker_color_var.set(color)

    def overlay_set_sticker_mode(self, shape: str):
        self.overlay_sticker_mode_var.set(shape)
        self.overlay_status_var.set(f"Mode sticker : {shape}. Double-clic dans la fenêtre image pour le placer.")

    def overlay_clear_stickers(self):
        self.overlay_stickers = []
        self.overlay_redraw()

    def overlay_open_viewer(self):
        if getattr(self, "overlay_viewer", None) is not None:
            try:
                self.overlay_viewer.deiconify()
                self.overlay_viewer.lift()
                return
            except Exception:
                self.overlay_viewer = None
        win = tk.Toplevel(self)
        win.title("Analyse d'image — Vue")
        win.geometry("1200x850")
        win.protocol("WM_DELETE_WINDOW", self.overlay_close_viewer)
        frame = ttk.Frame(win)
        frame.pack(fill="both", expand=True)
        canvas = tk.Canvas(frame, bg="#202020", highlightthickness=0)
        hbar = ttk.Scrollbar(frame, orient="horizontal", command=canvas.xview)
        vbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        canvas.configure(xscrollcommand=hbar.set, yscrollcommand=vbar.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vbar.grid(row=0, column=1, sticky="ns")
        hbar.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        canvas.bind("<Configure>", lambda _e: self.overlay_redraw())
        canvas.bind("<ButtonPress-1>", self.overlay_drag_start)
        canvas.bind("<B1-Motion>", self.overlay_drag_move)
        canvas.bind("<ButtonRelease-1>", lambda _e: setattr(self, "overlay_drag_last", None))
        canvas.bind("<Double-Button-1>", self.overlay_viewer_canvas_click)
        canvas.bind("<MouseWheel>", self.overlay_mouse_wheel)
        self.overlay_viewer = win
        self.overlay_canvas = canvas
        self.overlay_hbar = hbar
        self.overlay_vbar = vbar

    def overlay_close_viewer(self):
        try:
            if getattr(self, "overlay_viewer", None) is not None:
                self.overlay_viewer.destroy()
        except Exception:
            pass
        self.overlay_viewer = None
        self.overlay_canvas = None

    def overlay_viewer_canvas_click(self, event):
        shape = self.overlay_sticker_mode_var.get() if hasattr(self, "overlay_sticker_mode_var") else "none"
        if shape not in ("cross", "circle"):
            return
        if not getattr(self, "overlay_canvas", None):
            return
        cx = self.overlay_canvas.canvasx(event.x)
        cy = self.overlay_canvas.canvasy(event.y)
        img_xy = self.overlay_canvas_to_image_xy(cx, cy)
        if img_xy is None:
            return
        self.overlay_stickers.append({
            "shape": shape,
            "color": self.overlay_sticker_color_var.get(),
            "x": float(img_xy[0]),
            "y": float(img_xy[1]),
            "size": 14,
        })
        self.overlay_redraw()

    def overlay_pick_image(self):
        path = filedialog.askopenfilename(
            title="Choisir l'image optique",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff"), ("Tous les fichiers", "*.*")]
        )
        if not path:
            return
        try:
            self.overlay_add_layer_from_path(path, replace_current=bool(getattr(self, "overlay_layers", [])))
        except Exception as exc:
            messagebox.showerror("Image optique", f"Impossible de charger l'image :\n{exc}")

    def overlay_pick_mask(self):
        path = filedialog.askopenfilename(
            title="Choisir le masque du grillage",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff"), ("Tous les fichiers", "*.*")]
        )
        if not path:
            return
        self.overlay_mask_path_var.set(path)
        try:
            # On garde le masque en taille native, mais toute l'extraction se fait une seule fois en cache.
            self.overlay_mask_image = Image.open(path).convert("RGBA")
            self.overlay_invalidate_mask_cache()
            self.overlay_mask_mode_var.set("Bleu uniquement")
            self.overlay_roi_enabled_var.set(True)
            self.overlay_status_var.set(f"Masque chargé : {Path(path).name} — extraction légère prête")
            self.overlay_auto_fit()
            self.overlay_redraw()
        except Exception as exc:
            messagebox.showerror("Masque", f"Impossible de charger le masque :\n{exc}")

    def overlay_reset(self):
        layer = self.overlay_get_selected_layer()
        if layer is None:
            return
        layer["angle"] = 0.0
        layer["scale"] = 1.0
        layer["dx"] = 0.0
        layer["dy"] = 0.0
        layer["opacity"] = 1.0
        layer["flip_h"] = False
        layer["flip_v"] = False
        layer["detected_grid"] = None
        layer["detected_texture"] = None
        self.overlay_detected_grid = None
        self.overlay_detected_texture = None
        self.overlay_load_layer_to_vars(self.overlay_selected_index)
        self.overlay_view_zoom_var.set(1.0)
        self.overlay_redraw()

    def overlay_reset_view_zoom(self):
        self.overlay_view_zoom_var.set(1.0)
        self.overlay_redraw()

    def overlay_detect_optic_circle(self) -> tuple[float, float, float] | None:
        base = getattr(self, "overlay_base_image", None)
        if base is None:
            return None

        gray = base.convert("L")
        hist = gray.histogram()
        total = sum(hist)
        if total <= 0:
            return None

        # Seuillage Otsu simple : l'optique claire se sépare bien du fond noir.
        sum_total = sum(i * hist[i] for i in range(256))
        sum_b = 0.0
        w_b = 0.0
        max_var = -1.0
        threshold = 20
        for i in range(256):
            w_b += hist[i]
            if w_b == 0:
                continue
            w_f = total - w_b
            if w_f == 0:
                break
            sum_b += i * hist[i]
            m_b = sum_b / w_b
            m_f = (sum_total - sum_b) / w_f
            var_between = w_b * w_f * (m_b - m_f) * (m_b - m_f)
            if var_between > max_var:
                max_var = var_between
                threshold = i

        # Garde-fou pour les images peu contrastées.
        threshold = max(8, min(180, threshold))
        binary = gray.point(lambda p: 255 if p > threshold else 0)
        bbox = binary.getbbox()
        if bbox is None:
            return base.width / 2, base.height / 2, min(base.width, base.height) / 2

        x0, y0, x1, y1 = bbox
        cx = (x0 + x1) / 2.0
        cy = (y0 + y1) / 2.0
        radius = min((x1 - x0), (y1 - y0)) / 2.0
        if radius <= 1:
            return None
        return cx, cy, radius

    def overlay_detect_mask_roi_circle(self) -> tuple[float, float, float] | None:
        mask = getattr(self, "overlay_mask_image", None)
        if mask is None:
            return None

        # Important :
        # Le masque complet peut être beaucoup plus grand que l'optique.
        # La référence à faire correspondre avec le cercle optique est le cercle pointillé bleu
        # de la documentation, représenté ici par Rayon ROI doc.
        # On ne corrige plus avec la bbox alpha, car cela ramenait le masque à une mauvaise taille.
        w, h = mask.size
        cx = w / 2.0
        cy = h / 2.0
        radius = float(self.overlay_roi_radius_var.get()) * min(w, h)
        return cx, cy, radius

    def overlay_auto_fit(self):
        base = getattr(self, "overlay_base_image", None)
        mask = getattr(self, "overlay_mask_image", None)

        self.overlay_angle_var.set(0.0)
        self.overlay_view_zoom_var.set(1.0)

        if base is None or mask is None or not mask.width or not mask.height:
            self.overlay_dx_var.set(0.0)
            self.overlay_dy_var.set(0.0)
            self.overlay_redraw()
            return

        optic = self.overlay_detect_optic_circle()
        roi = self.overlay_detect_mask_roi_circle()

        if optic is None or roi is None:
            scale = min(base.width / mask.width, base.height / mask.height)
            self.overlay_scale_var.set(max(0.01, scale))
            self.overlay_dx_var.set(0.0)
            self.overlay_dy_var.set(0.0)
            self.overlay_status_var.set(f"Auto ajustement simple : échelle={scale:.4g}.")
            self.overlay_redraw()
            return

        optic_cx, optic_cy, optic_r = optic
        mask_cx, mask_cy, mask_r = roi

        scale = max(0.001, optic_r / max(mask_r, 1.0))
        dx = optic_cx - base.width / 2.0 + (mask.width / 2.0 - mask_cx) * scale
        dy = optic_cy - base.height / 2.0 + (mask.height / 2.0 - mask_cy) * scale

        self.overlay_scale_var.set(scale)
        self.overlay_dx_var.set(dx)
        self.overlay_dy_var.set(dy)
        self.overlay_status_var.set(
            f"Auto cercle : optique centre=({optic_cx:.0f},{optic_cy:.0f}) R={optic_r:.0f} | "
            f"masque R={mask_r:.0f} | échelle={scale:.4g}"
        )
        self.overlay_redraw()

    def overlay_get_mask_alpha(self) -> Image.Image | None:
        mask = getattr(self, "overlay_mask_image", None)
        if mask is None:
            return None

        key = (
            id(mask),
            self.overlay_mask_mode_var.get(),
            bool(self.overlay_roi_enabled_var.get()),
            round(float(self.overlay_roi_radius_var.get()), 4),
            int(round(float(self.overlay_line_thin_var.get()))),
        )
        if self.overlay_mask_alpha_cache is not None and self.overlay_mask_alpha_key == key:
            return self.overlay_mask_alpha_cache

        rgba = mask.convert("RGBA")
        r, g, b, a = rgba.split()
        mode = self.overlay_mask_mode_var.get()

        if mode == "Alpha existant":
            alpha = a
        elif mode == "Sombre":
            gray = mask.convert("L")
            alpha = gray.point(lambda p: 255 - p if p < 235 else 0)
        elif mode == "Lumineux":
            gray = mask.convert("L")
            alpha = gray.point(lambda p: 255 if p > 40 else 0)
        elif mode == "Bleu uniquement":
            # Très rapide : opérations Pillow en C, pas de boucle Python.
            blue_minus_red = ImageChops.subtract(b, r).point(lambda p: 255 if p > 40 else 0)
            green_minus_red = ImageChops.subtract(g, r).point(lambda p: 255 if p > 15 else 0)
            blue_level = b.point(lambda p: 255 if p > 80 else 0)
            alpha = ImageChops.multiply(ImageChops.multiply(blue_minus_red, green_minus_red), blue_level)
        else:
            # Couleur : saturation approximative = max(R,G,B)-min(R,G,B)
            max_rg = ImageChops.lighter(r, g)
            max_rgb = ImageChops.lighter(max_rg, b)
            min_rg = ImageChops.darker(r, g)
            min_rgb = ImageChops.darker(min_rg, b)
            sat = ImageChops.subtract(max_rgb, min_rgb)
            alpha = sat.point(lambda p: 255 if p > 35 else 0)

        if bool(self.overlay_roi_enabled_var.get()):
            w, h = alpha.size
            radius = float(self.overlay_roi_radius_var.get()) * min(w, h)
            cx = w / 2
            cy = h / 2
            roi = Image.new("L", (w, h), 0)
            draw = ImageDraw.Draw(roi)
            draw.ellipse((cx - radius, cy - radius, cx + radius, cy + radius), fill=255)
            alpha = ImageChops.multiply(alpha, roi)

        # Amincissement réel des traits, sans disparition du masque.
        # Principe : on érode les traits, mais on conserve toujours leur squelette central.
        # Donc si un trait est déjà à 1 px, il reste visible.
        try:
            thin = int(round(float(self.overlay_line_thin_var.get())))
        except Exception:
            thin = 0
        if thin > 0:
            skeleton = self.overlay_skeletonize_alpha(alpha)
            prev = alpha
            for _ in range(min(thin, 8)):
                candidate = prev.filter(ImageFilter.MinFilter(3))
                # Si l'érosion retire vraiment trop, on garde le squelette.
                if sum(candidate.histogram()[1:]) < max(sum(prev.histogram()[1:]), 1) * 0.08:
                    prev = skeleton
                    break
                prev = candidate
            alpha = ImageChops.lighter(prev, skeleton)

        self.overlay_mask_alpha_cache = alpha
        self.overlay_mask_alpha_key = key
        return alpha

    def overlay_mask_to_rgba(self, opacity: float | None = None) -> Image.Image | None:
        alpha = self.overlay_get_mask_alpha()
        if alpha is None:
            return None
        if opacity is None:
            opacity = float(self.overlay_opacity_var.get())
        opacity = max(0.0, min(1.0, opacity))
        use_alpha = alpha.point(lambda p: int(p * opacity))
        colored = Image.new("RGBA", alpha.size, (255, 0, 0, 0))
        colored.putalpha(use_alpha)
        return colored

    def overlay_composite_fullres(self) -> Image.Image | None:
        layers = getattr(self, "overlay_layers", [])
        if not layers:
            return None

        base_w = max(layer["image"].width for layer in layers)
        base_h = max(layer["image"].height for layer in layers)
        composite = Image.new("RGBA", (base_w, base_h), (0, 0, 0, 255))

        for layer in layers:
            img = layer["image"].convert("RGBA")
            if layer.get("flip_h"):
                img = img.transpose(Image.FLIP_LEFT_RIGHT)
            if layer.get("flip_v"):
                img = img.transpose(Image.FLIP_TOP_BOTTOM)

            layer_scale = max(0.01, float(layer.get("scale", 1.0)))
            if abs(layer_scale - 1.0) > 1e-9:
                img = img.resize(
                    (max(1, int(round(img.width * layer_scale))), max(1, int(round(img.height * layer_scale)))),
                    Image.Resampling.LANCZOS if hasattr(Image, "Resampling") else Image.LANCZOS
                )

            angle = float(layer.get("angle", 0.0))
            if abs(angle) > 1e-9:
                img = img.rotate(angle, expand=True, resample=Image.Resampling.BICUBIC if hasattr(Image, "Resampling") else Image.BICUBIC)

            opacity = max(0.0, min(1.0, float(layer.get("opacity", 1.0))))
            if opacity < 1.0:
                alpha = img.getchannel("A").point(lambda p: int(p * opacity))
                img.putalpha(alpha)

            x = int(round(base_w / 2 - img.width / 2 + float(layer.get("dx", 0.0))))
            y = int(round(base_h / 2 - img.height / 2 + float(layer.get("dy", 0.0))))
            composite.alpha_composite(img, (x, y))

            texture = layer.get("detected_texture")
            grid = layer.get("detected_grid")
            if texture is not None and np is not None:
                arr = np.asarray(texture, dtype=np.float32)
                vals = arr[arr > 0]
                if vals.size:
                    thr = float(np.percentile(vals, max(70.0, min(99.0, float(self.overlay_detect_percentile_var.get()) - 6.0))))
                    alpha_arr = np.zeros_like(arr, dtype=np.uint8)
                    sel = arr >= thr
                    if np.any(sel):
                        local = np.clip((arr - thr) / max(float(arr.max()) - thr, 1e-6), 0.0, 1.0)
                        alpha_arr = (local * 255.0 * 0.7).astype(np.uint8)
                        alpha_arr[~sel] = 0
                        alpha_img = Image.fromarray(alpha_arr, mode="L")
                        scale_back = 1.0 / max(float(grid.get("base_to_work_scale", 1.0)), 1e-9) if grid else 1.0
                        red = Image.new("RGBA", (
                            max(1, int(round(alpha_img.width * scale_back))),
                            max(1, int(round(alpha_img.height * scale_back)))
                        ), (255, 0, 0, 0))
                        red_alpha = alpha_img.resize(red.size, Image.Resampling.BILINEAR if hasattr(Image, "Resampling") else Image.BILINEAR)
                        red.putalpha(red_alpha)
                        if layer.get("flip_h"):
                            red = red.transpose(Image.FLIP_LEFT_RIGHT)
                        if layer.get("flip_v"):
                            red = red.transpose(Image.FLIP_TOP_BOTTOM)
                        if abs(layer_scale - 1.0) > 1e-9:
                            red = red.resize((max(1, int(round(red.width * layer_scale))), max(1, int(round(red.height * layer_scale)))), Image.Resampling.BILINEAR if hasattr(Image, "Resampling") else Image.BILINEAR)
                        if abs(angle) > 1e-9:
                            red = red.rotate(angle, expand=True, resample=Image.Resampling.BICUBIC if hasattr(Image, "Resampling") else Image.BICUBIC)
                        composite.alpha_composite(red, (x, y))

        draw = ImageDraw.Draw(composite)
        for st in getattr(self, "overlay_stickers", []):
            x = float(st.get("x", 0.0))
            y = float(st.get("y", 0.0))
            size = float(st.get("size", 14))
            color = st.get("color", "#ff0000")
            if st.get("shape") == "circle":
                draw.ellipse((x - size, y - size, x + size, y + size), outline=color, width=2)
            else:
                draw.line((x - size, y, x + size, y), fill=color, width=2)
                draw.line((x, y - size, x, y + size), fill=color, width=2)

        return composite.convert("RGB")

    def overlay_get_base_preview(self, composite_size: tuple[int, int]) -> tuple[Image.Image, int, int, float]:
        base = getattr(self, "overlay_base_image", None)
        if base is None:
            return None, 0, 0, 1.0

        cw, ch = composite_size
        fit_scale = min(cw / base.width, ch / base.height, 1.0)
        view_zoom = max(0.05, min(20.0, float(self.overlay_view_zoom_var.get())))
        scale = fit_scale * view_zoom
        disp_w = max(1, int(base.width * scale))
        disp_h = max(1, int(base.height * scale))
        key = (id(base), disp_w, disp_h)
        if self.overlay_base_preview_cache is None or self.overlay_base_preview_key != key:
            self.overlay_base_preview_cache = base.resize(
                (disp_w, disp_h),
                Image.Resampling.LANCZOS if hasattr(Image, "Resampling") else Image.LANCZOS
            ).convert("RGBA")
            self.overlay_base_preview_key = key
        return self.overlay_base_preview_cache.copy(), disp_w, disp_h, scale


    def overlay_line_rect_intersections(self, nx: float, ny: float, rho: float, w: float, h: float):
        pts = []
        cx = w / 2.0
        cy = h / 2.0

        for x in (0.0, w):
            if abs(ny) > 1e-9:
                y = (rho - nx * (x - cx)) / ny + cy
                if 0.0 <= y <= h:
                    pts.append((x, y))

        for y in (0.0, h):
            if abs(nx) > 1e-9:
                x = (rho - ny * (y - cy)) / nx + cx
                if 0.0 <= x <= w:
                    pts.append((x, y))

        unique = []
        for p in pts:
            if not any(abs(p[0] - q[0]) < 1e-6 and abs(p[1] - q[1]) < 1e-6 for q in unique):
                unique.append(p)
        if len(unique) >= 2:
            return unique[0], unique[1]
        return None

    def overlay_draw_detected_grid(self, canvas, x0: int, y0: int, display_scale: float):
        """Affiche la vraie carte de texture détectée.

        On évite volontairement de dessiner un pavage hexagonal infini :
        tant que la phase du motif n'est pas déterminée, cela donne un résultat trompeur.
        """
        grid = getattr(self, "overlay_detected_grid", None)
        texture = getattr(self, "overlay_detected_texture", None)
        base = getattr(self, "overlay_base_image", None)
        if base is None or texture is None:
            return

        try:
            edge_map = texture
            if np is None:
                return

            arr = np.asarray(edge_map, dtype=np.float32)
            if arr.size == 0:
                return

            vals = arr[arr > 0]
            if vals.size:
                # N'affiche que les points réellement significatifs pour éviter de remplir tout le champ.
                thr = float(np.percentile(vals, max(70.0, min(99.0, float(self.overlay_detect_percentile_var.get()) - 6.0))))
            else:
                thr = 1.0

            alpha = np.zeros_like(arr, dtype=np.uint8)
            sel = arr >= thr
            if np.any(sel):
                local = (arr - thr) / max(float(arr.max()) - thr, 1e-6)
                local = np.clip(local, 0.0, 1.0)
                opacity = max(0.05, min(1.0, float(self.overlay_opacity_var.get())))
                alpha = (local * 255.0 * opacity).astype(np.uint8)
                alpha[~sel] = 0

            alpha_img = Image.fromarray(alpha, mode="L")
            disp_w = getattr(self, "overlay_display_w", int(base.width * display_scale))
            disp_h = getattr(self, "overlay_display_h", int(base.height * display_scale))
            alpha_disp = alpha_img.resize(
                (max(1, disp_w), max(1, disp_h)),
                Image.Resampling.BILINEAR if hasattr(Image, "Resampling") else Image.BILINEAR
            )

            red = Image.new("RGBA", alpha_disp.size, (255, 0, 0, 0))
            red.putalpha(alpha_disp)
            self.overlay_detection_photo = ImageTk.PhotoImage(red)
            canvas.create_image(x0, y0, anchor="nw", image=self.overlay_detection_photo)

            # Cercle optique : pour visualiser la zone de contrôle.
            optic = self.overlay_detect_optic_circle()
            if optic is not None:
                cx, cy, radius = optic
                canvas.create_oval(
                    x0 + (cx - radius) * display_scale,
                    y0 + (cy - radius) * display_scale,
                    x0 + (cx + radius) * display_scale,
                    y0 + (cy + radius) * display_scale,
                    outline="#ff7070",
                    width=2,
                )

            if grid:
                spacing = float(grid.get("spacing", 0.0)) / max(float(grid.get("base_to_work_scale", 1.0)), 1e-9)
                canvas.create_text(
                    x0 + 10,
                    y0 + 10,
                    text=f"Texture grillage détectée | angle={float(grid.get('angle', 0.0)):.1f}° | pas≈{spacing:.1f}px | points={int(grid.get('points', 0))}",
                    fill="#ff7070",
                    anchor="nw",
                    font=("Arial", 10, "bold"),
                )
        except Exception as exc:
            canvas.create_text(x0 + 10, y0 + 10, text=f"Erreur affichage détection: {exc}", fill="red", anchor="nw")

    def overlay_redraw(self):
        if not getattr(self, "overlay_canvas", None):
            return

        canvas = self.overlay_canvas
        canvas.delete("all")

        layers = getattr(self, "overlay_layers", [])
        if not layers:
            canvas.create_text(
                max(canvas.winfo_width() // 2, 200),
                max(canvas.winfo_height() // 2, 120),
                text="Charge une ou plusieurs images optiques.",
                fill="white",
                font=("Arial", 14, "bold"),
            )
            return

        cw = max(canvas.winfo_width(), 1)
        ch = max(canvas.winfo_height(), 1)
        base_w = max(layer["image"].width for layer in layers)
        base_h = max(layer["image"].height for layer in layers)

        fit_scale = min(cw / base_w, ch / base_h, 1.0)
        display_scale = fit_scale * max(0.05, min(20.0, float(self.overlay_view_zoom_var.get())))
        disp_w = max(1, int(round(base_w * display_scale)))
        disp_h = max(1, int(round(base_h * display_scale)))

        # Optimisation importante : on compose directement à la taille écran.
        # Avant, toute la composition était faite en pleine résolution puis réduite,
        # ce qui devenait très lent dès qu'une deuxième image était ajoutée.
        preview = Image.new("RGBA", (disp_w, disp_h), (0, 0, 0, 255))

        for layer in layers:
            img_src = layer["image"]
            layer_scale = max(0.01, float(layer.get("scale", 1.0)))
            transform_scale = display_scale * layer_scale

            img = img_src.convert("RGBA")
            if layer.get("flip_h"):
                img = img.transpose(Image.FLIP_LEFT_RIGHT)
            if layer.get("flip_v"):
                img = img.transpose(Image.FLIP_TOP_BOTTOM)

            new_w = max(1, int(round(img_src.width * transform_scale)))
            new_h = max(1, int(round(img_src.height * transform_scale)))
            img = img.resize((new_w, new_h), Image.Resampling.BILINEAR if hasattr(Image, "Resampling") else Image.BILINEAR)

            angle = float(layer.get("angle", 0.0))
            if abs(angle) > 1e-9:
                img = img.rotate(angle, expand=True, resample=Image.Resampling.BICUBIC if hasattr(Image, "Resampling") else Image.BICUBIC)

            opacity = max(0.0, min(1.0, float(layer.get("opacity", 1.0))))
            if opacity < 1.0:
                alpha = img.getchannel("A").point(lambda p: int(p * opacity))
                img.putalpha(alpha)

            x = int(round(disp_w / 2 - img.width / 2 + float(layer.get("dx", 0.0)) * display_scale))
            y = int(round(disp_h / 2 - img.height / 2 + float(layer.get("dy", 0.0)) * display_scale))
            preview.alpha_composite(img, (x, y))

            texture = layer.get("detected_texture")
            grid = layer.get("detected_grid")
            if texture is not None and np is not None:
                try:
                    arr = np.asarray(texture, dtype=np.float32)
                    vals = arr[arr > 0]
                    if vals.size:
                        thr = float(np.percentile(vals, max(70.0, min(99.0, float(self.overlay_detect_percentile_var.get()) - 6.0))))
                        alpha_arr = np.zeros_like(arr, dtype=np.uint8)
                        sel = arr >= thr
                        if np.any(sel):
                            local = np.clip((arr - thr) / max(float(arr.max()) - thr, 1e-6), 0.0, 1.0)
                            alpha_arr = (local * 255.0 * 0.7).astype(np.uint8)
                            alpha_arr[~sel] = 0
                            alpha_img = Image.fromarray(alpha_arr, mode="L")
                            red = Image.new("RGBA", (new_w, new_h), (255, 0, 0, 0))
                            red_alpha = alpha_img.resize((new_w, new_h), Image.Resampling.BILINEAR if hasattr(Image, "Resampling") else Image.BILINEAR)
                            red.putalpha(red_alpha)
                            if layer.get("flip_h"):
                                red = red.transpose(Image.FLIP_LEFT_RIGHT)
                            if layer.get("flip_v"):
                                red = red.transpose(Image.FLIP_TOP_BOTTOM)
                            if abs(angle) > 1e-9:
                                red = red.rotate(angle, expand=True, resample=Image.Resampling.BICUBIC if hasattr(Image, "Resampling") else Image.BICUBIC)
                            preview.alpha_composite(red, (x, y))
                except Exception:
                    pass

        # Stickers dans le repère de la composition.
        draw = ImageDraw.Draw(preview)
        for st in getattr(self, "overlay_stickers", []):
            sx = float(st.get("x", 0.0)) * display_scale
            sy = float(st.get("y", 0.0)) * display_scale
            size = float(st.get("size", 14)) * max(display_scale, 0.5)
            color = st.get("color", "#ff0000")
            if st.get("shape") == "circle":
                draw.ellipse((sx - size, sy - size, sx + size, sy + size), outline=color, width=2)
            else:
                draw.line((sx - size, sy, sx + size, sy), fill=color, width=2)
                draw.line((sx, sy - size, sx, sy + size), fill=color, width=2)

        self.overlay_photo = ImageTk.PhotoImage(preview.convert("RGB"))
        self.overlay_display_scale = display_scale

        x0 = max((cw - disp_w) // 2, 0)
        y0 = max((ch - disp_h) // 2, 0)
        self.overlay_image_x0 = x0
        self.overlay_image_y0 = y0
        self.overlay_display_w = disp_w
        self.overlay_display_h = disp_h
        canvas.configure(scrollregion=(0, 0, max(cw, x0 + disp_w), max(ch, y0 + disp_h)))
        canvas.create_image(x0, y0, anchor="nw", image=self.overlay_photo)
        canvas.create_rectangle(x0, y0, x0 + disp_w, y0 + disp_h, outline="#808080")

        layer = self.overlay_get_selected_layer()
        info = "Aucune image sélectionnée"
        if layer is not None:
            grid = layer.get("detected_grid")
            extra = ""
            if grid:
                spacing = float(grid.get("spacing", 0.0)) / max(float(grid.get("base_to_work_scale", 1.0)), 1e-9)
                extra = f" | grille angle≈{float(grid.get('angle', 0.0)):.2f}° pas≈{spacing:.1f}px"
            info = (
                f"Image {self.overlay_selected_index + 1}/{len(getattr(self, 'overlay_layers', []))} | "
                f"rot={float(self.overlay_angle_var.get()):.2f}° | scale={float(self.overlay_scale_var.get()):.3f} | "
                f"dx={float(self.overlay_dx_var.get()):.0f}px | dy={float(self.overlay_dy_var.get()):.0f}px | "
                f"opacity={float(self.overlay_opacity_var.get()):.2f} | zoom vue={float(self.overlay_view_zoom_var.get()):.2f}x{extra}"
            )
        self.overlay_status_var.set(info)

    def overlay_drag_start(self, event):
        self.overlay_drag_last = (event.x, event.y)

    def overlay_drag_move(self, event):
        layer = self.overlay_get_selected_layer()
        if layer is None:
            return
        if self.overlay_drag_last is None:
            self.overlay_drag_last = (event.x, event.y)
            return
        last_x, last_y = self.overlay_drag_last
        dx_screen = event.x - last_x
        dy_screen = event.y - last_y
        self.overlay_drag_last = (event.x, event.y)

        scale = max(getattr(self, "overlay_display_scale", 1.0), 0.001)
        self.overlay_dx_var.set(float(self.overlay_dx_var.get()) + dx_screen / scale)
        self.overlay_dy_var.set(float(self.overlay_dy_var.get()) + dy_screen / scale)
        self.overlay_sync_current_layer_from_vars()
        self.overlay_redraw()

    def overlay_mouse_wheel(self, event):
        layer = self.overlay_get_selected_layer()
        if layer is None:
            return
        if event.state & 0x0004:
            factor = 1.15 if event.delta > 0 else 1 / 1.15
            self.overlay_view_zoom_var.set(max(0.05, min(20.0, float(self.overlay_view_zoom_var.get()) * factor)))
        elif event.state & 0x0001:
            step = 0.2 if event.delta > 0 else -0.2
            self.overlay_angle_var.set(float(self.overlay_angle_var.get()) + step)
        else:
            factor = 1.02 if event.delta > 0 else 1 / 1.02
            self.overlay_scale_var.set(max(0.01, float(self.overlay_scale_var.get()) * factor))
        self.overlay_sync_current_layer_from_vars()
        self.overlay_redraw()

    def overlay_build_optic_edge_map(self, max_dim: int = 640):
        """Carte de contraste du grillage inspirée de l'algo Halcon fourni.

        Adaptation :
        - MedianRect large / medianRect fin
        - SubImage autour de 128
        - neutralisation hors ROI optique
        - FFT + détection de pics de fréquence
        - notch filter sur pics + rotations 60°/120°
        - retour spatial pour isoler la texture chicken-wire
        """
        if np is None:
            messagebox.showerror("Détection grillage", "La détection automatique nécessite numpy.")
            return None

        base = getattr(self, "overlay_base_image", None)
        if base is None:
            return None

        sens = max(0.25, min(4.0, float(self.overlay_detect_sensitivity_var.get())))
        perc = max(70.0, min(99.5, float(self.overlay_detect_percentile_var.get())))
        blur_param = max(3.0, min(35.0, float(self.overlay_detect_blur_var.get())))

        gray = base.convert("L")
        scale = min(max_dim / max(gray.width, gray.height), 1.0)
        work_w = max(1, int(gray.width * scale))
        work_h = max(1, int(gray.height * scale))
        small = gray.resize((work_w, work_h), Image.Resampling.BILINEAR if hasattr(Image, "Resampling") else Image.BILINEAR)

        # Approximation MedianRect Halcon : deux médianes de tailles différentes.
        k_small = max(3, int(round(blur_param // 3)) | 1)
        k_large = max(k_small + 2, int(round(blur_param)) | 1)
        if k_large % 2 == 0:
            k_large += 1

        med_tube = small.filter(ImageFilter.MedianFilter(size=k_large))
        med_smooth = small.filter(ImageFilter.MedianFilter(size=k_small))

        arr_tube = np.asarray(med_tube, dtype=np.float32)
        arr_smooth = np.asarray(med_smooth, dtype=np.float32)

        # SubImage(medianTube, imageSmooth, 1, 128)
        sub = arr_tube - arr_smooth + 128.0

        optic = self.overlay_detect_optic_circle()
        if optic is not None:
            cx, cy, radius = optic
            cx *= scale
            cy *= scale
            radius *= scale * 0.93
            yy, xx = np.ogrid[:work_h, :work_w]
            roi = ((xx - cx) ** 2 + (yy - cy) ** 2) <= radius ** 2
        else:
            roi = np.ones((work_h, work_w), dtype=bool)

        # Equivalent PaintRegion hors regionControle à 128.
        sub = sub.copy()
        sub[~roi] = 128.0

        # FFT avec DC centré.
        zero_mean = sub - 128.0
        fft = np.fft.fftshift(np.fft.fft2(zero_mean))
        power = np.log1p(np.abs(fft)).astype(np.float32)

        fy = np.arange(work_h, dtype=np.float32) - work_h / 2.0
        fx = np.arange(work_w, dtype=np.float32) - work_w / 2.0
        FX, FY = np.meshgrid(fx, fy)
        FR = np.sqrt(FX * FX + FY * FY)

        # Recherche pics fréquence hors centre.
        center_exclusion = max(12.0, min(work_w, work_h) * 0.035)
        valid = FR > center_exclusion

        vals = power[valid]
        notch_mask = np.ones((work_h, work_w), dtype=np.float32)

        if vals.size > 0:
            thr = np.percentile(vals, min(99.7, max(perc, 85.0)))
            peak = valid & (power >= thr)

            # On garde peu de pics, les plus forts.
            ys, xs = np.where(peak)
            if ys.size > 0:
                strengths = power[ys, xs]
                order = np.argsort(strengths)[::-1][:10]
                radius_notch = max(1.5, min(work_w, work_h) * 0.004 * sens)

                cy0 = work_h / 2.0
                cx0 = work_w / 2.0

                for idx in order:
                    vx = float(xs[idx] - cx0)
                    vy = float(ys[idx] - cy0)

                    # Pic + symétrique + rotations 60/120° comme ton algo Halcon.
                    for base_angle in (0.0, math.pi / 3.0, 2.0 * math.pi / 3.0):
                        ca = math.cos(base_angle)
                        sa = math.sin(base_angle)
                        rx = ca * vx - sa * vy
                        ry = sa * vx + ca * vy
                        for sign in (1.0, -1.0):
                            px = cx0 + sign * rx
                            py = cy0 + sign * ry
                            d2 = (FX - (px - cx0)) ** 2 + (FY - (py - cy0)) ** 2
                            notch_mask[d2 <= radius_notch * radius_notch] = 0.0

        filtered_fft = fft * notch_mask
        image_result = np.real(np.fft.ifft2(np.fft.ifftshift(filtered_fft)))

        # ImageTextureFFT = ImageSub - imageResultFFT + 128
        texture = zero_mean - image_result + 128.0

        # Détection lignes : contraste local autour de 128.
        edge = np.abs(texture - 128.0)
        edge[~roi] = 0.0

        vals = edge[roi]
        if vals.size:
            p_low = np.percentile(vals, max(40.0, perc - 25.0))
            p_high = np.percentile(vals, min(99.9, perc + 5.0))
            edge = (edge - p_low) / max(p_high - p_low, 1.0)
        else:
            edge = edge / max(float(edge.max()), 1.0)

        edge = np.clip(edge * sens, 0.0, 1.0)
        edge = np.power(edge, 0.85)
        edge *= roi.astype(np.float32)

        return edge, scale, (work_w, work_h)

    def overlay_mask_map_for_angle(self, angle_deg: float, base_to_work_scale: float, work_size: tuple[int, int]):
        """Transforme le masque en carte basse résolution selon l'angle testé."""
        if np is None:
            return None

        alpha = self.overlay_get_mask_alpha()
        base = getattr(self, "overlay_base_image", None)
        if alpha is None or base is None:
            return None

        work_w, work_h = work_size
        s = max(0.001, float(self.overlay_scale_var.get()) * base_to_work_scale)
        new_w = max(1, int(round(alpha.width * s)))
        new_h = max(1, int(round(alpha.height * s)))

        alpha_scaled = alpha.resize((new_w, new_h), Image.Resampling.NEAREST if hasattr(Image, "Resampling") else Image.NEAREST)

        if abs(angle_deg) > 0.0001:
            alpha_scaled = alpha_scaled.rotate(
                angle_deg,
                expand=True,
                resample=Image.Resampling.NEAREST if hasattr(Image, "Resampling") else Image.NEAREST
            )

        dx = float(self.overlay_dx_var.get()) * base_to_work_scale
        dy = float(self.overlay_dy_var.get()) * base_to_work_scale

        x = int(round(work_w / 2 - alpha_scaled.width / 2 + dx))
        y = int(round(work_h / 2 - alpha_scaled.height / 2 + dy))

        canvas = Image.new("L", (work_w, work_h), 0)
        canvas.paste(alpha_scaled, (x, y))
        return np.asarray(canvas, dtype=np.float32) / 255.0

    def overlay_score_grid_angle(self, edge_map, angle_deg: float, base_to_work_scale: float, work_size: tuple[int, int]) -> float:
        mask_map = self.overlay_mask_map_for_angle(angle_deg, base_to_work_scale, work_size)
        if mask_map is None:
            return -1e9
        weight = float(mask_map.sum())
        if weight <= 1.0:
            return -1e9
        return float((edge_map * mask_map).sum() / weight)


    def overlay_crop_for_hex_features(self, source: str):
        """Retourne une image L centrée sur la zone utile pour estimer le pas/orientation hexagonal."""
        if source == "optic":
            base = getattr(self, "overlay_base_image", None)
            if base is None:
                return None, None

            gray = base.convert("L")
            circle = self.overlay_detect_optic_circle()
            if circle is not None:
                cx, cy, r = circle
                # Zone centrale : on évite le bord noir de l'optique et les gros défauts périphériques.
                rr = r * 0.62
                x0 = max(0, int(cx - rr))
                y0 = max(0, int(cy - rr))
                x1 = min(gray.width, int(cx + rr))
                y1 = min(gray.height, int(cy + rr))
                crop = gray.crop((x0, y0, x1, y1))
                meta = {"x0": x0, "y0": y0, "scale_to_original": 1.0}
                return crop, meta

            return gray, {"x0": 0, "y0": 0, "scale_to_original": 1.0}

        # Masque : on utilise directement l'alpha extrait, éventuellement restreint à la ROI.
        alpha = self.overlay_get_mask_alpha()
        if alpha is None:
            return None, None

        bbox = alpha.getbbox()
        if bbox is not None:
            x0, y0, x1, y1 = bbox
            # Garde une marge pour ne pas couper les hexagones du centre.
            mx = int((x1 - x0) * 0.08)
            my = int((y1 - y0) * 0.08)
            x0 = max(0, x0 - mx)
            y0 = max(0, y0 - my)
            x1 = min(alpha.width, x1 + mx)
            y1 = min(alpha.height, y1 + my)
            return alpha.crop((x0, y0, x1, y1)), {"x0": x0, "y0": y0, "scale_to_original": 1.0}

        return alpha, {"x0": 0, "y0": 0, "scale_to_original": 1.0}

    def overlay_fft_hex_features(self, img_l: Image.Image, source: str, max_dim: int = 620):
        """Estime orientation et pas du réseau hexagonal par FFT.

        On ne cherche pas à reconnaître chaque hexagone : on estime le pas moyen
        et l'orientation dominante à partir de quelques directions répétées.
        """
        if np is None or img_l is None:
            return None

        w, h = img_l.size
        if w < 32 or h < 32:
            return None

        resize_scale = min(max_dim / max(w, h), 1.0)
        ww = max(16, int(w * resize_scale))
        hh = max(16, int(h * resize_scale))

        small = img_l.resize((ww, hh), Image.Resampling.BILINEAR if hasattr(Image, "Resampling") else Image.BILINEAR)

        if source == "optic":
            # Le grillage est faible : on retire le fond lent et on amplifie les variations locales.
            blur_radius = max(3, int(max(ww, hh) * 0.018))
            blur = small.filter(ImageFilter.GaussianBlur(radius=blur_radius))
            work = ImageChops.difference(small, blur)
            work = work.filter(ImageFilter.GaussianBlur(radius=0.45))
        else:
            work = small

        arr = np.asarray(work, dtype=np.float32)
        arr = arr - float(arr.mean())
        std = float(arr.std())
        if std < 1e-6:
            return None
        arr = arr / std

        # Fenêtre Hann pour limiter les effets du bord de crop.
        wy = np.hanning(arr.shape[0]).astype(np.float32)
        wx = np.hanning(arr.shape[1]).astype(np.float32)
        arr = arr * wy[:, None] * wx[None, :]

        fft = np.fft.fftshift(np.fft.fft2(arr))
        mag = np.abs(fft).astype(np.float32)

        fy = np.fft.fftshift(np.fft.fftfreq(arr.shape[0]))
        fx = np.fft.fftshift(np.fft.fftfreq(arr.shape[1]))
        FX, FY = np.meshgrid(fx, fy)
        FR = np.sqrt(FX * FX + FY * FY)

        # Ignore le DC / fond lent et les fréquences trop hautes.
        valid = (FR > 0.012) & (FR < 0.33)
        if not np.any(valid):
            return None

        m = mag.copy()
        m[~valid] = 0.0

        # On cherche la fréquence dominante par histogramme radial pondéré.
        vals = m[valid]
        if vals.size < 10 or float(vals.max()) <= 0:
            return None

        # Seuls les pics forts sont utiles.
        threshold = float(np.percentile(vals, 99.4 if source == "optic" else 99.0))
        peak_mask = valid & (m >= threshold)

        if int(np.count_nonzero(peak_mask)) < 6:
            threshold = float(np.percentile(vals, 98.8))
            peak_mask = valid & (m >= threshold)

        r_vals = FR[peak_mask]
        weights = m[peak_mask]
        if r_vals.size < 3:
            return None

        bins = np.linspace(0.012, 0.33, 180)
        hist, edges = np.histogram(r_vals, bins=bins, weights=weights)
        best = int(np.argmax(hist))
        f_peak = float((edges[best] + edges[best + 1]) / 2.0)
        if f_peak <= 0:
            return None

        # Angle : on prend les points proches du rayon fréquentiel dominant.
        band = peak_mask & (np.abs(FR - f_peak) < max(0.008, f_peak * 0.22))
        if int(np.count_nonzero(band)) < 3:
            band = peak_mask

        # Angle des normales en fréquence, converti en angle de lignes.
        normal_angles = np.degrees(np.arctan2(FY[band], FX[band]))
        line_angles = (normal_angles + 90.0) % 60.0
        angle_weights = m[band]

        angle_bins = np.linspace(0.0, 60.0, 121)
        ahist, aedges = np.histogram(line_angles, bins=angle_bins, weights=angle_weights)
        abest = int(np.argmax(ahist))
        angle_mod = float((aedges[abest] + aedges[abest + 1]) / 2.0)

        # Le pas est inverse de la fréquence. On repasse à l'échelle de l'image originale du crop.
        pitch_in_crop_px = (1.0 / f_peak) / max(resize_scale, 1e-9)

        confidence = float(hist[best] / (float(hist.mean()) + 1e-9))
        return {
            "pitch": pitch_in_crop_px,
            "angle_mod": angle_mod,
            "frequency": f_peak,
            "confidence": confidence,
            "resize_scale": resize_scale,
            "source_size": (w, h),
            "work_size": (ww, hh),
        }

    def overlay_estimate_hex_features(self, source: str):
        crop, meta = self.overlay_crop_for_hex_features(source)
        if crop is None:
            return None
        features = self.overlay_fft_hex_features(crop, source)
        if features is None:
            return None
        features["meta"] = meta
        return features

    def overlay_mod60_delta(self, optic_angle: float, mask_angle: float) -> float:
        # Différence minimale dans un motif périodique tous les 60°.
        return ((optic_angle - mask_angle + 30.0) % 60.0) - 30.0

    def overlay_apply_hex_feature_fit(self, optic_features: dict, mask_features: dict):
        """Applique échelle + rotation issues du pas/orientation hexagonale, puis centre sur le cercle optique."""
        base = getattr(self, "overlay_base_image", None)
        mask = getattr(self, "overlay_mask_image", None)
        if base is None or mask is None:
            return False

        optic_pitch = float(optic_features["pitch"])
        mask_pitch = float(mask_features["pitch"])
        if optic_pitch <= 0 or mask_pitch <= 0:
            return False

        # Le facteur d'échelle du masque = pas optique / pas masque.
        scale = optic_pitch / mask_pitch

        # Rotation : orientation optique - orientation masque, modulo 60°.
        angle = self.overlay_mod60_delta(float(optic_features["angle_mod"]), float(mask_features["angle_mod"]))

        optic = self.overlay_detect_optic_circle()
        roi = self.overlay_detect_mask_roi_circle()

        if optic is not None and roi is not None:
            optic_cx, optic_cy, _optic_r = optic
            mask_cx, mask_cy, _mask_r = roi
            dx = optic_cx - base.width / 2.0 + (mask.width / 2.0 - mask_cx) * scale
            dy = optic_cy - base.height / 2.0 + (mask.height / 2.0 - mask_cy) * scale
        else:
            dx = 0.0
            dy = 0.0

        self.overlay_scale_var.set(max(0.001, min(20.0, scale)))
        self.overlay_angle_var.set(angle)
        self.overlay_dx_var.set(dx)
        self.overlay_dy_var.set(dy)
        return True


    def overlay_shift_array_zero(self, arr, dx: int, dy: int):
        """Décale un tableau sans retour circulaire."""
        if np is None:
            return arr
        h, w = arr.shape
        out = np.zeros_like(arr)

        if abs(dx) >= w or abs(dy) >= h:
            return out

        src_x0 = max(0, -dx)
        src_x1 = min(w, w - dx) if dx >= 0 else w
        dst_x0 = max(0, dx)
        dst_x1 = dst_x0 + (src_x1 - src_x0)

        src_y0 = max(0, -dy)
        src_y1 = min(h, h - dy) if dy >= 0 else h
        dst_y0 = max(0, dy)
        dst_y1 = dst_y0 + (src_y1 - src_y0)

        if src_x1 > src_x0 and src_y1 > src_y0:
            out[dst_y0:dst_y1, dst_x0:dst_x1] = arr[src_y0:src_y1, src_x0:src_x1]
        return out

    def overlay_mask_map_base_for_params(self, scale_abs: float, angle_deg: float,
                                         dx_full_px: float, dy_full_px: float,
                                         base_to_work_scale: float,
                                         work_size: tuple[int, int]):
        """Carte masque basse résolution pour une échelle/rotation/décalage donnés."""
        if np is None:
            return None

        alpha = self.overlay_get_mask_alpha()
        if alpha is None:
            return None

        work_w, work_h = work_size
        s = max(0.0005, scale_abs * base_to_work_scale)
        new_w = max(1, int(round(alpha.width * s)))
        new_h = max(1, int(round(alpha.height * s)))

        alpha_scaled = alpha.resize(
            (new_w, new_h),
            Image.Resampling.NEAREST if hasattr(Image, "Resampling") else Image.NEAREST
        )

        if abs(angle_deg) > 0.0001:
            alpha_scaled = alpha_scaled.rotate(
                angle_deg,
                expand=True,
                resample=Image.Resampling.NEAREST if hasattr(Image, "Resampling") else Image.NEAREST
            )

        dx = dx_full_px * base_to_work_scale
        dy = dy_full_px * base_to_work_scale
        x = int(round(work_w / 2 - alpha_scaled.width / 2 + dx))
        y = int(round(work_h / 2 - alpha_scaled.height / 2 + dy))

        canvas = Image.new("L", (work_w, work_h), 0)
        canvas.paste(alpha_scaled, (x, y))
        return np.asarray(canvas, dtype=np.float32) / 255.0

    def overlay_score_mask_against_edge(self, edge_map, mask_map) -> float:
        if np is None or mask_map is None:
            return -1e9

        weight = float(mask_map.sum())
        if weight <= 2.0:
            return -1e9

        # Score positif sur les traits. La normalisation limite le biais vers les masques très épais.
        raw = float((edge_map * mask_map).sum() / (weight ** 0.5))

        # Petit malus si trop de masque est hors signal utile.
        coverage = weight / max(float(mask_map.size), 1.0)
        return raw - 0.02 * coverage


    def overlay_find_hexalign_exe(self) -> Path | None:
        """Trouve le moteur C++ hexalign.exe inclus par PyInstaller ou placé à côté de main.py."""
        candidates = []
        try:
            candidates.append(Path(getattr(sys, "_MEIPASS")) / "hexalign.exe")
        except Exception:
            pass
        try:
            candidates.append(Path(__file__).resolve().parent / "hexalign.exe")
        except Exception:
            pass
        try:
            candidates.append(Path.cwd() / "hexalign.exe")
        except Exception:
            pass
        for p in candidates:
            if p.exists():
                return p
        return None

    def overlay_detect_grid_cpp(self, progress=None):
        """Lance le moteur C++ en mode détection de grillage uniquement."""
        if np is None:
            return None

        exe = self.overlay_find_hexalign_exe()
        if exe is None:
            if progress:
                progress(0, "Moteur C++ absent.")
            return None

        built = self.overlay_build_optic_edge_map(max_dim=640)
        if built is None:
            return None

        edge_map, base_to_work_scale, work_size = built
        work_w, work_h = work_size
        edge_u8 = np.clip(edge_map * 255.0, 0, 255).astype(np.uint8)

        with tempfile.TemporaryDirectory(prefix="hexdetect_") as tmp:
            tmpdir = Path(tmp)
            edge_path = tmpdir / "edge.raw"
            edge_u8.tofile(edge_path)

            cmd = [
                str(exe),
                str(edge_path),
                str(work_w),
                str(work_h),
                f"{float(self.overlay_detect_sensitivity_var.get()):.12g}",
                f"{float(self.overlay_detect_percentile_var.get()):.12g}",
                f"{float(self.overlay_pitch_min_var.get()):.12g}",
                f"{float(self.overlay_pitch_max_var.get()):.12g}",
            ]

            if progress:
                progress(1, "Moteur C++ : détection du grillage…")

            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                universal_newlines=True,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
            )

            result = None
            stderr_text = ""

            while True:
                line = proc.stdout.readline() if proc.stdout else ""
                if line:
                    line = line.strip()
                    if line.startswith("PROGRESS "):
                        parts = line.split(" ", 2)
                        try:
                            pct = float(parts[1])
                        except Exception:
                            pct = 0.0
                        msg = parts[2] if len(parts) > 2 else "Détection C++…"
                        if progress:
                            progress(pct, msg)
                    elif line.startswith("RESULT "):
                        parts = line.split()
                        if len(parts) >= 5:
                            result = {
                                "angle": float(parts[1]),
                                "spacing": float(parts[2]),
                                "score": float(parts[3]),
                                "points": float(parts[4]),
                                "base_to_work_scale": float(base_to_work_scale),
                                "work_size": work_size,
                                "engine": "cpp",
                                "edge_map": edge_map,
                            }
                elif proc.poll() is not None:
                    break

            if proc.stderr:
                stderr_text = proc.stderr.read()

            code = proc.wait()
            if code != 0:
                if stderr_text:
                    print("hexalign stderr:", stderr_text[:2000])
                return None

            return result

    def overlay_long_hex_search_cpp(self, progress=None):
        """Prépare les cartes en Python puis délègue la boucle lourde au moteur C++."""
        if np is None:
            return None
        exe = self.overlay_find_hexalign_exe()
        if exe is None:
            if progress:
                progress(0, "Moteur C++ hexalign.exe absent : fallback Python.")
            return None

        self.overlay_auto_fit()
        built = self.overlay_build_optic_edge_map(max_dim=430)
        if built is None:
            return None
        edge_map, base_to_work_scale, work_size = built
        work_w, work_h = work_size

        alpha = self.overlay_get_mask_alpha()
        if alpha is None:
            return None

        optic_features = self.overlay_estimate_hex_features("optic")
        mask_features = self.overlay_estimate_hex_features("mask")
        base_scale = float(self.overlay_scale_var.get())
        base_angle = float(self.overlay_angle_var.get())
        if optic_features is not None and mask_features is not None:
            try:
                base_angle = self.overlay_mod60_delta(float(optic_features["angle_mod"]), float(mask_features["angle_mod"]))
            except Exception:
                pass
        base_dx = float(self.overlay_dx_var.get())
        base_dy = float(self.overlay_dy_var.get())

        if optic_features is not None:
            pitch_work = max(6.0, min(80.0, float(optic_features["pitch"]) * base_to_work_scale))
        else:
            pitch_work = max(10.0, min(work_size) / 20.0)

        edge_u8 = np.clip(edge_map * 255.0, 0, 255).astype(np.uint8)
        mask_u8 = np.asarray(alpha, dtype=np.uint8)

        with tempfile.TemporaryDirectory(prefix="hexalign_") as tmp:
            tmpdir = Path(tmp)
            edge_path = tmpdir / "edge.raw"
            mask_path = tmpdir / "mask.raw"
            edge_u8.tofile(edge_path)
            mask_u8.tofile(mask_path)

            cmd = [
                str(exe), str(edge_path), str(mask_path), str(work_w), str(work_h),
                str(alpha.width), str(alpha.height), f"{base_scale:.12g}", f"{base_angle:.12g}",
                f"{base_dx:.12g}", f"{base_dy:.12g}", f"{base_to_work_scale:.12g}", f"{pitch_work:.12g}",
                f"{float(self.overlay_detect_sensitivity_var.get()):.12g}",
                f"{float(self.overlay_detect_percentile_var.get()):.12g}",
            ]
            if progress:
                progress(1, "Moteur C++ : recherche longue en cours…")

            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                universal_newlines=True,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
            )
            result = None
            while True:
                line = proc.stdout.readline() if proc.stdout else ""
                if line:
                    line = line.strip()
                    if line.startswith("PROGRESS "):
                        parts = line.split(" ", 2)
                        try:
                            pct = float(parts[1])
                        except Exception:
                            pct = 0.0
                        msg = parts[2] if len(parts) > 2 else "Moteur C++…"
                        if progress:
                            progress(pct, msg)
                    elif line.startswith("RESULT "):
                        parts = line.split()
                        if len(parts) >= 6:
                            result = {
                                "scale": float(parts[1]),
                                "angle": float(parts[2]),
                                "dx": float(parts[3]),
                                "dy": float(parts[4]),
                                "score": float(parts[5]),
                                "roi_locked_scale": base_scale,
                                "engine": "cpp",
                            }
                            if len(parts) >= 9:
                                result["optic_spacing"] = float(parts[6])
                                result["mask_spacing"] = float(parts[7])
                                result["hough_angle"] = float(parts[8])
                elif proc.poll() is not None:
                    break

            err = proc.stderr.read() if proc.stderr else ""
            code = proc.wait()
            if code != 0:
                if progress:
                    progress(0, "Moteur C++ en erreur : fallback Python.")
                if err:
                    print("hexalign stderr:", err[:2000])
                return None
            return result

    def overlay_long_hex_search(self, progress=None):
        """Recherche longue verrouillée sur le calage ROI.

        La taille et le centre viennent d'abord du cercle optique ↔ cercle ROI doc.
        Ensuite on cherche surtout la rotation et une petite phase X/Y du motif.
        On évite de changer fortement l'échelle, car le masque complet est naturellement
        plus grand que l'optique.
        """
        cpp_result = self.overlay_long_hex_search_cpp(progress=progress)
        if cpp_result is not None:
            return cpp_result

        if np is None:
            return None

        # 1) Calage géométrique fiable : cercle optique ↔ cercle ROI doc.
        self.overlay_auto_fit()

        built = self.overlay_build_optic_edge_map(max_dim=430)
        if built is None:
            return None
        edge_map, base_to_work_scale, work_size = built

        optic_features = self.overlay_estimate_hex_features("optic")
        mask_features = self.overlay_estimate_hex_features("mask")

        base_scale = float(self.overlay_scale_var.get())
        base_dx = float(self.overlay_dx_var.get())
        base_dy = float(self.overlay_dy_var.get())
        base_angle = float(self.overlay_angle_var.get())

        # L'orientation FFT est seulement une indication de départ.
        if optic_features is not None and mask_features is not None:
            try:
                base_angle = self.overlay_mod60_delta(
                    float(optic_features["angle_mod"]),
                    float(mask_features["angle_mod"]),
                )
            except Exception:
                pass

        # Le pas sert à définir une zone de translation locale, pas à imposer l'échelle.
        if optic_features is not None:
            pitch_work = max(6.0, min(80.0, float(optic_features["pitch"]) * base_to_work_scale))
        else:
            pitch_work = max(10.0, min(work_size) / 20.0)

        # Échelle verrouillée autour du calage ROI : variation très faible seulement.
        scale_factors = [0.94, 0.97, 0.99, 1.0, 1.01, 1.03, 1.06]
        scales = [max(0.001, min(20.0, base_scale * f)) for f in scale_factors]

        # Recherche d'angle large mais périodique sur 60°.
        angles = [base_angle - 30.0 + i * 3.0 for i in range(21)]

        # Phase locale : déplacement de l'ordre d'un pas hexagonal.
        offsets = [-0.75 * pitch_work, -0.38 * pitch_work, 0.0, 0.38 * pitch_work, 0.75 * pitch_work]

        total = len(scales) * len(angles) * len(offsets) * len(offsets)
        count = 0
        best = {
            "score": -1e9,
            "scale": base_scale,
            "angle": base_angle,
            "dx": base_dx,
            "dy": base_dy,
        }

        if progress:
            progress(0, "Recherche longue : taille fixée par cercle ROI, test rotation/phase…")

        for scale_abs in scales:
            for angle in angles:
                base_mask = self.overlay_mask_map_base_for_params(scale_abs, angle, base_dx, base_dy, base_to_work_scale, work_size)
                if base_mask is None:
                    continue
                for ox in offsets:
                    for oy in offsets:
                        shifted = self.overlay_shift_array_zero(base_mask, int(round(ox)), int(round(oy)))
                        score = self.overlay_score_mask_against_edge(edge_map, shifted)
                        if score > best["score"]:
                            best = {
                                "score": score,
                                "scale": scale_abs,
                                "angle": angle,
                                "dx": base_dx + ox / max(base_to_work_scale, 1e-9),
                                "dy": base_dy + oy / max(base_to_work_scale, 1e-9),
                            }

                        count += 1
                        if progress and count % 60 == 0:
                            progress(count / total * 65.0, f"Recherche rotation/phase : {count}/{total}")

        # Affinage très local.
        fine_scales = [max(0.001, min(20.0, best["scale"] * f)) for f in [0.985, 0.995, 1.0, 1.005, 1.015]]
        fine_angles = [best["angle"] - 3.0 + i * 0.25 for i in range(25)]
        fine_offsets = [-0.22 * pitch_work, -0.10 * pitch_work, 0.0, 0.10 * pitch_work, 0.22 * pitch_work]

        total2 = len(fine_scales) * len(fine_angles) * len(fine_offsets) * len(fine_offsets)
        count2 = 0
        base_dx2 = best["dx"]
        base_dy2 = best["dy"]

        if progress:
            progress(65, "Recherche longue : affinage local…")

        for scale_abs in fine_scales:
            for angle in fine_angles:
                base_mask = self.overlay_mask_map_base_for_params(scale_abs, angle, base_dx2, base_dy2, base_to_work_scale, work_size)
                if base_mask is None:
                    continue
                for ox in fine_offsets:
                    for oy in fine_offsets:
                        shifted = self.overlay_shift_array_zero(base_mask, int(round(ox)), int(round(oy)))
                        score = self.overlay_score_mask_against_edge(edge_map, shifted)
                        if score > best["score"]:
                            best = {
                                "score": score,
                                "scale": scale_abs,
                                "angle": angle,
                                "dx": base_dx2 + ox / max(base_to_work_scale, 1e-9),
                                "dy": base_dy2 + oy / max(base_to_work_scale, 1e-9),
                            }

                        count2 += 1
                        if progress and count2 % 75 == 0:
                            progress(65.0 + count2 / total2 * 35.0, f"Affinage local : {count2}/{total2}")

        best["roi_locked_scale"] = base_scale
        if optic_features is not None:
            best["optic_pitch"] = float(optic_features.get("pitch", 0.0))
        if mask_features is not None:
            best["mask_pitch"] = float(mask_features.get("pitch", 0.0))
        return best

    def overlay_detect_grid(self):
        """Détection automatique du grillage, sans calage de masque."""
        if np is None:
            messagebox.showerror(
                "Détection grillage",
                "Cette fonction nécessite numpy. Le fichier requirements.txt doit contenir numpy."
            )
            return

        if getattr(self, "overlay_base_image", None) is None:
            messagebox.showwarning("Détection grillage", "Charge d'abord l'image optique.")
            return

        win = tk.Toplevel(self)
        win.title("Détection grillage automatique")
        win.geometry("520x140")
        win.transient(self)
        win.grab_set()

        label_var = tk.StringVar(value="Préparation de la détection du grillage…")
        pct_var = tk.DoubleVar(value=0.0)

        ttk.Label(win, text="Détection automatique du chicken wire", font=("Arial", 10, "bold")).pack(anchor="w", padx=12, pady=(12, 4))
        ttk.Label(win, textvariable=label_var).pack(anchor="w", padx=12, pady=4)
        bar = ttk.Progressbar(win, maximum=100.0, variable=pct_var, mode="determinate")
        bar.pack(fill="x", padx=12, pady=8)

        def progress(pct, msg):
            pct_var.set(max(0.0, min(100.0, float(pct))))
            label_var.set(str(msg))
            self.overlay_status_var.set(str(msg))
            try:
                win.update_idletasks()
                win.update()
            except Exception:
                pass

        try:
            progress(1, "Construction de la carte de texture chicken-wire…")
            result = self.overlay_detect_grid_cpp(progress=progress)

            if result is None:
                messagebox.showwarning(
                    "Détection grillage",
                    "La détection n'a pas trouvé de grille exploitable. "
                    "Essaie d'augmenter Sensibilité ou de baisser Seuil %, puis relance."
                )
                return

            self.overlay_detected_grid = result
            self.overlay_detected_texture = result.get("edge_map")
            layer = self.overlay_get_selected_layer()
            if layer is not None:
                layer["detected_grid"] = self.overlay_detected_grid
                layer["detected_texture"] = self.overlay_detected_texture
            self.overlay_detection_photo = None
            progress(100, "Détection terminée.")
            self.overlay_status_var.set(
                f"Texture grillage détectée : angle≈{float(result['angle']):.2f}° | "
                f"pas≈{float(result['spacing']) / max(float(result['base_to_work_scale']), 1e-9):.1f}px image | "
                f"score={float(result['score']):.4g} | points={int(result.get('points', 0))}"
            )
            self.overlay_redraw()

        except Exception as exc:
            messagebox.showerror("Détection grillage", f"Erreur pendant la détection :\n{exc}")
        finally:
            try:
                win.grab_release()
                win.destroy()
            except Exception:
                pass

    def overlay_export(self):
        composite = self.overlay_composite_fullres()
        if composite is None:
            messagebox.showwarning("Export", "Charge au moins une image optique avant d'exporter.")
            return
        path = filedialog.asksaveasfilename(
            title="Exporter la composition",
            defaultextension=".png",
            filetypes=[("PNG", "*.png"), ("JPEG", "*.jpg"), ("Tous les fichiers", "*.*")]
        )
        if not path:
            return
        try:
            composite.save(path)
            self.overlay_status_var.set(f"Composition exportée : {path}")
        except Exception as exc:
            messagebox.showerror("Export", f"Impossible d'exporter :\n{exc}")

    def overlay_export_light_mask(self):
        mask_rgba = self.overlay_mask_to_rgba(opacity=1.0)
        if mask_rgba is None:
            messagebox.showwarning("Masque léger", "Charge un masque de rangement avant d'exporter.")
            return
        path = filedialog.asksaveasfilename(
            title="Sauver le masque léger transparent",
            defaultextension=".png",
            filetypes=[("PNG transparent", "*.png"), ("Tous les fichiers", "*.*")]
        )
        if not path:
            return
        try:
            mask_rgba.save(path)
            self.overlay_status_var.set(f"Masque léger exporté : {path}")
        except Exception as exc:
            messagebox.showerror("Masque léger", f"Impossible d'exporter :\n{exc}")

    def overlay_copy_params(self):
        layers = getattr(self, "overlay_layers", [])
        lines = [
            f"Nb_images={len(layers)}",
            f"Image_selectionnee={getattr(self, 'overlay_selected_index', 0) + 1}",
            f"View_zoom={float(self.overlay_view_zoom_var.get()):.6g}",
            f"Detection_sensitivity={float(self.overlay_detect_sensitivity_var.get()):.6g}",
            f"Detection_percentile={float(self.overlay_detect_percentile_var.get()):.6g}",
            f"Detection_blur={float(self.overlay_detect_blur_var.get()):.6g}",
            f"Pitch_min_px={float(self.overlay_pitch_min_var.get()):.6g}",
            f"Pitch_max_px={float(self.overlay_pitch_max_var.get()):.6g}",
        ]
        for i, layer in enumerate(layers, start=1):
            lines += [
                f"[Image {i}]",
                f"Path={layer.get('path', '')}",
                f"Rotation_deg={float(layer.get('angle', 0.0)):.6g}",
                f"Scale={float(layer.get('scale', 1.0)):.6g}",
                f"Offset_X_px={float(layer.get('dx', 0.0)):.6g}",
                f"Offset_Y_px={float(layer.get('dy', 0.0)):.6g}",
                f"Opacity={float(layer.get('opacity', 1.0)):.6g}",
                f"Flip_H={bool(layer.get('flip_h', False))}",
                f"Flip_V={bool(layer.get('flip_v', False))}",
                f"Detected_grid={layer.get('detected_grid')}",
            ]
        text = "\n".join(lines)
        try:
            self.clipboard_clear()
            self.clipboard_append(text)
            self.overlay_status_var.set("Paramètres copiés dans le presse-papier.")
        except Exception:
            messagebox.showinfo("Paramètres analyse image", text)

    def get_follow_targets(self) -> list[str]:
        if self.follow_targets_text is not None:
            raw = self.follow_targets_text.get("1.0", "end")
            targets = [line.strip() for line in raw.splitlines() if line.strip()]
            return list(dict.fromkeys(targets))
        return load_follow_targets()

    def save_follow_targets_from_ui(self):
        targets = self.get_follow_targets()
        save_follow_targets(targets)
        self.status_var.set(f"Suivi enregistré : {len(targets)} module(s).")
        messagebox.showinfo("Suivi", f"{len(targets)} numéro(s) module enregistré(s).")

    def render_follow_tab(self):
        self.clear_frame(self.follow_tab)
        frame = ttk.Frame(self.follow_tab, padding=10)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Suivi de modules cibles", font=("Arial", 12, "bold")).pack(anchor="w")
        ttk.Label(
            frame,
            text=(
                "Ajoute ici les numéros de module à surveiller, un par ligne. "
                "Quand le mode suivi est coché à côté de Scanner, le scan cherchera spécifiquement ces modules."
            ),
            wraplength=900,
        ).pack(anchor="w", pady=(4, 8))

        self.follow_targets_text = ScrolledText(frame, wrap="none", font=("Consolas", 11), height=18)
        self.follow_targets_text.pack(fill="both", expand=True, pady=4)
        self.follow_targets_text.delete("1.0", "end")
        self.follow_targets_text.insert("1.0", "\n".join(load_follow_targets()))

        btns = ttk.Frame(frame)
        btns.pack(fill="x", pady=8)
        ttk.Button(btns, text="Enregistrer la liste", command=self.save_follow_targets_from_ui).pack(side="left", padx=4)
        ttk.Button(btns, text="Scanner en mode suivi maintenant", command=self.scan).pack(side="left", padx=4)
        ttk.Checkbutton(btns, text="Mode suivi activé", variable=self.follow_mode_var).pack(side="left", padx=16)
        ttk.Checkbutton(btns, text="TXT uniquement", variable=self.txt_only_var).pack(side="left", padx=8)

    def render_settings(self):
        self.clear_frame(self.settings_tab)

        container = ttk.Frame(self.settings_tab, padding=10)
        container.pack(fill="both", expand=True)

        ttk.Label(container, text="Paramètres du rapport et du scan", font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 8))

        ttk.Label(container, text="Max fichiers à traiter par type").grid(row=1, column=0, sticky="w", pady=4)
        max_files_var = tk.StringVar(value=str(get_max_files_per_type()))
        ttk.Entry(container, textvariable=max_files_var, width=12).grid(row=1, column=1, sticky="w", pady=4)

        ttk.Label(container, text="Résolution caméra (µm/pixel)").grid(row=2, column=0, sticky="w", pady=4)
        res_var = tk.StringVar(value=str(APP_SETTINGS.get("resolution_um_per_pixel", 5.4)))
        ttk.Entry(container, textvariable=res_var, width=12).grid(row=2, column=1, sticky="w", pady=4)

        ttk.Label(container, text="Taille max prévisualisation images").grid(row=3, column=0, sticky="w", pady=4)
        preview_var = tk.StringVar(value=str(get_preview_max_size()))
        ttk.Entry(container, textvariable=preview_var, width=12).grid(row=3, column=1, sticky="w", pady=4)

        ttk.Label(container, text="Zones : Nom;Rayon max en µm. Laisse vide pour Hors zone.", font=("Arial", 10, "bold")).grid(row=4, column=0, columnspan=2, sticky="w", pady=(14, 2))
        zones_text = ScrolledText(container, width=42, height=7, font=("Consolas", 10))
        zones_text.grid(row=5, column=0, columnspan=2, sticky="nsew", padx=(0, 12))
        zones_text.insert("1.0", zones_to_text())

        ttk.Label(container, text="Tailles : Nom;Min µm;Max µm", font=("Arial", 10, "bold")).grid(row=4, column=2, columnspan=2, sticky="w", pady=(14, 2))
        sizes_text = ScrolledText(container, width=42, height=7, font=("Consolas", 10))
        sizes_text.grid(row=5, column=2, columnspan=2, sticky="nsew")
        sizes_text.insert("1.0", size_buckets_to_text())

        ttk.Label(container, text="Colonnes grade/statut : Nom colonne;clé dans [GRADE]", font=("Arial", 10, "bold")).grid(row=6, column=0, columnspan=2, sticky="w", pady=(14, 2))
        grades_text = ScrolledText(container, width=42, height=5, font=("Consolas", 10))
        grades_text.grid(row=7, column=0, columnspan=2, sticky="nsew", padx=(0, 12))
        grades_text.insert("1.0", grade_columns_to_text())

        help_text = (
            "Exemples :\n"
            "Zones :\n"
            "Z1;5800\nZ2;14700\nZ3;17700\nHors zone;\n\n"
            "Colonnes grade :\n"
            "GradeChickenWire;CW\nGradeInho;INHO\n"
            "Tu peux ajouter une ligne comme : GradeGrowing;GROWING SPOT"
        )
        ttk.Label(container, text=help_text, justify="left").grid(row=7, column=2, columnspan=2, sticky="nw")

        def apply_settings(reset: bool = False):
            try:
                if reset:
                    new_settings = normalize_settings(DEFAULT_SETTINGS)
                else:
                    new_settings = {
                        "resolution_um_per_pixel": float(res_var.get().replace(",", ".")),
                        "max_files_per_type": int(max_files_var.get()),
                        "preview_max_size": int(preview_var.get()),
                        "zones": parse_zones_text(zones_text.get("1.0", "end")),
                        "size_buckets": parse_size_buckets_text(sizes_text.get("1.0", "end")),
                        "grade_columns": parse_grade_columns_text(grades_text.get("1.0", "end")),
                    }
                    new_settings = normalize_settings(new_settings)

                apply_global_settings(new_settings)
                save_app_settings(new_settings)
                self.mark_views_dirty()
                self.status_var.set("Paramètres enregistrés. Recharge du rapport au prochain affichage.")
                messagebox.showinfo("Paramètres", "Paramètres enregistrés.")
                self.render_settings()
            except Exception as exc:
                messagebox.showerror("Paramètres invalides", str(exc))

        buttons = ttk.Frame(container)
        buttons.grid(row=8, column=0, columnspan=4, sticky="ew", pady=12)
        ttk.Button(buttons, text="Enregistrer / appliquer", command=lambda: apply_settings(False)).pack(side="left", padx=4)
        ttk.Button(buttons, text="Réinitialiser par défaut", command=lambda: apply_settings(True)).pack(side="left", padx=4)

        container.columnconfigure(0, weight=1)
        container.columnconfigure(2, weight=1)
        container.rowconfigure(5, weight=1)
        container.rowconfigure(7, weight=1)

    def render_text_compare(self, product: str):
        self.clear_frame(self.text_tab)
        runs = self.product_text_runs(product)

        if len(runs) < 2:
            total_unfiltered = self.unfiltered_product_text_count(product)
            ttk.Label(
                self.text_tab,
                text=f"Module {product} — {len(runs)} fichier TXT disponible. Affichage du fichier seul.",
                font=("Arial", 10, "bold"),
            ).pack(anchor="w", padx=8, pady=8)
            for r in runs:
                self.make_single_text_view(self.text_tab, r)
            return

        toolbar = ttk.Frame(self.text_tab, padding=6)
        toolbar.pack(fill="x")

        ttk.Label(toolbar, text=f"Module {product} — {len(runs)} fichiers TXT", font=("Arial", 10, "bold")).pack(side="left", padx=(0, 16))
        ttk.Label(toolbar, text="Fichier référence :").pack(side="left")

        self.ref_label_to_path = {run_label(r): str(r.path) for r in runs}
        labels = list(self.ref_label_to_path.keys())

        previous_ref_path = self.reference_by_product.get(product)
        default_label = labels[0]
        if previous_ref_path:
            for label, path in self.ref_label_to_path.items():
                if path == previous_ref_path:
                    default_label = label
                    break

        self.reference_var.set(default_label)
        ref_combo = ttk.Combobox(toolbar, textvariable=self.reference_var, values=labels, state="readonly", width=55)
        ref_combo.pack(side="left", padx=6)
        ref_combo.bind("<<ComboboxSelected>>", lambda _e, p=product: self.on_reference_changed(p))

        ttk.Label(toolbar, text="Les différences de la liste 2 sont en rouge. PARAMETERS est masqué.").pack(side="left", padx=12)

        ref = next((r for r in runs if str(r.path) == self.ref_label_to_path[self.reference_var.get()]), runs[0])
        self.reference_by_product[product] = str(ref.path)

        compare_tabs = ttk.Notebook(self.text_tab)
        compare_tabs.pack(fill="both", expand=True, padx=4, pady=4)

        for comp in runs:
            if comp.path == ref.path:
                continue
            tab = ttk.Frame(compare_tabs)
            compare_tabs.add(tab, text=date_label(comp.timestamp))
            self.make_text_diff_tab(tab, ref, comp)


    def make_single_text_view(self, parent, run: TextRun):
        top = ttk.Frame(parent, padding=4)
        top.pack(fill="x")
        ttk.Label(top, text=f"{date_label(run.timestamp)} — {run.path.name}").pack(side="left", anchor="w")
        ttk.Button(top, text="Visualisation", command=lambda rr=run: self.open_single_visual_report(rr)).pack(side="right", padx=4)

        txt = ScrolledText(parent, wrap="none", font=("Consolas", 10))
        txt.pack(fill="both", expand=True, padx=6, pady=6)
        txt.tag_configure("section", font=("Consolas", 10, "bold"))
        for section, lines in run.sections.items():
            if section in SKIP_SECTIONS_IN_COMPARE:
                continue
            txt.insert("end", f"[{section}]\n", "section")
            for line in lines:
                txt.insert("end", line + "\n")
            txt.insert("end", "\n")
        txt.configure(state="disabled")

    def on_reference_changed(self, product: str):
        selected_label = self.reference_var.get()
        ref_path = self.ref_label_to_path.get(selected_label)
        if ref_path:
            self.reference_by_product[product] = ref_path
        self.render_text_compare(product)

    def open_single_visual_report(self, run: TextRun):
        DefectReportWindow(self, run, run)

    def make_text_diff_tab(self, parent, ref: TextRun, comp: TextRun):
        top = ttk.Frame(parent, padding=4)
        top.pack(fill="x")
        info = ttk.Label(
            top,
            text=f"Liste 1 / référence : {ref.path.name} ({date_label(ref.timestamp)})    |    Liste 2 : {comp.path.name} ({date_label(comp.timestamp)})",
        )
        info.pack(side="left", anchor="w")
        ttk.Button(top, text="Visualisation", command=lambda: DefectReportWindow(self, ref, comp)).pack(side="right", padx=4)

        panes = ttk.PanedWindow(parent, orient="horizontal")
        panes.pack(fill="both", expand=True)

        left = ScrolledText(panes, wrap="none", font=("Consolas", 10))
        right = ScrolledText(panes, wrap="none", font=("Consolas", 10))
        panes.add(left, weight=1)
        panes.add(right, weight=1)

        left.tag_configure("section", font=("Consolas", 10, "bold"))
        right.tag_configure("section", font=("Consolas", 10, "bold"))
        left.tag_configure("missing", foreground="red")
        right.tag_configure("diff", foreground="red")
        right.tag_configure("missing", foreground="red")

        for section in section_order(ref, comp):
            left.insert("end", f"[{section}]\n", "section")
            right.insert("end", f"[{section}]\n", "section")

            ref_lines = ref.sections.get(section, [])
            comp_lines = comp.sections.get(section, [])
            pairs = align_lines(section, ref_lines, comp_lines)

            if not pairs:
                left.insert("end", "\n")
                right.insert("end", "\n")
                continue

            for ref_line, comp_line in pairs:
                if ref_line is None:
                    left.insert("end", "<nouveau dans liste 2>\n", "missing")
                else:
                    left.insert("end", ref_line + "\n")
                self.insert_compared_line(right, comp_line, ref_line)

            left.insert("end", "\n")
            right.insert("end", "\n")

        left.configure(state="disabled")
        right.configure(state="disabled")

    def insert_compared_line(self, widget: ScrolledText, comp_line: str | None, ref_line: str | None):
        if comp_line is None:
            widget.insert("end", "<absent dans liste 2>\n", "missing")
            return
        if ref_line is None:
            widget.insert("end", comp_line + "\n", "diff")
            return

        comp_fields, sep = split_fields(comp_line)
        ref_fields, _ = split_fields(ref_line)
        for i, field in enumerate(comp_fields):
            is_diff = i >= len(ref_fields) or field != ref_fields[i]
            widget.insert("end", field, "diff" if is_diff else None)
            if sep and i < len(comp_fields) - 1:
                widget.insert("end", sep)
        widget.insert("end", "\n")

    def render_images(self, product: str):
        self.clear_frame(self.image_tab)

        if Image is None:
            ttk.Label(self.image_tab, text="Installe Pillow : pip install pillow").pack(anchor="w", padx=8, pady=8)
            return

        images = self.product_image_runs(product)
        if not images:
            total_unfiltered = self.unfiltered_product_image_count(product)
            if total_unfiltered > 0 and self.date_filter_is_active():
                ttk.Label(
                    self.image_tab,
                    text=(
                        f"Aucune image dans l'intervalle choisi, mais {total_unfiltered} image(s) existent hors filtre date. "
                        "Clique sur 'Tout' dans l'intervalle de dates."
                    ),
                    foreground="red",
                    font=("Arial", 10, "bold"),
                ).pack(anchor="w", padx=8, pady=8)
            else:
                ttk.Label(
                    self.image_tab,
                    text="Aucune image trouvée pour ce module. Le numéro module doit être présent dans un dossier parent ou un nom de fichier.",
                ).pack(anchor="w", padx=8, pady=8)
            return

        groups: dict[str, list[ImageRun]] = {}
        for img in images:
            groups.setdefault(img.picture_name, []).append(img)

        ttk.Label(
            self.image_tab,
            text=(
                f"Module {product} — {len(images)} images indexées. "
                "Les images ne sont ouvertes que lorsque leur onglet picture_x est sélectionné. "
                "Affichage initial en prévisualisation rapide, bouton disponible pour pleine résolution."
            ),
        ).pack(anchor="w", padx=8, pady=4)

        pic_tabs = ttk.Notebook(self.image_tab)
        pic_tabs.pack(fill="both", expand=True, padx=4, pady=4)

        data = {"loaded": set(), "records": {}, "frames": {}}
        self._picture_notebooks[pic_tabs] = data

        for picture_name, records in sorted(groups.items()):
            tab = ttk.Frame(pic_tabs)
            pic_tabs.add(tab, text=f"{picture_name} ({len(records)})")
            data["records"][tab] = records
            data["frames"][tab] = tab
            ttk.Label(tab, text=f"{picture_name} — {len(records)} image(s). Sélectionne cet onglet pour charger les images.").pack(anchor="w", padx=8, pady=8)

        pic_tabs.bind("<<NotebookTabChanged>>", lambda _e, nb=pic_tabs: self.on_picture_tab_changed(nb))
        self.after(150, lambda nb=pic_tabs: self.load_selected_picture_tab(nb))

    def on_picture_tab_changed(self, notebook: ttk.Notebook):
        self.load_selected_picture_tab(notebook)

    def load_selected_picture_tab(self, notebook: ttk.Notebook):
        data = self._picture_notebooks.get(notebook)
        if not data:
            return
        try:
            tab_id = notebook.select()
            if not tab_id:
                return
            tab_widget = notebook.nametowidget(tab_id)
        except Exception:
            return

        if tab_widget in data["loaded"]:
            return

        records = data["records"].get(tab_widget, [])
        self.clear_frame(tab_widget)
        self.make_image_group_tab(tab_widget, records)
        data["loaded"].add(tab_widget)

    def make_image_group_tab(self, parent, records: list[ImageRun]):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)

        group = SyncedImageGroup()

        toolbar = ttk.Frame(parent, padding=4)
        toolbar.grid(row=0, column=0, sticky="ew")
        ttk.Label(toolbar, text=f"{len(records)} image(s) — zoom/pan synchronisés").pack(side="left")
        ttk.Button(toolbar, text="Ajuster", command=group.reset_fit).pack(side="right", padx=4)
        ttk.Button(toolbar, text="Zoom +", command=lambda: group.zoom_center(1.15)).pack(side="right", padx=4)
        ttk.Button(toolbar, text="Zoom -", command=lambda: group.zoom_center(1 / 1.15)).pack(side="right", padx=4)
        ttk.Button(toolbar, text="Pleine résolution", command=lambda: group.set_full_resolution(True)).pack(side="right", padx=4)
        ttk.Button(toolbar, text="Prévisualisation rapide", command=lambda: group.set_full_resolution(False)).pack(side="right", padx=4)

        body = ttk.Frame(parent)
        body.grid(row=1, column=0, sticky="nsew")

        n = len(records)
        if n <= 2:
            cols = max(n, 1)
        elif n <= 4:
            cols = 2
        else:
            cols = 3
        rows = math.ceil(n / cols) if n else 1

        for c in range(cols):
            body.columnconfigure(c, weight=1, uniform="img")
        for r in range(rows):
            body.rowconfigure(r, weight=1, uniform="img")

        for idx, rec in enumerate(records):
            viewer = SyncedImageCanvas(body, group, rec, self.image_rotations)
            viewer.grid(row=idx // cols, column=idx % cols, sticky="nsew", padx=3, pady=3)

        parent.after(250, group.fit_all)



if __name__ == "__main__":
    app = DefectCompareApp()
    app.mainloop()
